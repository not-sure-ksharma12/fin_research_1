# One-stop Python script: Fetch Bloomberg option data and calculate Black-Scholes + Heston prices and Greeks
import blpapi  # type: ignore
import pandas as pd
import QuantLib as ql
from datetime import datetime
import numpy as np
import time
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

# Bloomberg session setup
def connect_to_bloomberg():
    opts = blpapi.SessionOptions()
    opts.setServerHost("localhost")
    opts.setServerPort(8194)
    session = blpapi.Session(opts)
    if not session.start(): raise RuntimeError("Failed to start Bloomberg session")
    if not session.openService("//blp/refdata"): raise RuntimeError("Failed to open //blp/refdata")
    return session

def get_current_prices(session, tickers):
    ref = session.getService("//blp/refdata")
    request = ref.createRequest("ReferenceDataRequest")
    for t in tickers:
        request.append("securities", f"{t} US Equity")
    request.append("fields", "PX_LAST")
    session.sendRequest(request)
    prices = {}
    while True:
        ev = session.nextEvent(500)
        for msg in ev:
            if msg.messageType() == "ReferenceDataResponse":
                for s in msg.getElement("securityData").values():
                    prices[s.getElementAsString("security").split()[0]] = s.getElement("fieldData").getElementAsFloat("PX_LAST")
        if ev.eventType() == blpapi.Event.RESPONSE:
            break
    return prices

def get_risk_free_rate(session):
    ref = session.getService("//blp/refdata")
    request = ref.createRequest("ReferenceDataRequest")
    request.append("securities", "USGG3M Index")
    request.append("fields", "PX_LAST")
    session.sendRequest(request)
    while True:
        ev = session.nextEvent(500)
        for msg in ev:
            if msg.messageType() == "ReferenceDataResponse":
                field = msg.getElement("securityData").getValue(0).getElement("fieldData")
                return field.getElementAsFloat("PX_LAST") / 100
        if ev.eventType() == blpapi.Event.RESPONSE:
            break
    return 0.05

def generate_strikes(current_price):
    # Generate strikes around the current price (±30 strikes, 10 apart)
    base = int(current_price // 10) * 10  # Round to nearest 10
    strikes = []
    # Start from a reasonable minimum strike (at least 50% of current price)
    min_strike = max(1, int(current_price * 0.5))
    for i in range(-30, 31):
        strike = base + i * 10
        if strike >= min_strike:
            strikes.append(strike)
    return sorted(strikes)

def build_option_tickers(ticker, strikes, expiry="20251219"):
    # Add 'US' to ticker root for Bloomberg
    ticker_root = f"{ticker} US"
    puts = [(f"{ticker_root} {expiry}P{int(strike*100):08d} Equity", strike, "Put") for strike in strikes[:25]]
    calls = [(f"{ticker_root} {expiry}C{int(strike*100):08d} Equity", strike, "Call") for strike in strikes[-25:]]
    return puts + calls

def fetch_bloomberg_option_data(session, option_tickers):
    import re
    ref = session.getService("//blp/refdata")
    results = []
    for i in range(0, len(option_tickers), 50):
        batch = option_tickers[i:i+50]
        request = ref.createRequest("ReferenceDataRequest")
        for o in batch:
            request.append("securities", o[0])
        fields = ["PX_LAST", "BID", "ASK", "MID", "DELTA", "GAMMA", "VEGA", "THETA", "RHO", "IMPVOL_MID", "OPT_EXPIRE_DT", "OPT_CONTRACT_SIZE"]
        for f in fields:
            request.append("fields", f)
        session.sendRequest(request)
        while True:
            ev = session.nextEvent(500)
            for msg in ev:
                if msg.messageType() == "ReferenceDataResponse":
                    for idx, s in enumerate(msg.getElement("securityData").values()):
                        fd = s.getElement("fieldData")
                        opt_ticker = s.getElementAsString("security")
                        match = next((x for x in batch if x[0] == opt_ticker), None)
                        # --- Expiry fallback logic ---
                        if fd.hasElement("OPT_EXPIRE_DT"):
                            expiration = fd.getElementAsDatetime("OPT_EXPIRE_DT").date().isoformat()
                        else:
                            m = re.search(r" (\d{8})[CP]", opt_ticker)
                            if m:
                                expiry_str = m.group(1)
                                expiration = f"{expiry_str[:4]}-{expiry_str[4:6]}-{expiry_str[6:]}"
                            else:
                                expiration = None
                        if idx < 3:
                            print(f"[DEBUG] Fields for {opt_ticker}: {dict(fd)}")
                        if match:
                            results.append({
                                "Option Ticker": opt_ticker,
                                "Underlying": opt_ticker.split()[0],
                                "Strike": match[1],
                                "Option Type": match[2],
                                "PX_LAST": fd.getElementAsFloat("PX_LAST") if fd.hasElement("PX_LAST") else None,
                                "BID": fd.getElementAsFloat("BID") if fd.hasElement("BID") else None,
                                "ASK": fd.getElementAsFloat("ASK") if fd.hasElement("ASK") else None,
                                "MID": fd.getElementAsFloat("MID") if fd.hasElement("MID") else None,
                                "DELTA": fd.getElementAsFloat("DELTA") if fd.hasElement("DELTA") else None,
                                "GAMMA": fd.getElementAsFloat("GAMMA") if fd.hasElement("GAMMA") else None,
                                "VEGA": fd.getElementAsFloat("VEGA") if fd.hasElement("VEGA") else None,
                                "THETA": fd.getElementAsFloat("THETA") if fd.hasElement("THETA") else None,
                                "RHO": fd.getElementAsFloat("RHO") if fd.hasElement("RHO") else None,
                                "IMPVOL_MID": fd.getElementAsFloat("IMPVOL_MID") if fd.hasElement("IMPVOL_MID") else None,
                                "Expiration": expiration,
                                "Contract Size": fd.getElementAsFloat("OPT_CONTRACT_SIZE") if fd.hasElement("OPT_CONTRACT_SIZE") else None
                            })
            if ev.eventType() == blpapi.Event.RESPONSE:
                break
        time.sleep(0.1)
    return results

def test_with_known_tickers(session, underlying, expiry, current_price):
    """Test with known valid option ticker format"""
    # Generate some test tickers in Bloomberg format
    test_tickers = []
    
    # Generate strikes around the current price
    base_strike = int(current_price // 10) * 10  # Round to nearest 10
    strikes = [base_strike - 20, base_strike - 10, base_strike, base_strike + 10, base_strike + 20]
    
    # Test a few strikes around the current price
    for strike in strikes:
        if strike > 0:  # Only positive strikes
            # Bloomberg format: TICKER YYYYMMDDC/PSTRIKE Equity
            call_ticker = f"{underlying} US {expiry}C{strike:08d} Equity"
            put_ticker = f"{underlying} US {expiry}P{strike:08d} Equity"
            test_tickers.append((call_ticker, strike, "Call"))
            test_tickers.append((put_ticker, strike, "Put"))
    
    print(f"[DEBUG] Testing {len(test_tickers)} known tickers for {underlying} at price {current_price}")
    print(f"[DEBUG] Sample tickers: {test_tickers[:3]}")
    
    return test_tickers

def fetch_option_chain(session, underlying, expiry, current_price):
    # For now, use the test function instead of OPT_CHAIN
    return test_with_known_tickers(session, underlying, expiry, current_price)

def calc_T(expiration):
    d = datetime.strptime(expiration, "%Y-%m-%d")
    return max((d - datetime.today()).days / 365.25, 0.0001)

def enrich_with_models(data, prices, rfr):
    calendar = ql.UnitedStates(ql.UnitedStates.NYSE)
    day_count = ql.Actual365Fixed()
    today = ql.Date.todaysDate()
    ql.Settings.instance().evaluationDate = today
    out = []
    print(f"[DEBUG] Prices keys: {list(prices.keys())}")
    print(f"[DEBUG] Processing {len(data)} rows...")
    
    for i, row in enumerate(data):
        try:
            if i % 10 == 0:
                print(f"[DEBUG] Processing row {i}/{len(data)}")
            print(f"[DEBUG] Row {i} Underlying: {row.get('Underlying')}")

            # Defensive checks for required fields
            if row.get("PX_LAST") is None:
                print(f"[DEBUG] Skipping row {i}: PX_LAST is None")
                continue
            if row.get("IMPVOL_MID") is None or pd.isna(row.get("IMPVOL_MID")):
                print(f"[DEBUG] Skipping row {i}: IMPVOL_MID is None or NaN")
                continue
            if row.get("Expiration") is None:
                print(f"[DEBUG] Skipping row {i}: Expiration is None")
                continue

            S = prices.get(row["Underlying"], None)
            if S is None:
                print(f"[DEBUG] No price found for {row['Underlying']}")
                continue
            K = row["Strike"]
            T = calc_T(row["Expiration"])
            sigma = row["IMPVOL_MID"]
            q = 0.0

            spot = ql.SimpleQuote(S)
            spot_handle = ql.QuoteHandle(spot)
            rate_handle = ql.YieldTermStructureHandle(ql.FlatForward(today, rfr, day_count))
            div_handle = ql.YieldTermStructureHandle(ql.FlatForward(today, q, day_count))
            vol_handle = ql.BlackVolTermStructureHandle(ql.BlackConstantVol(today, calendar, sigma, day_count))

            payoff_type = ql.Option.Call if row["Option Type"] == "Call" else ql.Option.Put
            payoff = ql.PlainVanillaPayoff(payoff_type, K)
            expiry = today + int(T * 365)
            exercise = ql.EuropeanExercise(expiry)

            bs_process = ql.BlackScholesMertonProcess(spot_handle, div_handle, rate_handle, vol_handle)
            option = ql.VanillaOption(payoff, exercise)
            option.setPricingEngine(ql.AnalyticEuropeanEngine(bs_process))
            print(f"[DEBUG] Before BS NPV for row {i}")
            try:
                bs_price = option.NPV()
                bs_delta = option.delta()
                bs_gamma = option.gamma()
                bs_theta = option.theta() / 365
                bs_vega = option.vega() / 100
                bs_rho = option.rho() / 100
            except Exception as e:
                print(f"[DEBUG] Exception in BS pricing for row {i}: {e}")
                continue
            print(f"[DEBUG] After BS NPV for row {i}")
            try:
                if row["PX_LAST"] is not None:
                    print(f"[DEBUG] Before impliedVolatility for row {i}")
                    iv = option.impliedVolatility(row["PX_LAST"], bs_process, 1e-4, 100, 1e-7, 4.0)
                    print(f"[DEBUG] After impliedVolatility for row {i}")
                else:
                    iv = np.nan
            except Exception as e:
                print(f"[DEBUG] Implied volatility error for row {i}: {e}")
                iv = np.nan

            v0 = sigma**2
            kappa = 2.0
            theta = v0
            sigma_v = 0.3
            rho = -0.5
            heston_process = ql.HestonProcess(rate_handle, div_handle, spot_handle, v0, kappa, theta, sigma_v, rho)
            heston_model = ql.HestonModel(heston_process)
            heston_fd_engine = ql.FdHestonVanillaEngine(heston_model, 100, 400, 0, 0)
            option.setPricingEngine(heston_fd_engine)
            print(f"[DEBUG] Before Heston NPV for row {i}")
            try:
                heston_price = option.NPV()
                h_delta = option.delta()
                h_gamma = option.gamma()
                h_theta = option.theta() / 365
                h_vega = option.vega() / 100
                h_rho = option.rho() / 100
            except Exception as e:
                print(f"[DEBUG] Heston FD Greeks error for row {i}: {e}")
                heston_price = h_delta = h_gamma = h_theta = h_vega = h_rho = np.nan
            print(f"[DEBUG] After Heston NPV for row {i}")

            px_last = row["PX_LAST"]
            if px_last is None:
                bs_vs_market = heston_vs_market = np.nan
            else:
                bs_vs_market = px_last - bs_price
                heston_vs_market = px_last - heston_price

            out.append({
                **row,
                "Current Price": S,
                "Risk-Free Rate": rfr,
                "Time to Exp": T,
                "BS Price": bs_price,
                "BS Delta": bs_delta,
                "BS Gamma": bs_gamma,
                "BS Theta": bs_theta,
                "BS Vega": bs_vega,
                "BS Rho": bs_rho,
                "BS IV (from Market)": iv,
                "Heston Price": heston_price,
                "Heston Delta": h_delta,
                "Heston Gamma": h_gamma,
                "Heston Theta": h_theta,
                "Heston Vega": h_vega,
                "Heston Rho": h_rho,
                "BS vs Market": bs_vs_market,
                "Heston vs Market": heston_vs_market,
                "BS vs Heston": bs_price - heston_price
            })
        except Exception as e:
            print(f"[DEBUG] Exception in enrich_with_models row {i}: {e}")
            print(f"[DEBUG] Row data: {row}")
            continue
    print(f"[DEBUG] Finished enrich_with_models, output count: {len(out)}")
    return out

if __name__ == "__main__":
    print("🔗 Connecting to Bloomberg...")
    session = connect_to_bloomberg()
    print("✅ Bloomberg connection successful")
    
    tickers = ['NVDA', 'AAPL']
    expiry = "20240719"  # Use a near-term expiry
    
    print(f"📊 Fetching current prices for {tickers}...")
    prices = get_current_prices(session, tickers)
    print(f"💰 Current prices: {prices}")
    
    print("📈 Fetching risk-free rate...")
    rfr = get_risk_free_rate(session)
    print(f"💵 Risk-free rate: {rfr}")

    all_opts = []
    for t in tickers:
        spot = prices.get(t)
        if spot:
            print(f"🔍 Fetching option chain for {t}...")
            valid_opts = fetch_option_chain(session, t, expiry, spot)
            all_opts.extend(valid_opts)
            print(f"📋 Fetched {len(valid_opts)} valid options for {t}")
        else:
            print(f"⚠️ No price found for {t}")

    print(f"📝 Total valid options to fetch: {len(all_opts)}")
    print(f"📝 Sample of all valid tickers: {all_opts[:5]}")
    print(f"[DEBUG] All option tickers to request:")
    for opt in all_opts:
        print(opt[0])

    print("🔍 Fetching Bloomberg option data...")
    raw_data = fetch_bloomberg_option_data(session, all_opts)
    print(f"📊 Raw data count: {len(raw_data)}")
    
    if raw_data:
        print("🧮 Enriching with model calculations...")
        enriched = enrich_with_models(raw_data, prices, rfr)
        print(f"📈 Enriched data count: {len(enriched)}")
    else:
        print("⚠️ No raw data to enrich")
        enriched = []

    df = pd.DataFrame(enriched)
    print(f"📋 DataFrame shape: {df.shape}")
    
    output_file = "options_bs_heston_analysis.xlsx"
    df.to_excel(output_file, index=False)
    print(f"💾 Saved DataFrame to {output_file}")

    # Formatting
    wb = load_workbook(output_file)
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Format headers
    for col in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col)  # type: ignore
        cell.font = header_font
        cell.fill = header_fill

    # Add conditional formatting for comparison columns
    for col_name in ["BS vs Market", "Heston vs Market", "BS vs Heston"]:
        if col_name in df.columns:
            col_idx = df.columns.get_loc(col_name) + 1  # type: ignore
            col_letter = get_column_letter(col_idx)  # type: ignore
            color_rule = ColorScaleRule(start_type='min', start_color='F8696B',
                                        mid_type='percentile', mid_value=50, mid_color='FFEB84',
                                        end_type='max', end_color='63BE7B')
            ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{len(df) + 1}", color_rule)  # type: ignore

    wb.save(output_file)
    print(f"✅ Saved to {output_file} with comparisons and formatting")
