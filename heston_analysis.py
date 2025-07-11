import pandas as pd
import QuantLib as ql
from datetime import datetime
import numpy as np

# Load original Bloomberg market data
df = pd.read_excel("bloomberg_options_top50.xlsx")

# Setup
calendar = ql.UnitedStates(ql.UnitedStates.NYSE)
day_count = ql.Actual365Fixed()
today = ql.Date.todaysDate()
ql.Settings.instance().evaluationDate = today

def time_to_expiration(expiration_date_str):
    exp_date = datetime.strptime(expiration_date_str, "%Y-%m-%d")
    return max((exp_date - datetime.today()).days / 365.25, 0.0001)

def safe_value(value, default):
    return value if not pd.isna(value) else default

results = []

for _, row in df.iterrows():
    try:
        # Use original market data
        K = row["Strike"]
        option_type = row["Option Type"]
        
        # Estimate spot price based on option type and strike
        if option_type == "Call":
            S = K * 1.1  # 10% above strike as estimate
        else:
            S = K * 0.9  # 10% below strike as estimate
        
        r = safe_value(row["Interpolated RF Rate"], 0.05)
        T = time_to_expiration(row["Expiration"])
        sigma = safe_value(row["IVOL_MID"], 25) / 100  # Convert from percentage
        q = 0.0  # dividend yield

        # Heston parameters (assumed)
        v0 = sigma ** 2  # Initial variance
        kappa = 2.0      # Mean reversion speed
        theta = v0       # Long-term variance
        sigma_v = 0.3    # Volatility of volatility
        rho = -0.5       # Correlation between asset and volatility

        # Handles
        spot = ql.SimpleQuote(S)
        spot_handle = ql.QuoteHandle(spot)
        rate_handle = ql.YieldTermStructureHandle(ql.FlatForward(today, r, day_count))
        div_handle = ql.YieldTermStructureHandle(ql.FlatForward(today, q, day_count))

        heston_process = ql.HestonProcess(rate_handle, div_handle, spot_handle, v0, kappa, theta, sigma_v, rho)
        heston_model = ql.HestonModel(heston_process)
        engine = ql.AnalyticHestonEngine(heston_model)

        ql_option_type = ql.Option.Call if option_type == "Call" else ql.Option.Put
        payoff = ql.PlainVanillaPayoff(ql_option_type, K)
        expiry = today + int(T * 365)
        exercise = ql.EuropeanExercise(expiry)

        option = ql.VanillaOption(payoff, exercise)
        option.setPricingEngine(engine)

        heston_price = option.NPV()
        try:
            heston_delta = option.delta()
            heston_gamma = option.gamma()
            heston_theta = option.theta() / 365
            heston_vega = option.vega() / 100
            heston_rho = option.rho() / 100
        except RuntimeError:
            heston_delta = heston_gamma = heston_theta = heston_vega = heston_rho = np.nan

        # Get market data for comparison
        market_price = safe_value(row["PX_LAST"], np.nan)
        market_delta = safe_value(row["DELTA"], np.nan)
        market_gamma = safe_value(row["GAMMA"], np.nan)
        market_theta = safe_value(row["THETA"], np.nan)
        market_vega = safe_value(row["VEGA"], np.nan)
        market_rho = safe_value(row["RHO"], np.nan)

        results.append({
            "Option Ticker": row["Option Ticker"],
            "Strike": K,
            "Option Type": option_type,
            "Estimated Spot": S,
            "Market Price": market_price,
            "Heston Price": heston_price,
            "Price Diff (Heston - Market)": heston_price - market_price if not pd.isna(heston_price) and not pd.isna(market_price) else np.nan,
            "Heston Delta": heston_delta,
            "Market Delta": market_delta,
            "Delta Diff (Heston - Market)": heston_delta - market_delta if not pd.isna(heston_delta) and not pd.isna(market_delta) else np.nan,
            "Heston Gamma": heston_gamma,
            "Market Gamma": market_gamma,
            "Gamma Diff (Heston - Market)": heston_gamma - market_gamma if not pd.isna(heston_gamma) and not pd.isna(market_gamma) else np.nan,
            "Heston Theta": heston_theta,
            "Market Theta": market_theta,
            "Theta Diff (Heston - Market)": heston_theta - market_theta if not pd.isna(heston_theta) and not pd.isna(market_theta) else np.nan,
            "Heston Vega": heston_vega,
            "Market Vega": market_vega,
            "Vega Diff (Heston - Market)": heston_vega - market_vega if not pd.isna(heston_vega) and not pd.isna(market_vega) else np.nan,
            "Heston Rho": heston_rho,
            "Market Rho": market_rho,
            "Rho Diff (Heston - Market)": heston_rho - market_rho if not pd.isna(heston_rho) and not pd.isna(market_rho) else np.nan,
            "Vol Used": sigma,
            "Risk-Free Rate": r,
            "Time to Expiry": T,
            "Heston Parameters": f"v0={v0:.4f}, κ={kappa}, θ={theta:.4f}, σ_v={sigma_v}, ρ={rho}"
        })
    except Exception as e:
        print(f"Error processing {row['Option Ticker']}: {e}")
        continue

# Create DataFrame and sort
output_df = pd.DataFrame(results)
output_df = output_df.sort_values(by=['Option Ticker', 'Option Type', 'Strike'])

# Save with formatting
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook

def save_with_formatting(df, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Heston vs Market Analysis"
    
    # Write headers
    for col, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-size columns
    for column in ws.columns:
        max_length = max(len(str(cell.value) if cell.value else "") for cell in column)
        col_letter = column[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Freeze header row
    ws.freeze_panes = "A2"

    # Add conditional formatting for key difference columns
    color_columns = ['Price Diff (Heston - Market)', 'Delta Diff (Heston - Market)', 
                     'Gamma Diff (Heston - Market)', 'Theta Diff (Heston - Market)', 
                     'Vega Diff (Heston - Market)', 'Rho Diff (Heston - Market)']
    header_row = [cell.value for cell in ws[1]]
    for col_name in color_columns:
        if col_name in header_row:
            col_idx = header_row.index(col_name) + 1
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"
            ws.conditional_formatting.add(
                cell_range,
                ColorScaleRule(
                    start_type='min', start_color='63BE7B',
                    mid_type='percentile', mid_value=50, mid_color='FFEB84',
                    end_type='max', end_color='F8696B'
                )
            )

    wb.save(filename)
    print(f"✅ Excel file saved with formatting: {filename}")

save_with_formatting(output_df, "heston_vs_market_analysis.xlsx")
print("✅ Heston vs Market analysis completed!") 