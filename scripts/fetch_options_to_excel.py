import blpapi
import pandas as pd
from datetime import datetime
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from blpapi.event import Event
import os

def connect_to_bloomberg():
    from blpapi.sessionoptions import SessionOptions
    from blpapi.session import Session
    options = SessionOptions()
    options.setServerHost("localhost")
    options.setServerPort(8194)
    session = Session(options)
    if not session.start():
        raise RuntimeError("Failed to start session.")
    if not session.openService("//blp/refdata"):
        raise RuntimeError("Failed to open //blp/refdata")
    return session

def get_option_chain(session, ticker):
    ref_data = session.getService("//blp/refdata")
    request = ref_data.createRequest("ReferenceDataRequest")
    request.append("securities", f"{ticker} US Equity")
    request.append("fields", "OPT_CHAIN")
    session.sendRequest(request)
    chain_tickers = []
    while True:
        ev = session.nextEvent(500)
        for msg in ev:
            if msg.messageType() == "ReferenceDataResponse":
                sec_data = msg.getElement("securityData")
                for sec in sec_data.values():
                    field_data = sec.getElement("fieldData")
                    if field_data.hasElement("OPT_CHAIN"):
                        chain = field_data.getElement("OPT_CHAIN")
                        for i in range(chain.numValues()):
                            opt_elem = chain.getValue(i)
                            if opt_elem.hasElement("Security Description"):
                                chain_tickers.append(opt_elem.getElementAsString("Security Description"))
        if ev.eventType() == Event.RESPONSE:
            break
    return chain_tickers

def parse_option_ticker(ticker):
    parts = ticker.split()
    if len(parts) < 5:
        return None
    expiry = parts[2]
    cp_flag = parts[3][0]
    strike = float(parts[3][1:])
    exp_date = datetime.strptime(expiry, "%m/%d/%y").strftime("%Y-%m-%d")
    option_type = 'Call' if cp_flag == 'C' else 'Put'
    return {
        "Ticker": ticker,
        "Strike": strike,
        "Expiration": exp_date,
        "Option Type": option_type
    }

def select_top_25_each(option_list, target_exp="2025-12-19"):
    calls = []
    puts = []
    for o in option_list:
        parsed = parse_option_ticker(o)
        if parsed and parsed["Expiration"] == target_exp:
            if parsed["Option Type"] == "Call":
                calls.append((o, parsed["Strike"], "Call"))
            elif parsed["Option Type"] == "Put":
                puts.append((o, parsed["Strike"], "Put"))
    calls = sorted(calls, key=lambda x: x[1])[-25:]
    puts = sorted(puts, key=lambda x: x[1])[:25]
    return calls + puts

def get_current_price(session, ticker):
    ref_data = session.getService("//blp/refdata")
    request = ref_data.createRequest("ReferenceDataRequest")
    request.append("securities", ticker)
    for field in ["PX_LAST", "PX_MID", "ASK", "BID"]:
        request.append("fields", field)
    session.sendRequest(request)
    while True:
        ev = session.nextEvent(500)
        for msg in ev:
            if msg.messageType() == "ReferenceDataResponse":
                for sec in msg.getElement("securityData").values():
                    field_data = sec.getElement("fieldData")
                    for field in ["PX_LAST", "PX_MID", "ASK", "BID"]:
                        if field_data.hasElement(field):
                            return field_data.getElementAsFloat(field)
                    return None
        if ev.eventType() == Event.RESPONSE:
            break
    return None

def fetch_option_data(session, option_tickers, current_price=None):
    ref_data = session.getService("//blp/refdata")
    results = []
    for i in range(0, len(option_tickers), 50):
        batch = option_tickers[i:i+50]
        request = ref_data.createRequest("ReferenceDataRequest")
        for ticker, _, _ in batch:
            request.append("securities", ticker)
        # Try multiple implied vol fields
        fields = ["PX_LAST", "DELTA", "GAMMA", "VEGA", "THETA", "RHO", "IMPVOL_MID", "IVOL_MID", "IMPLIED_VOLATILITY", "OPT_EXPIRE_DT", "OPT_CONTRACT_SIZE"]
        for f in fields:
            request.append("fields", f)
        session.sendRequest(request)
        while True:
            ev = session.nextEvent(500)
            for msg in ev:
                if msg.messageType() == "ReferenceDataResponse":
                    sec_data = msg.getElement("securityData")
                    for j in range(sec_data.numValues()):
                        sec = sec_data.getValueAsElement(j)
                        name = sec.getElementAsString("security")
                        field_data = sec.getElement("fieldData")
                        parts = name.split()
                        if len(parts) < 5:
                            continue
                        cp_flag = parts[3][0]
                        strike = float(parts[3][1:])
                        option_type = 'Call' if cp_flag == 'C' else 'Put'
                        if field_data.hasElement("OPT_EXPIRE_DT"):
                            exp_val = field_data.getElementAsDatetime("OPT_EXPIRE_DT")
                            expiration = exp_val.isoformat() if hasattr(exp_val, 'isoformat') else str(exp_val)
                        else:
                            expiration = None
                        # Try multiple implied vol fields
                        implied_vol = None
                        for vol_field in ["IMPVOL_MID", "IVOL_MID", "IMPLIED_VOLATILITY"]:
                            if field_data.hasElement(vol_field):
                                implied_vol = field_data.getElementAsFloat(vol_field)
                                break
                        results.append({
                            "Option Ticker": name,
                            "Strike": strike,
                            "Option Type": option_type,
                            "PX_LAST": field_data.getElementAsFloat("PX_LAST") if field_data.hasElement("PX_LAST") else None,
                            "DELTA": field_data.getElementAsFloat("DELTA") if field_data.hasElement("DELTA") else None,
                            "GAMMA": field_data.getElementAsFloat("GAMMA") if field_data.hasElement("GAMMA") else None,
                            "VEGA": field_data.getElementAsFloat("VEGA") if field_data.hasElement("VEGA") else None,
                            "THETA": field_data.getElementAsFloat("THETA") if field_data.hasElement("THETA") else None,
                            "RHO": field_data.getElementAsFloat("RHO") if field_data.hasElement("RHO") else None,
                            "Implied Volatility": implied_vol,
                            "Expiration": expiration,
                            "Contract Size": field_data.getElementAsFloat("OPT_CONTRACT_SIZE") if field_data.hasElement("OPT_CONTRACT_SIZE") else None,
                            "Current Price": current_price
                        })
            if ev.eventType() == Event.RESPONSE:
                break
        time.sleep(0.1)
    return results

def save_with_formatting(df, file_path):
    df.to_excel(file_path, index=False)
    wb = load_workbook(file_path)
    ws = wb.active
    if ws is None:
        raise ValueError("Active worksheet is None. Check the Excel file and openpyxl version.")

    header_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    # Format header row: fill, bold, center
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    # Auto-size columns
    for idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), 1):
        max_length = max(len(str(cell.value) if cell.value else "") for cell in col_cells)
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = max_length + 2
    # Freeze the top row (header)
    ws.freeze_panes = ws["A2"]
    wb.save(file_path)
    print(f"✅ Excel file saved: {file_path}")

if __name__ == "__main__":
    session = connect_to_bloomberg()
    ticker = 'NVDA'
    expiry = "2025-09-19"
    chain = get_option_chain(session, ticker)
    # Select only call options for the given expiry
    call_options = []
    for o in chain:
        parsed = parse_option_ticker(o)
        if parsed and parsed["Expiration"] == expiry and parsed["Option Type"] == "Call":
            call_options.append((o, parsed["Strike"], parsed["Option Type"]))
    print(f"Total call options selected: {len(call_options)}")
    print("Fetching market data for selected call options...")
    # Fetch the underlying's current price
    underlying_ticker = f"{ticker} US Equity"
    current_price = get_current_price(session, underlying_ticker)
    data = fetch_option_data(session, call_options, current_price=current_price)
    df = pd.DataFrame(data)
    # Ensure 'Current Price' is filled for all rows
    if "Current Price" not in df.columns or df["Current Price"].isna().any():
        df["Current Price"] = current_price
    # Sort call options by ascending strike
    df_sorted = df.sort_values(by=["Strike"], ascending=True)
    output_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "excels", "nvda_call_options_2025-09-19.xlsx"))
    save_with_formatting(df_sorted, output_path) 