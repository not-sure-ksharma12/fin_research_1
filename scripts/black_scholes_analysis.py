import pandas as pd
import QuantLib as ql
from datetime import datetime
import numpy as np

# Load your Excel file
df = pd.read_excel("analysis/bloomberg_options_top50.xlsx")

# Prepare output list
results = []

# Set up QuantLib calendar
calendar = ql.UnitedStates(ql.UnitedStates.NYSE)
day_count = ql.Actual365Fixed()
today = ql.Date.todaysDate()
ql.Settings.instance().evaluationDate = today

def time_to_expiration(expiration_date_str):
    exp_date = datetime.strptime(expiration_date_str, "%Y-%m-%d")
    t = (exp_date - datetime.today()).days / 365.25
    return max(t, 0.0001)

for _, row in df.iterrows():
    try:
        # Use spot price from the data (you may need to add this column)
        # For now, we'll use a reasonable estimate based on strike and option type
        K = row["Strike"]
        option_type = row["Option Type"]
        
        # Estimate spot price based on option type and strike
        # This is a rough estimate - ideally you'd have the actual spot price
        if option_type == "Call":
            # For calls, spot is likely above strike
            S = K * 1.1  # 10% above strike as estimate
        else:
            # For puts, spot is likely below strike
            S = K * 0.9  # 10% below strike as estimate
        
        # Helper function to safely check for NaN values
        def safe_value(value, default):
            return value if not pd.isna(value) else default
        
        r = safe_value(row["Interpolated RF Rate"], 0.05)
        T = time_to_expiration(row["Expiration"])
        sigma = safe_value(row["IVOL_MID"], 25) / 100  # Convert from percentage
        q = 0.0  # Add dividend yield if you have it

        # Create handles
        spot = ql.SimpleQuote(S)
        spot_handle = ql.QuoteHandle(spot)
        rate_handle = ql.YieldTermStructureHandle(ql.FlatForward(today, r, day_count))
        div_handle = ql.YieldTermStructureHandle(ql.FlatForward(today, q, day_count))
        vol_handle = ql.BlackVolTermStructureHandle(ql.BlackConstantVol(today, calendar, sigma, day_count))

        process = ql.BlackScholesMertonProcess(spot_handle, div_handle, rate_handle, vol_handle)

        ql_option_type = ql.Option.Call if option_type == "Call" else ql.Option.Put
        payoff = ql.PlainVanillaPayoff(ql_option_type, K)
        expiry = today + int(T * 365)
        exercise = ql.EuropeanExercise(expiry)

        option = ql.VanillaOption(payoff, exercise)
        option.setPricingEngine(ql.AnalyticEuropeanEngine(process))

        price = option.NPV()
        delta = option.delta()
        gamma = option.gamma()
        theta = option.theta() / 365
        vega = option.vega() / 100
        rho = option.rho() / 100

        # Implied volatility from market price
        market_price = safe_value(row["PX_LAST"], np.nan)
        try:
            iv = option.impliedVolatility(market_price, process, 1e-4, 100, 0.01, 2.0)
        except RuntimeError:
            iv = np.nan

        # Calculate differences safely
        market_delta = safe_value(row["DELTA"], np.nan)
        market_gamma = safe_value(row["GAMMA"], np.nan)
        market_theta = safe_value(row["THETA"], np.nan)
        market_vega = safe_value(row["VEGA"], np.nan)
        market_rho = safe_value(row["RHO"], np.nan)

        results.append({
            "Option Ticker": row["Option Ticker"],
            "Strike": K,
            "Option Type": option_type,
            "Estimated Spot": S,
            "BS Price": price,
            "Market Price": market_price,
            "Price Difference": market_price - price if not pd.isna(market_price) else np.nan,
            "BS Delta": delta,
            "Market Delta": market_delta,
            "Delta Difference": market_delta - delta if not pd.isna(market_delta) else np.nan,
            "BS Gamma": gamma,
            "Market Gamma": market_gamma,
            "Gamma Difference": market_gamma - gamma if not pd.isna(market_gamma) else np.nan,
            "BS Theta": theta,
            "Market Theta": market_theta,
            "Theta Difference": market_theta - theta if not pd.isna(market_theta) else np.nan,
            "BS Vega": vega,
            "Market Vega": market_vega,
            "Vega Difference": market_vega - vega if not pd.isna(market_vega) else np.nan,
            "BS Rho": rho,
            "Market Rho": market_rho,
            "Rho Difference": market_rho - rho if not pd.isna(market_rho) else np.nan,
            "Implied Vol (from Market)": iv,
            "Vol Used": sigma,
            "Time to Expiry": T,
            "Risk-Free Rate": r,
        })
    except Exception as e:
        print(f"Error processing {row['Option Ticker']}: {e}")
        continue

# Save to Excel with formatting
output_df = pd.DataFrame(results)

# Sort by ticker, option type, and strike
def get_sort_key(row):
    ticker = row['Option Ticker'].split()[0]
    opt_type = 0 if row['Option Type'] == 'Call' else 1
    return (ticker, opt_type, row['Strike'])

output_df = output_df.sort_values(by=['Option Ticker', 'Option Type', 'Strike'])

# Save with formatting
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook

def save_with_formatting(df, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Black-Scholes Analysis"
    
    # Write headers
    for col, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-size columns
    for column in ws.columns:
        max_length = max(len(str(cell.value) if cell.value else "") for cell in column)
        col_letter = column[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Freeze header row
    ws.freeze_panes = "A2"

    # Add conditional formatting for key columns
    color_columns = ['Price Difference', 'Delta Difference', 'Gamma Difference', 'Theta Difference', 'Vega Difference']
    header_row = [cell.value for cell in ws[1]]
    for col_name in color_columns:
        if col_name in header_row:
            col_idx = header_row.index(col_name) + 1
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"
            ws.conditional_formatting.add(
                cell_range,
                ColorScaleRule(
                    start_type='min', start_color='63BE7B',
                    mid_type='percentile', mid_value=50, mid_color='FFEB84',
                    end_type='max', end_color='F8696B'
                )
            )

    wb.save(filename)
    print(f"✅ Excel file saved with formatting: {filename}")

save_with_formatting(output_df, "analysis/black_scholes_quantlib_output.xlsx")
print("✅ Black-Scholes analysis completed!") 