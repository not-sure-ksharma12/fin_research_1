import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import logging
import os
from datetime import datetime

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def read_trade_log_comprehensive(log_filename):
    """Read trade log and return both active and exited trades"""
    active_trades = {}
    exited_trades = {}
    
    if not os.path.exists(log_filename):
        logger.warning(f"Trade log not found: {log_filename}")
        return active_trades, exited_trades
    
    try:
        with open(log_filename, 'r') as f:
            lines = f.readlines()
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # Parse log entry: [timestamp] ACTION - COMPANY - OPTION_ID - TRADE_TYPE - AMOUNT
            parts = line.split(' - ')
            if len(parts) >= 5:
                # First part contains timestamp and action: [timestamp] ACTION
                first_part = parts[0]
                if ']' in first_part:
                    timestamp_str = first_part.split(']')[0].strip('[')
                    action = first_part.split(']')[1].strip()
                else:
                    timestamp_str = first_part
                    action = parts[1].strip() if len(parts) > 1 else ""
                
                company = parts[1].strip() if ']' not in first_part else parts[1].strip()
                option_id = parts[2].strip() if ']' not in first_part else parts[2].strip()
                trade_type_part = parts[3].strip() if ']' not in first_part else parts[3].strip()
                
                # Extract amount (remove $ sign)
                amount = trade_type_part.split('$')[1].split()[0] if '$' in trade_type_part else "25.00"
                trade_type = trade_type_part.split('$')[0].strip()
                
                if action == 'ENTER':
                    active_trades[option_id] = {
                        'company': company,
                        'trade_type': trade_type,
                        'position_size': float(amount),
                        'entry_time': timestamp_str
                    }
                elif action == 'EXIT':
                    # Remove from active trades and add to exited trades
                    if option_id in active_trades:
                        trade_info = active_trades.pop(option_id)
                        trade_info['exit_time'] = timestamp_str
                        
                        # Parse PnL if available
                        if len(parts) >= 6:
                            pnl_part = parts[5]
                            if 'PnL:' in pnl_part:
                                pnl_str = pnl_part.split('PnL: $')[1].split()[0]
                                trade_info['pnl'] = float(pnl_str)
                        
                        exited_trades[option_id] = trade_info
        
        logger.info(f"Found {len(active_trades)} active trades and {len(exited_trades)} exited trades in {log_filename}")
        return active_trades, exited_trades
        
    except Exception as e:
        logger.error(f"Error reading trade log {log_filename}: {e}")
        return active_trades, exited_trades

def fix_excel_file_comprehensive(excel_filename, active_trades, exited_trades):
    """Fix Excel file with both active and exited trades"""
    if not os.path.exists(excel_filename):
        logger.warning(f"Excel file not found: {excel_filename}")
        return
    
    try:
        wb = openpyxl.load_workbook(excel_filename)
        sheets_updated = 0
        total_trades_marked = 0
        
        # Define fill colors
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # BUY trades
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # SELL trades
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')   # Exited trades
        
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith('Hour_'):
                ws = wb[sheet_name]
                updated = False
                trades_marked_in_sheet = 0
                
                # Extract hour number from sheet name
                try:
                    hour = int(sheet_name.split('_')[1])
                except:
                    continue
                
                # Determine which trades should be marked in this hour
                trades_for_this_hour = {}
                
                # Add active trades that were entered in this hour or earlier
                for option_id, trade in active_trades.items():
                    # Parse timestamp: "2025-08-05 08:01:20" -> hour = 8
                    entry_hour = int(trade['entry_time'].split(' ')[1].split(':')[0])
                    if entry_hour <= hour:
                        trades_for_this_hour[option_id] = {'trade': trade, 'status': 'Entered'}
                
                # Add exited trades that were exited in this hour
                for option_id, trade in exited_trades.items():
                    # Parse timestamp: "2025-08-05 09:02:09" -> hour = 9
                    exit_hour = int(trade['exit_time'].split(' ')[1].split(':')[0])
                    if exit_hour == hour:
                        trades_for_this_hour[option_id] = {'trade': trade, 'status': 'Exited'}
                    elif exit_hour > hour:
                        # Trade was still active in this hour
                        trades_for_this_hour[option_id] = {'trade': trade, 'status': 'Entered'}
                
                # Find column indices
                option_id_col = None
                trade_status_col = None
                trade_type_col = None
                position_size_col = None
                strike_col = None
                option_type_col = None
                
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value == 'Option_ID':
                        option_id_col = col
                    elif cell_value == 'Trade_Status':
                        trade_status_col = col
                    elif cell_value == 'Trade_Type':
                        trade_type_col = col
                    elif cell_value == 'Position_Size':
                        position_size_col = col
                    elif cell_value == 'Strike':
                        strike_col = col
                    elif cell_value == 'Option Type':
                        option_type_col = col
                
                if option_id_col and trade_status_col and trade_type_col and position_size_col:
                    # First, clear all trade status values in this sheet
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=trade_status_col).value = None
                        ws.cell(row=row, column=trade_type_col).value = None
                        ws.cell(row=row, column=position_size_col).value = None
                    
                    # Now update each row with the correct trades for this hour
                    for row in range(2, ws.max_row + 1):
                        option_id = ws.cell(row=row, column=option_id_col).value
                        
                        # Check if this trade should be marked in this hour
                        if option_id in trades_for_this_hour:
                            trade_info = trades_for_this_hour[option_id]
                            trade = trade_info['trade']
                            status = trade_info['status']
                            
                            # Update trade status
                            ws.cell(row=row, column=trade_status_col).value = status
                            ws.cell(row=row, column=trade_type_col).value = trade['trade_type']
                            ws.cell(row=row, column=position_size_col).value = trade['position_size']
                            
                            # Apply color coding
                            if status == 'Exited':
                                fill_color = green_fill
                            else:
                                fill_color = yellow_fill if trade['trade_type'] == 'BUY' else orange_fill
                            
                            for col in range(1, ws.max_column + 1):
                                ws.cell(row=row, column=col).fill = fill_color
                            
                            updated = True
                            trades_marked_in_sheet += 1
                            
                        else:
                            # Try to match by components if we have strike and option type columns
                            if strike_col and option_type_col:
                                try:
                                    strike = ws.cell(row=row, column=strike_col).value
                                    option_type = ws.cell(row=row, column=option_type_col).value
                                    
                                    if strike is not None and option_type is not None:
                                        # Try to find matching trade by components
                                        for trade_option_id, trade_info in trades_for_this_hour.items():
                                            try:
                                                # Parse option_id format: "COMPANY_STRIKE_OPTION_TYPE"
                                                parts = trade_option_id.split('_')
                                                if len(parts) >= 3:
                                                    trade_strike = float(parts[1])
                                                    trade_option_type = parts[2]
                                                    
                                                    if (strike == trade_strike and
                                                        option_type == trade_option_type):
                                                        
                                                        trade = trade_info['trade']
                                                        status = trade_info['status']
                                                        
                                                        # Update trade status
                                                        ws.cell(row=row, column=trade_status_col).value = status
                                                        ws.cell(row=row, column=trade_type_col).value = trade['trade_type']
                                                        ws.cell(row=row, column=position_size_col).value = trade['position_size']
                                                        
                                                        # Apply color coding
                                                        if status == 'Exited':
                                                            fill_color = green_fill
                                                        else:
                                                            fill_color = yellow_fill if trade['trade_type'] == 'BUY' else orange_fill
                                                        
                                                        for col in range(1, ws.max_column + 1):
                                                            ws.cell(row=row, column=col).fill = fill_color
                                                        
                                                        updated = True
                                                        trades_marked_in_sheet += 1
                                                        logger.info(f"Fixed: Matched trade by components: {trade_option_id} -> Strike:{strike}, Type:{option_type}, Status:{status}")
                                                        break
                                            except Exception as e:
                                                continue
                                except Exception as e:
                                    continue
                
                if updated:
                    sheets_updated += 1
                    total_trades_marked += trades_marked_in_sheet
                    logger.info(f"Updated sheet {sheet_name}: {trades_marked_in_sheet} trades marked")
        
        if sheets_updated > 0:
            wb.save(excel_filename)
            logger.info(f"Updated {sheets_updated} sheets in {excel_filename} with {total_trades_marked} total trades marked")
        else:
            logger.info(f"No updates needed for {excel_filename}")
            
    except Exception as e:
        logger.error(f"Error fixing Excel file {excel_filename}: {e}")

def main():
    """Main function to fix Excel files comprehensively"""
    
    # Define companies
    companies = ['NVDA', 'AUR', 'TSLA', 'SOFI', 'SOUN', 'AMD', 'AVGO', 'CRCL', 'BBAI', 'SLDB']
    
    # Fix main strategy files
    logger.info("=" * 80)
    logger.info("🔧 FIXING MAIN STRATEGY EXCEL FILES (COMPREHENSIVE)")
    logger.info("=" * 80)
    
    for company in companies:
        # Main strategy
        trade_log = f"realtime_output/multi_company_sep19/{company}_trades.log"
        excel_file = f"realtime_output/multi_company_sep19/{company}_hourly_data.xlsx"
        
        active_trades, exited_trades = read_trade_log_comprehensive(trade_log)
        fix_excel_file_comprehensive(excel_file, active_trades, exited_trades)
    
    logger.info("=" * 80)
    logger.info("✅ EXCEL FILES FIXED COMPREHENSIVELY")
    logger.info("=" * 80)

if __name__ == "__main__":
    main() 