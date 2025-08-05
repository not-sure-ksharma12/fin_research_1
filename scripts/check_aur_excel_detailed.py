import pandas as pd
import openpyxl
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def check_aur_excel_detailed():
    """Check the AUR Excel file in detail to see all trade status values"""
    
    excel_file = "realtime_output/multi_company_sep19/AUR_hourly_data.xlsx"
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file)
        
        logger.info(f"Excel file loaded: {excel_file}")
        
        # Check each hour sheet
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith('Hour_'):
                ws = wb[sheet_name]
                logger.info(f"\n=== Checking {sheet_name} ===")
                
                # Find column indices
                option_id_col = None
                trade_status_col = None
                trade_type_col = None
                position_size_col = None
                
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value == 'Option_ID':
                        option_id_col = col
                    elif cell_value == 'Trade_Status':
                        trade_status_col = col
                    elif cell_value == 'Trade_Type':
                        trade_type_col = col
                    elif cell_value == 'Position_Size':
                        position_size_col = col
                
                logger.info(f"Columns found - Option_ID: {option_id_col}, Trade_Status: {trade_status_col}, Trade_Type: {trade_type_col}")
                
                # Check all rows for any trade status values
                logger.info("All rows with trade status values:")
                for row in range(2, ws.max_row + 1):
                    option_id = ws.cell(row=row, column=option_id_col).value if option_id_col else None
                    trade_status = ws.cell(row=row, column=trade_status_col).value if trade_status_col else None
                    trade_type = ws.cell(row=row, column=trade_type_col).value if trade_type_col else None
                    position_size = ws.cell(row=row, column=position_size_col).value if position_size_col else None
                    
                    if trade_status:  # Show any row that has a trade status
                        logger.info(f"  Row {row}: {option_id} - Status: '{trade_status}' - Type: '{trade_type}' - Size: {position_size}")
                
    except Exception as e:
        logger.error(f"Error checking Excel file: {e}")

if __name__ == "__main__":
    check_aur_excel_detailed() 