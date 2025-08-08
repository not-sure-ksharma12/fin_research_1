import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import os

def test_color_coding():
    """Test the color coding logic for profits and losses"""
    
    # Create test data with exited trades
    test_data = pd.DataFrame({
        'Option_ID': ['AMD_100_Call', 'AMD_110_Call', 'AMD_120_Call', 'AMD_130_Call'],
        'Strike': [100, 110, 120, 130],
        'PX_LAST': [5.50, 3.25, 2.10, 1.50],
        'Trade_Status': ['No_Trade', 'Exited', 'Exited', 'Entered'],
        'Trade_Type': ['', 'BUY', 'SELL', 'BUY'],
        'Position_Size': [0.0, 100.0, 150.0, 200.0],
        'pnl': [0.0, 25.50, -15.75, 0.0]  # Positive PnL for BUY, Negative for SELL
    })
    
    # Test filename
    filename = "test_color_coding.xlsx"
    
    # Define colors
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # BUY entered
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # SELL entered
    light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # BUY profitable exit
    light_red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # BUY loss exit
    dark_green_fill = PatternFill(start_color='006400', end_color='006400', fill_type='solid')  # SELL profitable exit
    dark_red_fill = PatternFill(start_color='8B0000', end_color='8B0000', fill_type='solid')  # SELL loss exit
    
    print("Testing color coding logic...")
    print(f"Test data:")
    print(test_data[['Option_ID', 'Trade_Status', 'Trade_Type', 'pnl']])
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test_Color_Coding"
    
    # Write data
    for r in test_data.itertuples(index=False):
        ws.append(r)
    
    # Apply color coding (same logic as in multi_company_strategy.py)
    for row in range(2, len(test_data) + 2):  # Skip header
        trade_status = ws.cell(row=row, column=test_data.columns.get_loc('Trade_Status') + 1).value
        trade_type = ws.cell(row=row, column=test_data.columns.get_loc('Trade_Type') + 1).value
        
        print(f"\nRow {row}: Status={trade_status}, Type={trade_type}")
        
        if trade_status == 'Entered':
            if trade_type == 'BUY':
                for col in range(1, len(test_data.columns) + 1):
                    ws.cell(row=row, column=col).fill = yellow_fill
                print(f"  Applied YELLOW fill for BUY entered")
            elif trade_type == 'SELL':
                for col in range(1, len(test_data.columns) + 1):
                    ws.cell(row=row, column=col).fill = orange_fill
                print(f"  Applied ORANGE fill for SELL entered")
        
        elif trade_status == 'Exited':
            # Check if profitable by looking at PnL
            pnl_col = test_data.columns.get_loc('pnl') + 1 if 'pnl' in test_data.columns else None
            if pnl_col:
                pnl_value = ws.cell(row=row, column=pnl_col).value
                print(f"  PnL value: {pnl_value}")
                if pnl_value is not None:
                    if trade_type == 'BUY':
                        fill_color = light_green_fill if pnl_value > 0 else light_red_fill
                        color_name = "LIGHT GREEN" if pnl_value > 0 else "LIGHT RED"
                    else:  # SELL
                        fill_color = dark_green_fill if pnl_value > 0 else dark_red_fill
                        color_name = "DARK GREEN" if pnl_value > 0 else "DARK RED"
                    
                    for col in range(1, len(test_data.columns) + 1):
                        ws.cell(row=row, column=col).fill = fill_color
                    
                    print(f"  Applied {color_name} fill for {trade_type} exited with PnL: ${pnl_value:.2f}")
                else:
                    print(f"  No PnL value found")
            else:
                print(f"  PnL column not found")
    
    # Save file
    wb.save(filename)
    print(f"\n✅ Test completed! Excel file saved: {filename}")
    
    # Verify the file was created and has the expected structure
    if os.path.exists(filename):
        print(f"✅ File exists and can be opened")
        
        # Try to read it back
        wb_read = openpyxl.load_workbook(filename)
        ws_read = wb_read.active
        
        print(f"\nVerification - Reading back the file:")
        for row in range(2, len(test_data) + 2):
            option_id = ws_read.cell(row=row, column=1).value
            trade_status = ws_read.cell(row=row, column=4).value  # Trade_Status column
            trade_type = ws_read.cell(row=row, column=5).value    # Trade_Type column
            pnl = ws_read.cell(row=row, column=6).value           # pnl column
            
            # Check if any cell in the row has fill
            has_fill = any(ws_read.cell(row=row, column=col).fill.start_color.rgb != '00000000' 
                          for col in range(1, len(test_data.columns) + 1))
            
            print(f"Row {row}: {option_id} - Status: {trade_status}, Type: {trade_type}, PnL: ${pnl:.2f}, Has Fill: {has_fill}")
        
        wb_read.close()
    
    # Clean up
    if os.path.exists(filename):
        os.remove(filename)
        print(f"\nCleaned up test file: {filename}")
    
    print("\n🎨 Color coding test completed successfully!")

if __name__ == "__main__":
    test_color_coding() 