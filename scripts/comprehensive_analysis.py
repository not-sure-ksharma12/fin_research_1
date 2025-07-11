import pandas as pd
import QuantLib as ql
from datetime import datetime
import numpy as np

# Setup
calendar = ql.UnitedStates(ql.UnitedStates.NYSE)
day_count = ql.Actual365Fixed()
today = ql.Date.todaysDate()
ql.Settings.instance().evaluationDate = today

def time_to_expiration(expiration_date_str):
    exp_date = datetime.strptime(expiration_date_str, "%Y-%m-%d")
    return max((exp_date - datetime.today()).days / 365.25, 0.0001)

def safe_value(value, default):
    return value if not pd.isna(value) else default

def run_heston_analysis(input_df, is_original_data=True):
    """Run Heston analysis on given data"""
    results = []
    
    for _, row in input_df.iterrows():
        try:
            K = row["Strike"]
            option_type = row["Option Type"]
            
            if is_original_data:
                # Use original market data
                if option_type == "Call":
                    S = K * 1.1  # 10% above strike as estimate
                else:
                    S = K * 0.9  # 10% below strike as estimate
                
                r = safe_value(row["Interpolated RF Rate"], 0.05)
                T = time_to_expiration(row["Expiration"])
                sigma = safe_value(row["IVOL_MID"], 25) / 100  # Convert from percentage
                
                # Market data
                market_price = safe_value(row["PX_LAST"], np.nan)
                market_delta = safe_value(row["DELTA"], np.nan)
                market_gamma = safe_value(row["GAMMA"], np.nan)
                market_theta = safe_value(row["THETA"], np.nan)
                market_vega = safe_value(row["VEGA"], np.nan)
                market_rho = safe_value(row["RHO"], np.nan)
            else:
                # Use Black-Scholes data
                S = row["Estimated Spot"]
                r = safe_value(row["Risk-Free Rate"], 0.05)
                T = row["Time to Expiry"]
                sigma = safe_value(row["Vol Used"], 0.25)
                
                # BS and Market data
                bs_price = safe_value(row["BS Price"], np.nan)
                bs_delta = safe_value(row["BS Delta"], np.nan)
                bs_gamma = safe_value(row["BS Gamma"], np.nan)
                bs_theta = safe_value(row["BS Theta"], np.nan)
                bs_vega = safe_value(row["BS Vega"], np.nan)
                bs_rho = safe_value(row["BS Rho"], np.nan)
                
                market_price = safe_value(row["Market Price"], np.nan)
                market_delta = safe_value(row["Market Delta"], np.nan)
                market_gamma = safe_value(row["Market Gamma"], np.nan)
                market_theta = safe_value(row["Market Theta"], np.nan)
                market_vega = safe_value(row["Market Vega"], np.nan)
                market_rho = safe_value(row["Market Rho"], np.nan)
            
            q = 0.0  # dividend yield

            # Heston parameters (assumed)
            v0 = sigma ** 2  # Initial variance
            kappa = 2.0      # Mean reversion speed
            theta = v0       # Long-term variance
            sigma_v = 0.3    # Volatility of volatility
            rho = -0.5       # Correlation between asset and volatility

            # Handles
            spot = ql.SimpleQuote(S)
            spot_handle = ql.QuoteHandle(spot)
            rate_handle = ql.YieldTermStructureHandle(ql.FlatForward(today, r, day_count))
            div_handle = ql.YieldTermStructureHandle(ql.FlatForward(today, q, day_count))

            heston_process = ql.HestonProcess(rate_handle, div_handle, spot_handle, v0, kappa, theta, sigma_v, rho)
            heston_model = ql.HestonModel(heston_process)
            engine = ql.AnalyticHestonEngine(heston_model)

            ql_option_type = ql.Option.Call if option_type == "Call" else ql.Option.Put
            payoff = ql.PlainVanillaPayoff(ql_option_type, K)
            expiry = today + int(T * 365)
            exercise = ql.EuropeanExercise(expiry)

            option = ql.VanillaOption(payoff, exercise)
            option.setPricingEngine(engine)

            heston_price = option.NPV()
            try:
                heston_delta = option.delta()
                heston_gamma = option.gamma()
                heston_theta = option.theta() / 365
                heston_vega = option.vega() / 100
                heston_rho = option.rho() / 100
            except RuntimeError:
                heston_delta = heston_gamma = heston_theta = heston_vega = heston_rho = np.nan

            if is_original_data:
                # Heston vs Market analysis
                result = {
                    "Option Ticker": row["Option Ticker"],
                    "Strike": K,
                    "Option Type": option_type,
                    "Estimated Spot": S,
                    "Market Price": market_price,
                    "Heston Price": heston_price,
                    "Price Diff (Heston - Market)": heston_price - market_price if not pd.isna(heston_price) and not pd.isna(market_price) else np.nan,
                    "Heston Delta": heston_delta,
                    "Market Delta": market_delta,
                    "Delta Diff (Heston - Market)": heston_delta - market_delta if not pd.isna(heston_delta) and not pd.isna(market_delta) else np.nan,
                    "Heston Gamma": heston_gamma,
                    "Market Gamma": market_gamma,
                    "Gamma Diff (Heston - Market)": heston_gamma - market_gamma if not pd.isna(heston_gamma) and not pd.isna(market_gamma) else np.nan,
                    "Heston Theta": heston_theta,
                    "Market Theta": market_theta,
                    "Theta Diff (Heston - Market)": heston_theta - market_theta if not pd.isna(heston_theta) and not pd.isna(market_theta) else np.nan,
                    "Heston Vega": heston_vega,
                    "Market Vega": market_vega,
                    "Vega Diff (Heston - Market)": heston_vega - market_vega if not pd.isna(heston_vega) and not pd.isna(market_vega) else np.nan,
                    "Heston Rho": heston_rho,
                    "Market Rho": market_rho,
                    "Rho Diff (Heston - Market)": heston_rho - market_rho if not pd.isna(heston_rho) and not pd.isna(market_rho) else np.nan,
                    "Vol Used": sigma,
                    "Risk-Free Rate": r,
                    "Time to Expiry": T,
                    "Heston Parameters": f"v0={v0:.4f}, κ={kappa}, θ={theta:.4f}, σ_v={sigma_v}, ρ={rho}"
                }
            else:
                # Heston vs BS vs Market analysis
                result = {
                    "Option Ticker": row["Option Ticker"],
                    "Strike": K,
                    "Option Type": option_type,
                    "Estimated Spot": S,
                    "Market Price": market_price,
                    "BS Price": bs_price,
                    "Heston Price": heston_price,
                    "Price Diff (BS - Market)": bs_price - market_price if not pd.isna(bs_price) and not pd.isna(market_price) else np.nan,
                    "Price Diff (Heston - Market)": heston_price - market_price if not pd.isna(heston_price) and not pd.isna(market_price) else np.nan,
                    "Price Diff (Heston - BS)": heston_price - bs_price if not pd.isna(heston_price) and not pd.isna(bs_price) else np.nan,
                    "BS Delta": bs_delta,
                    "Heston Delta": heston_delta,
                    "Market Delta": market_delta,
                    "Delta Diff (BS - Market)": bs_delta - market_delta if not pd.isna(bs_delta) and not pd.isna(market_delta) else np.nan,
                    "Delta Diff (Heston - Market)": heston_delta - market_delta if not pd.isna(heston_delta) and not pd.isna(market_delta) else np.nan,
                    "Delta Diff (Heston - BS)": heston_delta - bs_delta if not pd.isna(heston_delta) and not pd.isna(bs_delta) else np.nan,
                    "BS Gamma": bs_gamma,
                    "Heston Gamma": heston_gamma,
                    "Market Gamma": market_gamma,
                    "Gamma Diff (BS - Market)": bs_gamma - market_gamma if not pd.isna(bs_gamma) and not pd.isna(market_gamma) else np.nan,
                    "Gamma Diff (Heston - Market)": heston_gamma - market_gamma if not pd.isna(heston_gamma) and not pd.isna(market_gamma) else np.nan,
                    "Gamma Diff (Heston - BS)": heston_gamma - bs_gamma if not pd.isna(heston_gamma) and not pd.isna(bs_gamma) else np.nan,
                    "BS Theta": bs_theta,
                    "Heston Theta": heston_theta,
                    "Market Theta": market_theta,
                    "Theta Diff (BS - Market)": bs_theta - market_theta if not pd.isna(bs_theta) and not pd.isna(market_theta) else np.nan,
                    "Theta Diff (Heston - Market)": heston_theta - market_theta if not pd.isna(heston_theta) and not pd.isna(market_theta) else np.nan,
                    "Theta Diff (Heston - BS)": heston_theta - bs_theta if not pd.isna(heston_theta) and not pd.isna(bs_theta) else np.nan,
                    "BS Vega": bs_vega,
                    "Heston Vega": heston_vega,
                    "Market Vega": market_vega,
                    "Vega Diff (BS - Market)": bs_vega - market_vega if not pd.isna(bs_vega) and not pd.isna(market_vega) else np.nan,
                    "Vega Diff (Heston - Market)": heston_vega - market_vega if not pd.isna(heston_vega) and not pd.isna(market_vega) else np.nan,
                    "Vega Diff (Heston - BS)": heston_vega - bs_vega if not pd.isna(heston_vega) and not pd.isna(bs_vega) else np.nan,
                    "BS Rho": bs_rho,
                    "Heston Rho": heston_rho,
                    "Market Rho": market_rho,
                    "Rho Diff (BS - Market)": bs_rho - market_rho if not pd.isna(bs_rho) and not pd.isna(market_rho) else np.nan,
                    "Rho Diff (Heston - Market)": heston_rho - market_rho if not pd.isna(heston_rho) and not pd.isna(market_rho) else np.nan,
                    "Rho Diff (Heston - BS)": heston_rho - bs_rho if not pd.isna(heston_rho) and not pd.isna(bs_rho) else np.nan,
                    "Vol Used": sigma,
                    "Risk-Free Rate": r,
                    "Time to Expiry": T,
                    "Heston Parameters": f"v0={v0:.4f}, κ={kappa}, θ={theta:.4f}, σ_v={sigma_v}, ρ={rho}"
                }
            
            results.append(result)
        except Exception as e:
            print(f"Error processing {row['Option Ticker']}: {e}")
            continue
    
    return pd.DataFrame(results)

def save_with_formatting(df, filename, title):
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Font, PatternFill
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = title
    
    # Write headers
    for col, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-size columns
    for column in ws.columns:
        max_length = max(len(str(cell.value) if cell.value else "") for cell in column)
        col_letter = column[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Freeze header row
    ws.freeze_panes = "A2"

    # Add conditional formatting for key difference columns
    if "Price Diff (Heston - Market)" in df.columns:
        # Heston vs Market analysis
        color_columns = ['Price Diff (Heston - Market)', 'Delta Diff (Heston - Market)', 
                         'Gamma Diff (Heston - Market)', 'Theta Diff (Heston - Market)', 
                         'Vega Diff (Heston - Market)', 'Rho Diff (Heston - Market)']
    else:
        # Comprehensive analysis
        color_columns = ['Price Diff (BS - Market)', 'Price Diff (Heston - Market)', 'Price Diff (Heston - BS)',
                         'Delta Diff (BS - Market)', 'Delta Diff (Heston - Market)', 'Delta Diff (Heston - BS)',
                         'Gamma Diff (BS - Market)', 'Gamma Diff (Heston - Market)', 'Gamma Diff (Heston - BS)',
                         'Theta Diff (BS - Market)', 'Theta Diff (Heston - Market)', 'Theta Diff (Heston - BS)',
                         'Vega Diff (BS - Market)', 'Vega Diff (Heston - Market)', 'Vega Diff (Heston - BS)']
    
    header_row = [cell.value for cell in ws[1]]
    for col_name in color_columns:
        if col_name in header_row:
            col_idx = header_row.index(col_name) + 1
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"
            ws.conditional_formatting.add(
                cell_range,
                ColorScaleRule(
                    start_type='min', start_color='63BE7B',
                    mid_type='percentile', mid_value=50, mid_color='FFEB84',
                    end_type='max', end_color='F8696B'
                )
            )

    wb.save(filename)
    print(f"✅ Excel file saved with formatting: {filename}")

# Main execution
print("🔄 Starting comprehensive Heston analysis...")

# 1. Heston vs Market (using original data)
print("📊 Running Heston vs Market analysis...")
original_data = pd.read_excel("analysis/bloomberg_options_top50.xlsx")
heston_vs_market = run_heston_analysis(original_data, is_original_data=True)
heston_vs_market = heston_vs_market.sort_values(by=['Option Ticker', 'Option Type', 'Strike'])
save_with_formatting(heston_vs_market, "analysis/heston_vs_market_direct.xlsx", "Heston vs Market")

# 2. Comprehensive BS vs Heston vs Market (using BS data)
print("📊 Running comprehensive BS vs Heston vs Market analysis...")
bs_data = pd.read_excel("analysis/black_scholes_quantlib_output.xlsx")
comprehensive_analysis = run_heston_analysis(bs_data, is_original_data=False)
comprehensive_analysis = comprehensive_analysis.sort_values(by=['Option Ticker', 'Option Type', 'Strike'])
save_with_formatting(comprehensive_analysis, "analysis/bs_heston_market_comparison.xlsx", "BS vs Heston vs Market")

print("✅ Comprehensive analysis completed!")
print("📁 Generated files:")
print("   - heston_vs_market_direct.xlsx (Heston vs Market)")
print("   - bs_heston_market_comparison.xlsx (BS vs Heston vs Market)") 