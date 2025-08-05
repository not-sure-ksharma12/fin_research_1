import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def read_trade_log(log_filename):
    """Read trade log and extract active trades"""
    active_trades = {}

    if not os.path.exists(log_filename):
        logger.warning(f"Trade log not found: {log_filename}")
        return active_trades

    try:
        with open(log_filename, 'r') as f:
            lines = f.readlines()

        for line in lines:
            if 'ENTER' in line and 'EXIT' not in line:
                # Parse trade entry
                parts = line.strip().split(' - ')
                if len(parts) >= 5:
                    timestamp = parts[0].strip('[]')
                    company = parts[2]
                    option_id = parts[3]
                    trade_type = parts[4]

                    # Extract position size if available
                    position_size = 0.0
                    for part in parts:
                        if part.startswith('$') and part.endswith('0'):
                            try:
                                position_size = float(part.replace('$', ''))
                                break
                            except:
                                continue

                    active_trades[option_id] = {
                        'trade_type': trade_type,
                        'position_size': position_size,
                        'entry_time': timestamp
                    }

            elif 'EXIT' in line:
                # Remove from active trades when exited
                parts = line.strip().split(' - ')
                if len(parts) >= 4:
                    option_id = parts[3]
                    if option_id in active_trades:
                        del active_trades[option_id]

        logger.info(f"Found {len(active_trades)} active trades in {log_filename}")
        return active_trades

    except Exception as e:
        logger.error(f"Error reading trade log {log_filename}: {e}")
        return active_trades

def fix_excel_file(excel_filename, active_trades):
    """Fix Excel file to show active trades with improved matching"""
    if not os.path.exists(excel_filename):
        logger.warning(f"Excel file not found: {excel_filename}")
        return

    try:
        wb = load_workbook(excel_filename)

        # Define colors
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # BUY entered
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # SELL entered

        sheets_updated = 0
        total_trades_marked = 0

        for sheet_name in wb.sheetnames:
            if sheet_name.startswith('Hour_'):
                ws = wb[sheet_name]
                updated = False
                trades_marked_in_sheet = 0

                # Find column indices
                option_id_col = None
                trade_status_col = None
                trade_type_col = None
                position_size_col = None
                strike_col = None
                option_type_col = None

                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value == 'Option_ID':
                        option_id_col = col
                    elif cell_value == 'Trade_Status':
                        trade_status_col = col
                    elif cell_value == 'Trade_Type':
                        trade_type_col = col
                    elif cell_value == 'Position_Size':
                        position_size_col = col
                    elif cell_value == 'Strike':
                        strike_col = col
                    elif cell_value == 'Option Type':
                        option_type_col = col

                if option_id_col and trade_status_col and trade_type_col and position_size_col:
                    # Update each row
                    for row in range(2, ws.max_row + 1):
                        option_id = ws.cell(row=row, column=option_id_col).value

                        if option_id in active_trades:
                            trade = active_trades[option_id]

                            # Update trade status
                            ws.cell(row=row, column=trade_status_col).value = 'Entered'
                            ws.cell(row=row, column=trade_type_col).value = trade['trade_type']
                            ws.cell(row=row, column=position_size_col).value = trade['position_size']

                            # Apply color coding
                            fill_color = yellow_fill if trade['trade_type'] == 'BUY' else orange_fill
                            for col in range(1, ws.max_column + 1):
                                ws.cell(row=row, column=col).fill = fill_color

                            updated = True
                            trades_marked_in_sheet += 1
                        else:
                            # Try to match by components if we have strike and option type columns
                            if strike_col and option_type_col:
                                try:
                                    strike = ws.cell(row=row, column=strike_col).value
                                    option_type = ws.cell(row=row, column=option_type_col).value

                                    if strike is not None and option_type is not None:
                                        # Try to find matching trade by components
                                        for trade_option_id, trade in active_trades.items():
                                            try:
                                                # Parse option_id format: "COMPANY_STRIKE_OPTION_TYPE"
                                                parts = trade_option_id.split('_')
                                                if len(parts) >= 3:
                                                    trade_strike = float(parts[1])
                                                    trade_option_type = parts[2]

                                                    if (strike == trade_strike and
                                                        option_type == trade_option_type):

                                                        # Update trade status
                                                        ws.cell(row=row, column=trade_status_col).value = 'Entered'
                                                        ws.cell(row=row, column=trade_type_col).value = trade['trade_type']
                                                        ws.cell(row=row, column=position_size_col).value = trade['position_size']

                                                        # Apply color coding
                                                        fill_color = yellow_fill if trade['trade_type'] == 'BUY' else orange_fill
                                                        for col in range(1, ws.max_column + 1):
                                                            ws.cell(row=row, column=col).fill = fill_color

                                                        updated = True
                                                        trades_marked_in_sheet += 1
                                                        logger.info(f"Fixed: Matched trade by components: {trade_option_id} -> Strike:{strike}, Type:{option_type}")
                                                        break
                                            except Exception as e:
                                                continue
                                except Exception as e:
                                    continue

                if updated:
                    sheets_updated += 1
                    total_trades_marked += trades_marked_in_sheet
                    logger.info(f"Updated sheet {sheet_name}: {trades_marked_in_sheet} trades marked")

        if sheets_updated > 0:
            wb.save(excel_filename)
            logger.info(f"Updated {sheets_updated} sheets in {excel_filename} with {total_trades_marked} total trades marked")
        else:
            logger.info(f"No updates needed for {excel_filename}")

    except Exception as e:
        logger.error(f"Error fixing Excel file {excel_filename}: {e}")

def main():
    """Main function to fix Excel files for both strategies"""

    # Define companies
    companies = ['NVDA', 'AUR', 'TSLA', 'SOFI', 'SOUN', 'AMD', 'AVGO', 'CRCL', 'BBAI', 'SLDB']

    # Fix main strategy files
    logger.info("=" * 80)
    logger.info("🔧 FIXING MAIN STRATEGY EXCEL FILES")
    logger.info("=" * 80)

    for company in companies:
        # Main strategy
        trade_log = f"realtime_output/multi_company_sep19/{company}_trades.log"
        excel_file = f"realtime_output/multi_company_sep19/{company}_hourly_data.xlsx"

        active_trades = read_trade_log(trade_log)
        fix_excel_file(excel_file, active_trades)

    # Fix puts strategy files
    logger.info("=" * 80)
    logger.info("🔧 FIXING PUTS STRATEGY EXCEL FILES")
    logger.info("=" * 80)

    for company in companies:
        # Puts strategy
        trade_log = f"realtime_output/multi_company_puts_sep19/{company}_puts_trades.log"
        excel_file = f"realtime_output/multi_company_puts_sep19/{company}_puts_hourly_data.xlsx"

        active_trades = read_trade_log(trade_log)
        fix_excel_file(excel_file, active_trades)

    logger.info("=" * 80)
    logger.info("✅ EXCEL FILES FIXED FOR BOTH STRATEGIES")
    logger.info("=" * 80)

if __name__ == "__main__":
    main() 