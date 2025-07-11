import blpapi
import pandas as pd
from datetime import datetime
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from blpapi.sessionoptions import SessionOptions
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.formatting.rule import ColorScaleRule
import os
os.makedirs('analysis', exist_ok=True)

# ---------- Bloomberg Setup ----------
def connect_to_bloomberg():
    options = SessionOptions()
    options.setServerHost("localhost")
    options.setServerPort(8194)
    session = blpapi.Session(options)
    if not session.start():
        raise RuntimeError("Failed to start session.")
    if not session.openService("//blp/refdata"):
        raise RuntimeError("Failed to open //blp/refdata")
    return session

# ---------- Get spot prices ----------
def get_current_price(session, tickers):
    prices = {}
    ref_data = session.getService("//blp/refdata")
    request = ref_data.createRequest("ReferenceDataRequest")
    for t in tickers:
        request.append("securities", f"{t} US Equity")
    request.append("fields", "PX_LAST")
    session.sendRequest(request)

    while True:
        ev = session.nextEvent(500)
        for msg in ev:
            if msg.messageType() == "ReferenceDataResponse":
                for sec in msg.getElement("securityData").values():
                    name = sec.getElementAsString("security").split()[0]
                    px = sec.getElement("fieldData").getElementAsFloat("PX_LAST")
                    prices[name] = px
        if ev.eventType() == blpapi.Event.RESPONSE:
            break
    return prices

# ---------- Get risk-free rate ----------
def get_risk_free_rate(session):
    ref_data = session.getService("//blp/refdata")
    request = ref_data.createRequest("ReferenceDataRequest")
    request.append("securities", "USGG3M Index")
    request.append("fields", "PX_LAST")
    session.sendRequest(request)

    while True:
        ev = session.nextEvent(500)
        for msg in ev:
            if msg.messageType() == "ReferenceDataResponse":
                sec_data = msg.getElement("securityData")
                field_data = sec_data.getValueAsElement(0).getElement("fieldData")
                return field_data.getElementAsFloat("PX_LAST") / 100
        if ev.eventType() == blpapi.Event.RESPONSE:
            break
    return 0.05

# ---------- Get valid Bloomberg option tickers from OPT_CHAIN ----------
def get_option_chain(session, ticker):
    ref_data = session.getService("//blp/refdata")
    request = ref_data.createRequest("ReferenceDataRequest")
    request.append("securities", f"{ticker} US Equity")
    request.append("fields", "OPT_CHAIN")
    session.sendRequest(request)
    chain_tickers = []

    while True:
        ev = session.nextEvent(500)
        for msg in ev:
            if msg.messageType() == "ReferenceDataResponse":
                sec_data = msg.getElement("securityData")
                for sec in sec_data.values():
                    field_data = sec.getElement("fieldData")
                    if field_data.hasElement("OPT_CHAIN"):
                        chain = field_data.getElement("OPT_CHAIN")
                        for i in range(chain.numValues()):
                            opt_elem = chain.getValue(i)
                            print(f"[DEBUG] OPT_CHAIN element {i} fields:")
                            for j in range(opt_elem.numElements()):
                                try:
                                    name = opt_elem.getElement(j).name()
                                    value = opt_elem.getElement(j).getValueAsString() if opt_elem.getElement(j).isNull() == False else str(opt_elem.getElement(j))
                                except Exception:
                                    name = str(j)
                                    value = str(opt_elem.getElement(j))
                                print(f"    {name}: {value}")
                            if opt_elem.hasElement("Security Description"):
                                chain_tickers.append(opt_elem.getElementAsString("Security Description"))
        if ev.eventType() == blpapi.Event.RESPONSE:
            break
    return chain_tickers

# ---------- Parse ticker structure ----------
def parse_option_ticker(ticker):
    parts = ticker.split()
    if len(parts) < 5:
        return None
    expiry = parts[2]
    cp_flag = parts[3][0]
    strike = float(parts[3][1:])
    exp_date = datetime.strptime(expiry, "%m/%d/%y").strftime("%Y-%m-%d")
    option_type = 'Call' if cp_flag == 'C' else 'Put'
    return {
        "Ticker": ticker,
        "Strike": strike,
        "Expiration": exp_date,
        "Option Type": option_type
    }

# ---------- Filter top 25 calls and puts expiring on target ----------
def select_top_25_each(option_list, spot, target_exp="2025-12-19"):
    calls_above = []
    calls_below = []
    puts_above = []
    puts_below = []
    for o in option_list:
        parsed = parse_option_ticker(o)
        if parsed and parsed["Expiration"] == target_exp:
            strike = parsed["Strike"]
            if parsed["Option Type"] == "Call":
                if strike > spot:
                    calls_above.append((o, strike, "Call"))
                elif strike < spot:
                    calls_below.append((o, strike, "Call"))
            elif parsed["Option Type"] == "Put":
                if strike > spot:
                    puts_above.append((o, strike, "Put"))
                elif strike < spot:
                    puts_below.append((o, strike, "Put"))
    # Sort by distance from spot
    calls_above = sorted(calls_above, key=lambda x: x[1] - spot)[:25]
    calls_below = sorted(calls_below, key=lambda x: spot - x[1])[:25]
    puts_above = sorted(puts_above, key=lambda x: x[1] - spot)[:25]
    puts_below = sorted(puts_below, key=lambda x: spot - x[1])[:25]
    return calls_above + calls_below + puts_above + puts_below

def get_treasury_yields(session):
    # Maturities in years and their Bloomberg tickers
    maturities = [
        (1/12, "USGG1M Index"),
        (0.25, "USGG3M Index"),
        (0.5, "USGG6M Index"),
        (1, "USGG1Y Index"),
        (2, "USGG2Y Index"),
        (5, "USGG5Y Index"),
        (10, "USGG10Y Index"),
        (30, "USGG30Y Index"),
    ]
    ref_data = session.getService("//blp/refdata")
    request = ref_data.createRequest("ReferenceDataRequest")
    for _, ticker in maturities:
        request.append("securities", ticker)
    request.append("fields", "PX_LAST")
    request.append("fields", "LAST_PRICE")
    request.append("fields", "YLD_YTM_MID")
    session.sendRequest(request)
    yields = {}
    while True:
        ev = session.nextEvent(500)
        for msg in ev:
            if msg.messageType() == "ReferenceDataResponse":
                for sec in msg.getElement("securityData").values():
                    ticker = sec.getElementAsString("security")
                    field_data = sec.getElement("fieldData")
                    px = None
                    # Try different field names
                    for field_name in ["PX_LAST", "LAST_PRICE", "YLD_YTM_MID"]:
                        if field_data.hasElement(field_name):
                            try:
                                px = field_data.getElementAsFloat(field_name)
                                break
                            except:
                                continue
                    if px is not None:
                        for mat, tkr in maturities:
                            if tkr.startswith(ticker.split()[0]):
                                yields[mat] = px / 100
                                print(f"[DEBUG] Treasury yield for {tkr}: {px/100:.4f}")
                                break
        if ev.eventType() == blpapi.Event.RESPONSE:
            break
    if not yields:
        print("[WARNING] No Treasury yields fetched. Using default rates.")
        yields = {0.08: 0.052, 0.25: 0.053, 0.5: 0.054, 1: 0.055, 2: 0.057, 5: 0.06, 10: 0.062, 30: 0.065}
    return yields

def interpolate_rf_rate(yields, tte):
    # tte: time to expiry in years
    mats = sorted(yields.keys())
    if tte <= mats[0]:
        return yields[mats[0]]
    if tte >= mats[-1]:
        return yields[mats[-1]]
    for i in range(1, len(mats)):
        if tte <= mats[i]:
            x0, y0 = mats[i-1], yields[mats[i-1]]
            x1, y1 = mats[i], yields[mats[i]]
            return y0 + (y1 - y0) * (tte - x0) / (x1 - x0)
    return yields[mats[-1]]

# ---------- Fetch option data ----------
def fetch_option_data(session, option_tickers, yields):
    ref_data = session.getService("//blp/refdata")
    results = []

    for i in range(0, len(option_tickers), 50):
        batch = option_tickers[i:i+50]
        request = ref_data.createRequest("ReferenceDataRequest")
        for ticker, _, _ in batch:
            request.append("securities", ticker)
        fields = [
            "PX_LAST", "BID", "ASK", "MID", "DELTA", "GAMMA", "VEGA", "THETA", "RHO",
            "IMPVOL_MID", "IMPLIED_VOLATILITY", "IVOL_MID", "OPT_EXPIRE_DT", "OPT_CONTRACT_SIZE"
        ]
        for f in fields:
            request.append("fields", f)
        session.sendRequest(request)

        while True:
            ev = session.nextEvent(500)
            for msg in ev:
                if msg.messageType() == "ReferenceDataResponse":
                    sec_data = msg.getElement("securityData")
                    for j in range(sec_data.numValues()):
                        sec = sec_data.getValueAsElement(j)
                        name = sec.getElementAsString("security")
                        field_data = sec.getElement("fieldData")
                        parts = name.split()
                        if len(parts) < 5:
                            continue
                        cp_flag = parts[3][0]
                        strike = float(parts[3][1:])
                        option_type = 'Call' if cp_flag == 'C' else 'Put'
                        # Fix: handle OPT_EXPIRE_DT as date or datetime
                        if field_data.hasElement("OPT_EXPIRE_DT"):
                            exp_val = field_data.getElementAsDatetime("OPT_EXPIRE_DT")
                            if hasattr(exp_val, 'isoformat'):
                                expiration = exp_val.isoformat()
                                exp_date = exp_val if hasattr(exp_val, 'year') else None
                            else:
                                expiration = str(exp_val)
                                exp_date = None
                        else:
                            expiration = None
                            exp_date = None
                        # Calculate time to expiry in years
                        if exp_date:
                            tte = (exp_date - datetime.today().date()).days / 365.25
                            tte = max(tte, 0.0001)
                            rf_rate = interpolate_rf_rate(yields, tte)
                        else:
                            rf_rate = None
                        result_row = {
                            "Option Ticker": name,
                            "Strike": strike,
                            "Option Type": option_type,
                            "PX_LAST": field_data.getElementAsFloat("PX_LAST") if field_data.hasElement("PX_LAST") else None,
                            "BID": field_data.getElementAsFloat("BID") if field_data.hasElement("BID") else None,
                            "ASK": field_data.getElementAsFloat("ASK") if field_data.hasElement("ASK") else None,
                            "MID": field_data.getElementAsFloat("MID") if field_data.hasElement("MID") else None,
                            "DELTA": field_data.getElementAsFloat("DELTA") if field_data.hasElement("DELTA") else None,
                            "GAMMA": field_data.getElementAsFloat("GAMMA") if field_data.hasElement("GAMMA") else None,
                            "VEGA": field_data.getElementAsFloat("VEGA") if field_data.hasElement("VEGA") else None,
                            "THETA": field_data.getElementAsFloat("THETA") if field_data.hasElement("THETA") else None,
                            "RHO": field_data.getElementAsFloat("RHO") if field_data.hasElement("RHO") else None,
                            "IMPVOL_MID": field_data.getElementAsFloat("IMPVOL_MID") if field_data.hasElement("IMPVOL_MID") else None,
                            "IMPLIED_VOLATILITY": field_data.getElementAsFloat("IMPLIED_VOLATILITY") if field_data.hasElement("IMPLIED_VOLATILITY") else None,
                            "IVOL_MID": field_data.getElementAsFloat("IVOL_MID") if field_data.hasElement("IVOL_MID") else None,
                            "Expiration": expiration,
                            "Contract Size": field_data.getElementAsFloat("OPT_CONTRACT_SIZE") if field_data.hasElement("OPT_CONTRACT_SIZE") else None,
                            "Interpolated RF Rate": rf_rate,
                        }
                        results.append(result_row)
            if ev.eventType() == blpapi.Event.RESPONSE:
                break
        time.sleep(0.1)
    # Print a sample of the returned data
    print("[DEBUG] Sample returned option data:")
    for row in results[:5]:
        print(row)
    return results

# ---------- Save with formatting ----------
def save_with_formatting(df, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Options Data"
    
    # Write headers
    for col, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-size columns
    for column in ws.columns:
        max_length = max(len(str(cell.value) if cell.value else "") for cell in column)
        col_letter = column[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Freeze header row
    ws.freeze_panes = "A2"

    # 🎨 Add conditional formatting
    color_columns = ['PX_LAST', 'IMPVOL_MID', 'DELTA', 'VEGA', 'THETA']
    header_row = [cell.value for cell in ws[1]]
    for col_name in color_columns:
        if col_name in header_row:
            col_idx = header_row.index(col_name) + 1
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"
            ws.conditional_formatting.add(
                cell_range,
                ColorScaleRule(
                    start_type='min', start_color='63BE7B',
                    mid_type='percentile', mid_value=50, mid_color='FFEB84',
                    end_type='max', end_color='F8696B'
                )
            )

    wb.save(filename)
    print(f"✅ Excel file saved with formatting: {filename}")

# ========== MAIN ==========
if __name__ == "__main__":
    session = connect_to_bloomberg()
    tickers = ['NVDA', 'SOUN', 'BBAI', 'CRCL', 'AVGO', 'AMD']
    all_options = []

    print("Fetching current prices...")
    prices = get_current_price(session, tickers)
    print("Fetching risk-free rate...")
    rfr = get_risk_free_rate(session)

    print("Fetching real option tickers from OPT_CHAIN...")
    for ticker in tickers:
        spot = prices.get(ticker)
        if spot is None:
            print(f"Could not find spot price for {ticker}. Skipping.")
            continue
        chain = get_option_chain(session, ticker)
        selected = select_top_25_each(chain, spot)
        for entry in selected:
            all_options.append(entry)

    print(f"Total options selected: {len(all_options)}")
    print("Fetching market data for selected options...")
    yields = get_treasury_yields(session)
    data = fetch_option_data(session, all_options, yields)

    for row in data:
        row["Underlying"] = row["Option Ticker"].split()[0]
        row["Current Price"] = prices.get(row["Underlying"])
        row["Risk-Free Rate"] = rfr

    # Sort data by ticker, option type (Call before Put), and strike price
    def get_ticker_type_strike(row):
        ticker = row['Option Ticker'].split()[0]
        # Option Type: 'Call' before 'Put'
        opt_type = 0 if row['Option Type'].lower() == 'call' else 1
        strike = row['Strike']
        return (ticker, opt_type, strike)
    data.sort(key=get_ticker_type_strike)
    
    # Create DataFrame and save to Excel
    df = pd.DataFrame(data)
    save_with_formatting(df, "analysis/bloomberg_options_top50.xlsx")
