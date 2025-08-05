#!/usr/bin/env python3
"""
Script to check Excel files and fix trade status
"""

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def check_and_fix_excel_trades():
    """Check Excel files and fix trade status"""
    
    # Read trade logs to see what trades were entered
    trade_logs = {}
    companies = ['NVDA', 'AUR', 'TSLA', 'SOFI', 'SOUN', 'AMD', 'AVGO', 'CRCL', 'BBAI', 'SLDB']
    
    for company in companies:
        log_file = f"realtime_output/multi_company_sep19/{company}_trades.log"
        if os.path.exists(log_file):
            with open(log_file, 'r') as f:
                lines = f.readlines()
                entered_trades = []
                for line in lines:
                    if 'ENTER' in line:
                        # Parse the trade info
                        parts = line.strip().split(' - ')
                        if len(parts) >= 4:
                            option_id = parts[2]  # e.g., "AMD_60.0_Call"
                            trade_type = parts[3]  # e.g., "BUY"
                            entered_trades.append({
                                'option_id': option_id,
                                'trade_type': trade_type
                            })
                trade_logs[company] = entered_trades
    
    print("=" * 80)
    print("📊 TRADE LOG ANALYSIS:")
    print("=" * 80)
    
    for company, trades in trade_logs.items():
        print(f"{company}: {len(trades)} trades entered")
        for trade in trades:
            print(f"  - {trade['trade_type']} {trade['option_id']}")
    
    # Now fix the Excel files
    print("\n" + "=" * 80)
    print("🔧 FIXING EXCEL FILES...")
    print("=" * 80)
    
    for company in companies:
        excel_file = f"realtime_output/multi_company_sep19/{company}_hourly_data.xlsx"
        
        if not os.path.exists(excel_file):
            print(f"❌ Excel file not found for {company}")
            continue
        
        try:
            # Load the workbook
            wb = load_workbook(excel_file)
            
            # Check each sheet
            for sheet_name in wb.sheetnames:
                if sheet_name.startswith('Hour_'):
                    ws = wb[sheet_name]
                    
                    # Find the columns
                    trade_status_col = None
                    trade_type_col = None
                    option_id_col = None
                    
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=1, column=col).value
                        if cell_value == 'Trade_Status':
                            trade_status_col = col
                        elif cell_value == 'Trade_Type':
                            trade_type_col = col
                        elif cell_value == 'Option_ID':
                            option_id_col = col
                    
                    if trade_status_col and trade_type_col and option_id_col:
                        # Check if this company has entered trades
                        if company in trade_logs:
                            for trade in trade_logs[company]:
                                # Find the row with this option_id
                                for row in range(2, ws.max_row + 1):
                                    option_id_cell = ws.cell(row=row, column=option_id_col).value
                                    if option_id_cell == trade['option_id']:
                                        # Mark as entered
                                        ws.cell(row=row, column=trade_status_col).value = 'Entered'
                                        ws.cell(row=row, column=trade_type_col).value = trade['trade_type']
                                        
                                        # Apply color coding
                                        if trade['trade_type'] == 'BUY':
                                            fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                                        else:  # SELL
                                            fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
                                        
                                        # Color the entire row
                                        for col in range(1, ws.max_column + 1):
                                            ws.cell(row=row, column=col).fill = fill
                                        
                                        print(f"✅ Fixed {company} {trade['option_id']} in {sheet_name}")
                                        break
            
            # Save the workbook
            wb.save(excel_file)
            print(f"💾 Saved updated Excel file for {company}")
            
        except Exception as e:
            print(f"❌ Error processing {company}: {e}")
    
    print("\n" + "=" * 80)
    print("✅ EXCEL FILES FIXED!")
    print("=" * 80)
    print("Check the Excel files now - the trades should be visible with proper color coding.")

if __name__ == "__main__":
    check_and_fix_excel_trades() 