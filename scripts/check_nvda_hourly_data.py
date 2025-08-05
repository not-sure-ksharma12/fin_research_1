import openpyxl
import os

EXCEL_PATH = os.path.join('realtime_output', 'multi_company_sep19', 'NVDA_hourly_data.xlsx')

if not os.path.exists(EXCEL_PATH):
    print(f"File not found: {EXCEL_PATH}")
    exit(1)

wb = openpyxl.load_workbook(EXCEL_PATH)

# Strike prices to check
target_strikes = [280, 290, 300]

for sheet_name in wb.sheetnames:
    if not sheet_name.startswith('Hour_'):
        continue
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} ===")
    headers = [cell.value for cell in ws[1]]
    col_idx = {h: i for i, h in enumerate(headers)}
    
    # Check for target strike prices
    for row in ws.iter_rows(min_row=2, values_only=True):
        option_id = row[col_idx.get('Option_ID', -1)]
        strike_price = row[col_idx.get('Strike', -1)]
        trade_status = row[col_idx.get('Trade_Status', -1)]
        trade_type = row[col_idx.get('Trade_Type', -1)]
        market_vs_heston = row[col_idx.get('Market_vs_Heston', -1)]
        option_type = row[col_idx.get('Option Type', -1)]
        
        if strike_price in target_strikes:
            print(f"Strike {strike_price} {option_type}: Market_vs_Heston = {market_vs_heston}, Trade Status = {trade_status}, Trade Type = {trade_type}")