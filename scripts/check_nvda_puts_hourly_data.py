import openpyxl
import os

EXCEL_PATH = os.path.join('realtime_output', 'multi_company_sep19', 'NVDA_hourly_data.xlsx')

if not os.path.exists(EXCEL_PATH):
    print(f"File not found: {EXCEL_PATH}")
    exit(1)

wb = openpyxl.load_workbook(EXCEL_PATH)

for sheet_name in wb.sheetnames:
    if not sheet_name.startswith('Hour_'):
        continue
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} (PUTS ONLY) ===")
    headers = [cell.value for cell in ws[1]]
    col_idx = {h: i for i, h in enumerate(headers)}
    
    # Collect all Put options for this hour
    puts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        option_type = row[col_idx.get('Option Type', -1)]
        if option_type != 'Put':
            continue
        option_id = row[col_idx.get('Option_ID', -1)]
        strike_price = row[col_idx.get('Strike', -1)]
        trade_status = row[col_idx.get('Trade_Status', -1)]
        trade_type = row[col_idx.get('Trade_Type', -1)]
        market_vs_heston = row[col_idx.get('Market_vs_Heston', -1)]
        puts.append({
            'option_id': option_id,
            'strike_price': strike_price,
            'market_vs_heston': market_vs_heston,
            'trade_status': trade_status,
            'trade_type': trade_type
        })
    if not puts:
        print("No Put options found for this hour.")
        continue
    # Sort puts by Market_vs_Heston
    sorted_puts = sorted(puts, key=lambda x: (x['market_vs_heston'] if x['market_vs_heston'] is not None else 0))
    print("All Put options sorted by Market_vs_Heston:")
    for p in sorted_puts:
        print(f"  Strike {p['strike_price']}: {p['option_id']}, Market_vs_Heston = {p['market_vs_heston']}, Trade Status = {p['trade_status']}, Trade Type = {p['trade_type']}")
    # Most underpriced and overpriced
    min_mv = min([p['market_vs_heston'] for p in sorted_puts if p['market_vs_heston'] is not None])
    max_mv = max([p['market_vs_heston'] for p in sorted_puts if p['market_vs_heston'] is not None])
    for p in sorted_puts:
        if p['market_vs_heston'] == min_mv:
            print(f"  --> MOST UNDERPRICED PUT: Strike {p['strike_price']} (Market_vs_Heston = {p['market_vs_heston']})")
        if p['market_vs_heston'] == max_mv:
            print(f"  --> MOST OVERPRICED PUT: Strike {p['strike_price']} (Market_vs_Heston = {p['market_vs_heston']})")