import pandas as pd
import openpyxl
import os
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def check_excel_contents():
    """Check what data is in the Excel files"""
    
    # Check multiple possible directories
    directories_to_check = [
        "realtime_output/multi_company_sep19",
        "scripts/realtime_output/multi_company_sep19", 
        "excels/New folder",
        "realtime_output/multi_company_aug15",
        "realtime_output/improved_multi_company_aug15"
    ]
    
    for excel_dir in directories_to_check:
        logger.info(f"\n=== Checking directory: {excel_dir} ===")
        
        if not os.path.exists(excel_dir):
            logger.warning(f"Directory {excel_dir} does not exist")
            continue
        
        excel_files = [f for f in os.listdir(excel_dir) if f.endswith('_hourly_data.xlsx')]
        
        if not excel_files:
            logger.warning(f"No Excel files found in {excel_dir}")
            continue
        
        logger.info(f"Found {len(excel_files)} Excel files in {excel_dir}")
        
        for excel_file in excel_files[:3]:  # Check first 3 files to avoid too much output
            file_path = os.path.join(excel_dir, excel_file)
            logger.info(f"\n--- Checking {excel_file} ---")
            
            try:
                # Load the workbook
                wb = openpyxl.load_workbook(file_path)
                
                # Check what sheets exist
                sheet_names = wb.sheetnames
                logger.info(f"Sheets: {sheet_names}")
                
                # Check each sheet
                for sheet_name in sheet_names:
                    if sheet_name.startswith('Hour_'):
                        ws = wb[sheet_name]
                        logger.info(f"  {sheet_name}: {ws.max_row - 1} rows of data")
                        
                        # Check if there's any data
                        if ws.max_row > 1:
                            # Get the first few rows to see what data is there
                            logger.info(f"    Sample data from {sheet_name}:")
                            for row in range(2, min(4, ws.max_row + 1)):  # Show first 2 data rows
                                row_data = []
                                for col in range(1, min(8, ws.max_column + 1)):  # Show first 7 columns
                                    cell_value = ws.cell(row=row, column=col).value
                                    row_data.append(str(cell_value) if cell_value is not None else "None")
                                logger.info(f"      Row {row}: {row_data}")
                        else:
                            logger.info(f"    {sheet_name}: No data rows")
                
            except Exception as e:
                logger.error(f"Error reading {excel_file}: {e}")

if __name__ == "__main__":
    check_excel_contents() 