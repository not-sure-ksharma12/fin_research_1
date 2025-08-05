import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime, timedelta
import logging
import os
import sys
import time
import warnings
from typing import Dict, List, Tuple, Optional
import QuantLib as ql
import pytz
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Add scripts directory to path
sys.path.append(r"C:\Users\ksharma12\fin_research\scripts")

# Import the existing modules
from fetch_options_to_excel import (
    connect_to_bloomberg, get_option_chain, parse_option_ticker, 
    get_current_price, fetch_option_data, save_with_formatting
)
from heston_calculator import calibrate_heston, heston_price_row

warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class CompanyRealTimeTrading:
    """Individual company real-time trading system for PUT OPTIONS"""
    
    def __init__(self, company: str, initial_capital: float = 100.0):
        """
        Initialize real-time trading system for a single company
        
        Args:
            company: Company ticker symbol
            initial_capital: Initial capital allocation
        """
        self.company = company
        self.initial_capital = initial_capital
        self.current_capital = initial_capital
        self.active_trades = {}  # {option_id: trade_info}
        self.trade_history = []
        self.last_processed_time = None
        self.session = None  # Bloomberg session
        
        # Trade tracking
        self.total_trades = 0
        self.profitable_trades = 0
        self.total_pnl = 0.0
        
        # Data saving and logging
        self.hourly_data_history = []
        self.daily_trades = []
        self.daily_pnl = 0.0
        self.daily_trade_count = 0
        
        # Market hours (PST)
        self.pst_tz = pytz.timezone('US/Pacific')
        
        logger.info(f"Real-time PUT trading system initialized for {self.company} with ${initial_capital} capital")
    
    def is_market_open(self) -> bool:
        """
        Check if market is open (PST timezone)
        Market hours: 6:30 AM - 1:00 PM PST (9:30 AM - 4:00 PM EST)
        """
        now_pst = datetime.now(self.pst_tz)
        current_time = now_pst.time()
        
        # Market hours: 6:30 AM - 1:00 PM PST
        market_open = datetime.strptime("06:30:00", "%H:%M:%S").time()
        market_close = datetime.strptime("13:00:00", "%H:%M:%S").time()
        
        # Check if it's a weekday and within market hours
        is_weekday = now_pst.weekday() < 5  # Monday = 0, Friday = 4
        is_market_hours = market_open <= current_time <= market_close
        
        return is_weekday and is_market_hours
    
    def fetch_live_data(self, session, expiry_date: str = "2025-09-19") -> pd.DataFrame:
        """
        Fetch live PUT options data from Bloomberg for this company
        
        Args:
            session: Bloomberg session
            expiry_date: Options expiration date
            
        Returns:
            DataFrame with current PUT options data including Heston prices
        """
        try:
            logger.info(f"Fetching live {self.company} PUT options data for expiry {expiry_date}")
            
            # Get current option chain
            chain = get_option_chain(session, self.company)
            
            # Select PUT options for the given expiry
            all_options = []
            for option in chain:
                parsed = parse_option_ticker(option)
                if parsed and parsed["Expiration"] == expiry_date and parsed["Option Type"] == "Put":
                    all_options.append((option, parsed["Strike"], parsed["Option Type"]))
            
            logger.info(f"Found {len(all_options)} live PUT options for {self.company}")
            
            # Get current stock price
            underlying_ticker = f"{self.company} US Equity"
            current_price = get_current_price(session, underlying_ticker)
            
            if not current_price:
                logger.warning(f"Could not get current price for {self.company}")
                return pd.DataFrame()
            
            # Fetch live PUT option data
            option_data = fetch_option_data(session, all_options, current_price=current_price)
            df = pd.DataFrame(option_data)
            
            if df.empty:
                logger.warning(f"No live PUT option data retrieved for {self.company}")
                return df
            
            # Ensure current price is filled
            df["Current Price"] = current_price
            
            # Sort by strike price (PUTS)
            df_sorted = df.sort_values(by=["Strike"], ascending=True)
            
            # Calculate live Heston prices
            df_with_heston = self.calculate_live_heston_prices(df_sorted)
            
            # Add current timestamp
            current_time = datetime.now()
            df_with_heston['Timestamp'] = current_time
            df_with_heston['Hour'] = current_time.hour
            
            return df_with_heston
            
        except Exception as e:
            logger.error(f"Error fetching live {self.company} PUT data: {e}")
            return pd.DataFrame()
    
    def calculate_live_heston_prices(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Calculate live Heston model prices for PUT options
        
        Args:
            df: DataFrame with live PUT options data
            
        Returns:
            DataFrame with live Heston prices added
        """
        try:
            if df.empty:
                return df
            
            # Prepare data for Heston calculation
            valuation_date = datetime.today()
            risk_free_rate = 0.0453  # 4.53% from Bloomberg
            
            # Ensure required columns exist
            if "Implied Volatility" not in df.columns or df["Implied Volatility"].isna().all():
                logger.warning(f"No implied volatility data for {self.company} PUTs, using default")
                df["Implied Volatility"] = 0.3 * 100  # 30% default
            
            # Calibrate Heston model with live PUT data
            params, heston_model, helpers, model_prices, market_prices = calibrate_heston(
                df, valuation_date, risk_free_rate=risk_free_rate
            )
            
            logger.info(f"Live Heston params for {self.company} PUTs: v0={params[0]:.4f}, kappa={params[1]:.4f}, theta={params[2]:.4f}, sigma={params[3]:.4f}, rho={params[4]:.4f}")
            
            # Calculate live Heston prices for all PUT options
            heston_results = df.apply(
                lambda row: heston_price_row(row, heston_model, valuation_date, risk_free_rate), 
                axis=1
            )
            
            # Add Heston prices to DataFrame
            df_with_heston = pd.concat([df, heston_results], axis=1)
            
            # Calculate live mispricing metrics
            df_with_heston['Market_vs_Heston'] = df_with_heston['PX_LAST'] - df_with_heston['Heston_Price']
            df_with_heston['Heston_vs_Market'] = df_with_heston['Heston_Price'] - df_with_heston['PX_LAST']
            
            # Add unique option identifier
            df_with_heston['Option_ID'] = df_with_heston.apply(
                lambda row: f"{self.company}_{row['Strike']}_{row['Option Type']}", 
                axis=1
            )
            
            return df_with_heston
        
        except Exception as e:
            logger.error(f"Error calculating live Heston prices for {self.company} PUTs: {e}")
            return df
    
    def select_live_trading_opportunities(self, data: pd.DataFrame) -> Tuple[List[str], List[str]]:
        """
        Select PUT trading opportunities based on current active positions
        - If no active trades: select 2 BUY and 2 SELL opportunities
        - If some trades active: only select opportunities to fill remaining slots
        
        Args:
            data: Current live PUT options data
            
        Returns:
            Tuple of (undervalued_option_ids, overvalued_option_ids)
        """
        # Filter for PUT options only
        put_data = data[data['Option Type'] == 'Put'].copy()
        
        # Filter for valid data (non-null prices)
        valid_data = put_data.dropna(subset=['PX_LAST', 'Heston_Price'])
        
        if len(valid_data) < 4:
            logger.warning(f"Insufficient live PUT data for {self.company}: only {len(valid_data)} valid PUT options")
            return [], []
        
        # Find undervalued PUT options (Market < Heston)
        undervalued = valid_data[valid_data['Market_vs_Heston'] < 0].copy()
        undervalued = undervalued.sort_values('Heston_vs_Market', ascending=False)
        
        # Find overvalued PUT options (Market > Heston)
        overvalued = valid_data[valid_data['Market_vs_Heston'] > 0].copy()
        overvalued = overvalued.sort_values('Market_vs_Heston', ascending=False)
        
        # Count current active trades by type
        active_buy_trades = sum(1 for trade in self.active_trades.values() if trade['trade_type'] == 'BUY')
        active_sell_trades = sum(1 for trade in self.active_trades.values() if trade['trade_type'] == 'SELL')
        
        # Calculate how many more trades we need of each type
        needed_buy_trades = max(0, 2 - active_buy_trades)
        needed_sell_trades = max(0, 2 - active_sell_trades)
        
        # Select opportunities based on what we need
        undervalued_ids = undervalued['Option_ID'].head(needed_buy_trades).tolist()
        overvalued_ids = overvalued['Option_ID'].head(needed_sell_trades).tolist()
        
        logger.info(f"Live PUT selection for {self.company}: {len(undervalued_ids)} BUY opportunities, {len(overvalued_ids)} SELL opportunities")
        logger.info(f"Current active PUT trades: {active_buy_trades} BUY, {active_sell_trades} SELL")
        
        return undervalued_ids, overvalued_ids

    def save_hourly_data(self, data: pd.DataFrame, hour: int):
        """
        Save hourly PUT data with color coding for trades - updates single file
        """
        if data.empty:
            return
        
        # Create a copy for saving
        save_data = data.copy()
        
        # Add hour and timestamp
        save_data['Hour'] = hour
        save_data['Timestamp'] = datetime.now()
        save_data['Date'] = datetime.now().strftime('%Y-%m-%d')
        
        # Add trade status columns
        save_data['Trade_Status'] = 'No_Trade'
        save_data['Trade_Type'] = ''
        save_data['Position_Size'] = 0.0
        
        # Mark all active trades (regardless of entry hour)
        for option_id, trade in self.active_trades.items():
            mask = save_data['Option_ID'] == option_id
            save_data.loc[mask, 'Trade_Status'] = 'Entered'
            save_data.loc[mask, 'Trade_Type'] = trade['trade_type']
            save_data.loc[mask, 'Position_Size'] = trade['position_size']
        
        # Mark trades that were exited this hour
        for trade in self.trade_history:
            if trade.get('exit_hour') == hour:
                mask = save_data['Option_ID'] == trade['option_id']
                save_data.loc[mask, 'Trade_Status'] = 'Exited'
                save_data.loc[mask, 'Trade_Type'] = trade['trade_type']
                save_data.loc[mask, 'Position_Size'] = trade['position_size']
        
        # Add to history
        self.hourly_data_history.append(save_data)
        
        # Update single Excel file with new hour data
        self.update_hourly_excel(save_data, hour)
    
    def update_hourly_excel(self, data: pd.DataFrame, hour: int):
        """
        Update single Excel file with new hour PUT data and color coding
        """
        # Single filename for the company (PUTS)
        filename = f"scripts/realtime_output/multi_company_puts_sep19/{self.company}_puts_hourly_data.xlsx"
        
        try:
            # Try to load existing workbook
            from openpyxl import load_workbook
            wb = load_workbook(filename)
            
            # Check if hour sheet already exists
            if f"Hour_{hour}" in wb.sheetnames:
                # Remove existing sheet for this hour
                wb.remove(wb[f"Hour_{hour}"])
            
        except FileNotFoundError:
            # Create new workbook if file doesn't exist
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
        
        # Create new sheet for this hour
        ws = wb.create_sheet(f"Hour_{hour}")
        
        # Define colors
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # BUY entered
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # SELL entered
        light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # BUY profitable exit
        light_red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # BUY loss exit
        dark_green_fill = PatternFill(start_color='006400', end_color='006400', fill_type='solid')  # SELL profitable exit
        dark_red_fill = PatternFill(start_color='8B0000', end_color='8B0000', fill_type='solid')  # SELL loss exit
        
        # Write data
        for r in dataframe_to_rows(data, index=False, header=True):
            ws.append(r)
        
        # Apply color coding
        for row in range(2, len(data) + 2):  # Skip header
            trade_status = ws.cell(row=row, column=data.columns.get_loc('Trade_Status') + 1).value
            trade_type = ws.cell(row=row, column=data.columns.get_loc('Trade_Type') + 1).value
            
            if trade_status == 'Entered':
                if trade_type == 'BUY':
                    for col in range(1, len(data.columns) + 1):
                        ws.cell(row=row, column=col).fill = yellow_fill
                elif trade_type == 'SELL':
                    for col in range(1, len(data.columns) + 1):
                        ws.cell(row=row, column=col).fill = orange_fill
            
            elif trade_status == 'Exited':
                # Check if profitable by looking at PnL
                pnl_col = data.columns.get_loc('pnl') + 1 if 'pnl' in data.columns else None
                if pnl_col:
                    pnl_value = ws.cell(row=row, column=pnl_col).value
                    if pnl_value is not None:
                        if trade_type == 'BUY':
                            fill_color = light_green_fill if pnl_value > 0 else light_red_fill
                        else:  # SELL
                            fill_color = dark_green_fill if pnl_value > 0 else dark_red_fill
                        
                        for col in range(1, len(data.columns) + 1):
                            ws.cell(row=row, column=col).fill = fill_color
        
        # Save file
        wb.save(filename)
        logger.info(f"Hourly PUT data updated for {self.company} hour {hour}: {filename}")

    def log_trade_to_text(self, trade_info: dict, action: str):
        """
        Log PUT trade information to text file
        """
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log_entry = f"[{timestamp}] {action.upper()} - {self.company} PUT - {trade_info['option_id']} - {trade_info['trade_type']} - ${trade_info['position_size']:.2f}"
        
        if action == 'EXIT':
            log_entry += f" - PnL: ${trade_info.get('pnl', 0):.2f} - Return: {trade_info.get('return_pct', 0):.2f}%"
        
        # Save to text file
        log_filename = f"scripts/realtime_output/multi_company_puts_sep19/{self.company}_puts_trades.log"
        with open(log_filename, 'a') as f:
            f.write(log_entry + '\n')
        
        logger.info(log_entry)
    
    def update_daily_stats(self, trade_info: dict = None):
        """
        Update daily PUT trading statistics
        """
        today = datetime.now().strftime('%Y-%m-%d')
        
        # Reset daily stats if it's a new day
        if not self.daily_trades or self.daily_trades[0].get('date') != today:
            self.daily_trades = []
            self.daily_pnl = 0.0
            self.daily_trade_count = 0
        
        if trade_info:
            self.daily_trades.append(trade_info)
            self.daily_pnl += trade_info.get('pnl', 0)
            self.daily_trade_count += 1
    
    def enter_live_trade(self, option_id: str, trade_type: str, data: pd.DataFrame) -> bool:
        """
        Enter a new live PUT trade if not already in position and max 4 trades not reached
        Position size is dynamic based on PUT option price and risk characteristics
        
        Args:
            option_id: Unique PUT option identifier
            trade_type: 'BUY' for undervalued, 'SELL' for overvalued
            data: Current live PUT data
            
        Returns:
            True if trade entered, False otherwise
        """
        if option_id in self.active_trades:
            logger.debug(f"Already in live PUT trade for {option_id}")
            return False
        
        # Check if we already have 4 active trades (max limit)
        if len(self.active_trades) >= 4:
            logger.debug(f"Maximum 4 PUT trades already active for {self.company}, cannot enter new trade")
            return False
        
        option_data = data[data['Option_ID'] == option_id].iloc[0]
        put_price = option_data['PX_LAST']
        
        # PUT-SPECIFIC CAPITAL ALLOCATION LOGIC
        # For PUT options, we need to consider:
        # 1. PUT option price (can be much higher than calls)
        # 2. Risk characteristics (delta, gamma)
        # 3. Maximum loss potential
        
        # Calculate available capital for this PUT trade
        is_last_trade = (len(self.active_trades) == 3)  # 0-based, so 3 means 4th trade
        
        if self.current_capital >= self.initial_capital:
            # We have profits or break-even
            if trade_type == 'BUY':
                # For BUY PUTS: Maximum loss is the premium paid
                # Use standard $25 but ensure we can afford the PUT
                position_size = min(25.0, self.current_capital * 0.25)  # 25% of capital max
            else:  # SELL PUTS
                # For SELL PUTS: Maximum loss is unlimited (stock can go to zero)
                # Be more conservative with position sizing
                position_size = min(20.0, self.current_capital * 0.20)  # 20% of capital max
        else:
            # We have losses - be more conservative
            if is_last_trade:
                # 4th trade: use remaining capital more conservatively
                allocated_capital = sum(trade['position_size'] for trade in self.active_trades.values())
                remaining_capital = self.current_capital - allocated_capital
                
                if trade_type == 'BUY':
                    position_size = min(20.0, remaining_capital * 0.8)  # 80% of remaining
                else:  # SELL PUTS
                    position_size = min(15.0, remaining_capital * 0.6)  # 60% of remaining
                
                logger.info(f"4th PUT trade detected for {self.company}: allocated=${allocated_capital:.2f}, remaining=${remaining_capital:.2f}, position_size=${position_size:.2f}")
            else:
                # First 3 trades: conservative allocation
                safe_per_trade = self.current_capital / 5  # Divide by 5 instead of 4 for safety
                
                if trade_type == 'BUY':
                    position_size = min(20.0, safe_per_trade)
                else:  # SELL PUTS
                    position_size = min(15.0, safe_per_trade * 0.8)  # 20% less for SELL PUTS
        
        # Ensure we have minimum capital to trade
        if position_size < 1.0:
            logger.warning(f"Insufficient capital for {self.company} PUT trade: ${self.current_capital:.2f}, minimum required: $1.00")
            return False
        
        # Calculate number of contracts based on PUT price
        # For PUT options, we need to be more careful about contract sizing
        if put_price > 0:
            # Calculate contracts based on position size and PUT price
            # Ensure we don't over-leverage
            max_contracts_by_price = int(position_size / put_price)
            max_contracts_by_capital = int(position_size / 100)  # $100 per contract minimum
            
            contracts = min(max_contracts_by_price, max_contracts_by_capital, 10)  # Max 10 contracts
            if contracts == 0:
                contracts = 1  # Minimum 1 contract
        else:
            contracts = 1
        
        # Adjust position size based on actual contracts
        actual_position_size = contracts * put_price
        
        trade_info = {
            'option_id': option_id,
            'trade_type': trade_type,
            'entry_time': option_data['Timestamp'],
            'entry_hour': option_data['Hour'],
            'entry_market_price': option_data['PX_LAST'],
            'entry_heston_price': option_data['Heston_Price'],
            'strike_price': option_data['Strike'],
            'stock_price': option_data['Current Price'],
            'contracts': contracts,
            'position_size': actual_position_size,
            'entry_mispricing': option_data['Market_vs_Heston'],
            'capital_at_entry': self.current_capital,
            'put_price': put_price
        }
        
        self.active_trades[option_id] = trade_info
        self.total_trades += 1
        
        # Log trade to text file
        self.log_trade_to_text(trade_info, 'ENTER')
        
        # Enhanced logging for PUT trade entry
        logger.info("=" * 80)
        logger.info(f"🎯 PUT TRADE ENTERED: {trade_type} {option_id}")
        logger.info(f"   Company: {self.company}")
        logger.info(f"   Entry Time: {option_data['Timestamp'].strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"   Entry Price: ${option_data['PX_LAST']:.2f}")
        logger.info(f"   Heston Price: ${option_data['Heston_Price']:.2f}")
        logger.info(f"   Strike Price: ${option_data['Strike']:.2f}")
        logger.info(f"   Stock Price: ${option_data['Current Price']:.2f}")
        logger.info(f"   PUT Price: ${put_price:.2f}")
        logger.info(f"   Contracts: {contracts}")
        logger.info(f"   Position Size: ${actual_position_size:.2f}")
        logger.info(f"   Mispricing: ${option_data['Market_vs_Heston']:.2f}")
        logger.info(f"   Active PUT Trades: {len(self.active_trades)}/4")
        logger.info(f"   Current Capital: ${self.current_capital:.2f}")
        logger.info(f"   Capital Status: {'PROFIT' if self.current_capital >= self.initial_capital else 'LOSS'}")
        if is_last_trade and self.current_capital < self.initial_capital:
            logger.info(f"   🎯 4TH PUT TRADE: Using remaining capital allocation")
        logger.info("=" * 80)
        
        return True
    
    def check_live_exit_conditions(self, option_id: str, data: pd.DataFrame) -> bool:
        """
        Check if exit conditions are met for a live PUT trade
        
        Args:
            option_id: PUT option identifier
            data: Current live PUT data
            
        Returns:
            True if should exit, False otherwise
        """
        if option_id not in self.active_trades:
            return False
        
        trade = self.active_trades[option_id]
        option_data = data[data['Option_ID'] == option_id]
        
        if len(option_data) == 0:
            return False
        
        current_data = option_data.iloc[0]
        current_market_price = current_data['PX_LAST']
        current_heston_price = current_data['Heston_Price']
        
        # Exit conditions for PUT options
        if trade['trade_type'] == 'BUY':
            # Exit when Market >= Heston (no longer undervalued)
            should_exit = current_market_price >= current_heston_price
        else:  # SELL
            # Exit when Market <= Heston (no longer overvalued)
            should_exit = current_market_price <= current_heston_price
        
        return should_exit
    
    def exit_live_trade(self, option_id: str, data: pd.DataFrame) -> bool:
        """
        Exit a live PUT trade and calculate PnL
        
        Args:
            option_id: PUT option identifier
            data: Current live PUT data
            
        Returns:
            True if trade exited, False otherwise
        """
        if option_id not in self.active_trades:
            return False
        
        trade = self.active_trades[option_id]
        option_data = data[data['Option_ID'] == option_id]
        
        if len(option_data) == 0:
            return False
        
        current_data = option_data.iloc[0]
        current_market_price = current_data['PX_LAST']
        
        # Calculate PnL for PUT options
        if trade['trade_type'] == 'BUY':
            # Long PUT position: profit = (exit_price - entry_price) * contracts
            pnl = (current_market_price - trade['entry_market_price']) * trade['contracts']
        else:  # SELL
            # Short PUT position: profit = (entry_price - exit_price) * contracts
            pnl = (trade['entry_market_price'] - current_market_price) * trade['contracts']
        
        # Update trade info
        trade.update({
            'exit_time': current_data['Timestamp'],
            'exit_hour': current_data['Hour'],
            'exit_market_price': current_market_price,
            'exit_heston_price': current_data['Heston_Price'],
            'pnl': pnl,
            'return_pct': (pnl / trade['position_size']) * 100,
            'duration_hours': current_data['Hour'] - trade['entry_hour']
        })
        
        # Update portfolio statistics
        self.total_pnl += pnl
        if pnl > 0:
            self.profitable_trades += 1
        
        # Update capital
        self.current_capital += pnl
        
        # Move to trade history
        self.trade_history.append(trade)
        del self.active_trades[option_id]
        
        # Update daily stats
        self.update_daily_stats(trade)
        
        # Log trade to text file
        self.log_trade_to_text(trade, 'EXIT')
        
        # Enhanced logging for PUT trade exit
        logger.info("=" * 80)
        logger.info(f"💰 PUT TRADE EXITED: {trade['trade_type']} {option_id}")
        logger.info(f"   Company: {self.company}")
        logger.info(f"   Exit Time: {current_data['Timestamp'].strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"   Entry Time: {trade['entry_time'].strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"   Duration: {trade['duration_hours']} hours")
        logger.info(f"   Entry Price: ${trade['entry_market_price']:.2f}")
        logger.info(f"   Exit Price: ${current_market_price:.2f}")
        logger.info(f"   Entry Heston: ${trade['entry_heston_price']:.2f}")
        logger.info(f"   Exit Heston: ${current_data['Heston_Price']:.2f}")
        logger.info(f"   PnL: ${pnl:.2f}")
        logger.info(f"   Return: {trade['return_pct']:.2f}%")
        logger.info(f"   Contracts: {trade['contracts']}")
        logger.info(f"   Position Size: ${trade['position_size']:.2f}")
        logger.info(f"   Capital at Entry: ${trade['capital_at_entry']:.2f}")
        logger.info(f"   Capital at Exit: ${self.current_capital:.2f}")
        active_buy_trades = sum(1 for trade in self.active_trades.values() if trade['trade_type'] == 'BUY')
        active_sell_trades = sum(1 for trade in self.active_trades.values() if trade['trade_type'] == 'SELL')
        logger.info(f"   Active PUT Trades: {len(self.active_trades)}/4 ({active_buy_trades} BUY, {active_sell_trades} SELL)")
        logger.info(f"   Total PnL: ${self.total_pnl:.2f}")
        logger.info(f"   Capital Status: {'PROFIT' if self.current_capital >= self.initial_capital else 'LOSS'}")
        logger.info("=" * 80)
        
        return True
    
    def process_live_hour(self, session):
        """
        Process one live PUT trading hour for this company
        """
        current_time = datetime.now()
        current_hour = current_time.hour
        
        # Check if market is open
        if not self.is_market_open():
            logger.info(f"Market is closed for {self.company} PUTs at {current_time.strftime('%Y-%m-%d %H:%M:%S')} PST")
            return
        
        logger.info(f"Processing live PUT hour for {self.company} at {current_time.strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Fetch fresh live PUT data
        live_data = self.fetch_live_data(session)
        if live_data.empty:
            logger.error(f"No live PUT data available for {self.company}, skipping this hour")
            return
        
        # Save hourly PUT data with color coding
        self.save_hourly_data(live_data, current_hour)
        
        # Check exit conditions for existing PUT trades
        options_to_exit = []
        for option_id in list(self.active_trades.keys()):
            if self.check_live_exit_conditions(option_id, live_data):
                options_to_exit.append(option_id)
        
        # Exit PUT trades
        for option_id in options_to_exit:
            self.exit_live_trade(option_id, live_data)
        
        # Select new PUT trading opportunities
        undervalued_ids, overvalued_ids = self.select_live_trading_opportunities(live_data)
        
        # Enter new PUT trades
        for option_id in undervalued_ids:
            self.enter_live_trade(option_id, 'BUY', live_data)
        
        for option_id in overvalued_ids:
            self.enter_live_trade(option_id, 'SELL', live_data)
        
        # Log current PUT portfolio status with detailed breakdown
        active_trades_count = len(self.active_trades)
        active_buy_trades = sum(1 for trade in self.active_trades.values() if trade['trade_type'] == 'BUY')
        active_sell_trades = sum(1 for trade in self.active_trades.values() if trade['trade_type'] == 'SELL')
        
        logger.info(f"Live PUT hour complete for {self.company}: {active_trades_count}/4 active PUT trades ({active_buy_trades} BUY, {active_sell_trades} SELL), Capital: ${self.current_capital:.2f}")
        
        # Update last processed time
        self.last_processed_time = current_time
    
    def regenerate_hourly_excel_with_active_trades(self):
        """
        Regenerate hourly Excel file to show all active PUT trades
        This is useful to fix the issue where trades don't show up in Excel
        """
        if not self.hourly_data_history:
            logger.warning(f"No hourly PUT data history for {self.company}")
            return
        
        # Get the most recent hourly PUT data
        latest_data = self.hourly_data_history[-1].copy()
        
        # Mark all active PUT trades
        for option_id, trade in self.active_trades.items():
            mask = latest_data['Option_ID'] == option_id
            latest_data.loc[mask, 'Trade_Status'] = 'Entered'
            latest_data.loc[mask, 'Trade_Type'] = trade['trade_type']
            latest_data.loc[mask, 'Position_Size'] = trade['position_size']
        
        # Update the Excel file
        hour = latest_data['Hour'].iloc[0] if 'Hour' in latest_data.columns else datetime.now().hour
        self.update_hourly_excel(latest_data, hour)
        
        logger.info(f"Regenerated hourly PUT Excel for {self.company} with {len(self.active_trades)} active PUT trades")
    
    def save_daily_stats_to_excel(self):
        """
        Save daily PUT trading statistics to an Excel file.
        """
        daily_stats_filename = f"scripts/realtime_output/multi_company_puts_sep19/{self.company}_daily_puts_stats.xlsx"
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(daily_stats_filename)
            
            # Check if sheet already exists
            if 'Daily Stats' in wb.sheetnames:
                wb.remove(wb['Daily Stats'])
            
        except FileNotFoundError:
            wb = Workbook()
            wb.remove(wb.active)
        
        ws = wb.create_sheet('Daily Stats')
        
        # Define colors
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # BUY entered
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # SELL entered
        light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # BUY profitable exit
        light_red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # BUY loss exit
        dark_green_fill = PatternFill(start_color='006400', end_color='006400', fill_type='solid')  # SELL profitable exit
        dark_red_fill = PatternFill(start_color='8B0000', end_color='8B0000', fill_type='solid')  # SELL loss exit
        
        # Write header
        headers = ['Date', 'Company', 'Total Trades', 'Profitable Trades', 'Total PnL', 'Daily Trades', 'Daily PnL', 'Current Capital']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col).value = header
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        # Write data
        row = 2
        ws.cell(row=row, column=1).value = datetime.now().strftime('%Y-%m-%d')
        ws.cell(row=row, column=2).value = self.company
        ws.cell(row=row, column=3).value = self.total_trades
        ws.cell(row=row, column=4).value = self.profitable_trades
        ws.cell(row=row, column=5).value = self.total_pnl
        ws.cell(row=row, column=6).value = self.daily_trade_count
        ws.cell(row=row, column=7).value = self.daily_pnl
        ws.cell(row=row, column=8).value = self.current_capital
        
        # Apply color coding
        if self.daily_pnl > 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = light_green_fill
        elif self.daily_pnl < 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = light_red_fill
        
        # Save file
        wb.save(daily_stats_filename)
        logger.info(f"Daily PUT stats updated for {self.company}: {daily_stats_filename}")

class MultiCompanyRealTimeTrading:
    """Multi-company real-time PUT trading system"""
    
    def __init__(self, companies: List[str], capital_per_company: float = 100.0):
        """
        Initialize multi-company real-time PUT trading system
        
        Args:
            companies: List of company ticker symbols
            capital_per_company: Capital allocation per company
        """
        self.companies = companies
        self.capital_per_company = capital_per_company
        self.session = None
        self.is_running = False
        
        # Initialize individual company PUT trading systems
        self.company_systems = {}
        for company in companies:
            self.company_systems[company] = CompanyRealTimeTrading(company, capital_per_company)
        
        # Create output directories for PUTS
        self.output_dir = "scripts/realtime_output/multi_company_puts_sep19"
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Portfolio tracking
        self.total_portfolio_capital = capital_per_company * len(companies)
        self.current_portfolio_capital = self.total_portfolio_capital
        self.portfolio_trade_history = []
        
        logger.info(f"Multi-company real-time PUT trading system initialized for {len(companies)} companies")
        logger.info(f"Companies: {', '.join(companies)}")
        logger.info(f"Capital per company: ${capital_per_company}")
        logger.info(f"Total portfolio capital: ${self.total_portfolio_capital}")
        logger.info(f"Portfolio allocation: {len(companies)} companies × ${capital_per_company} = ${self.total_portfolio_capital}")

    def run_realtime_trading(self):
        """
        Main loop for real-time PUT trading across all companies.
        """
        self.is_running = True
        while self.is_running:
            for company in self.companies:
                company_system = self.company_systems[company]
                
                # Connect to Bloomberg session if not already connected
                if not company_system.session:
                    company_system.session = connect_to_bloomberg()
                    if not company_system.session:
                        logger.error(f"Could not connect to Bloomberg for {company}. Exiting.")
                        self.is_running = False
                        return
                
                # Process the current hour for this company
                company_system.process_live_hour(company_system.session)
                
                # Regenerate Excel to show active trades
                company_system.regenerate_hourly_excel_with_active_trades()
                
                # Save daily stats to Excel
                company_system.save_daily_stats_to_excel()
                
                # Check if any company needs to exit all trades (e.g., market closed)
                if not company_system.is_market_open():
                    logger.info(f"Market closed for {company}. Exiting all PUT trades.")
                    for option_id in list(company_system.active_trades.keys()):
                        company_system.exit_live_trade(option_id, company_system.fetch_live_data(company_system.session))
                    company_system.regenerate_hourly_excel_with_active_trades()
                    company_system.save_daily_stats_to_excel()
                    time.sleep(60) # Wait for market to reopen
                    continue
                
                time.sleep(60) # Wait for the next minute

    def save_daily_stats_to_excel(self):
        """
        Save daily PUT trading statistics to an Excel file.
        """
        daily_stats_filename = f"scripts/realtime_output/multi_company_puts_sep19/daily_puts_stats.xlsx"
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(daily_stats_filename)
            
            # Check if sheet already exists
            if 'Daily Stats' in wb.sheetnames:
                wb.remove(wb['Daily Stats'])
            
        except FileNotFoundError:
            wb = Workbook()
            wb.remove(wb.active)
        
        ws = wb.create_sheet('Daily Stats')
        
        # Define colors
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # BUY entered
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # SELL entered
        light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # BUY profitable exit
        light_red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # BUY loss exit
        dark_green_fill = PatternFill(start_color='006400', end_color='006400', fill_type='solid')  # SELL profitable exit
        dark_red_fill = PatternFill(start_color='8B0000', end_color='8B0000', fill_type='solid')  # SELL loss exit
        
        # Write header
        headers = ['Date', 'Company', 'Total Trades', 'Profitable Trades', 'Total PnL', 'Daily Trades', 'Daily PnL']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col).value = header
            ws.cell(row=1, column=col).fill = Font(bold=True)
        
        # Write data
        row = 2
        for company in self.companies:
            company_system = self.company_systems[company]
            ws.cell(row=row, column=1).value = datetime.now().strftime('%Y-%m-%d')
            ws.cell(row=row, column=2).value = company
            ws.cell(row=row, column=3).value = company_system.total_trades
            ws.cell(row=row, column=4).value = company_system.profitable_trades
            ws.cell(row=row, column=5).value = company_system.total_pnl
            ws.cell(row=row, column=6).value = company_system.daily_trade_count
            ws.cell(row=row, column=7).value = company_system.daily_pnl
            
            # Apply color coding
            if company_system.daily_pnl > 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = light_green_fill
            elif company_system.daily_pnl < 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = dark_red_fill
            
            row += 1
        
        # Save file
        wb.save(daily_stats_filename)
        logger.info(f"Daily PUT stats updated: {daily_stats_filename}")

def main():
    """Main function to run the multi-company real-time PUT trading system"""
    
    # Define the 9 companies
    companies = ['NVDA', 'AUR', 'TSLA', 'SOFI', 'SOUN', 'AMD', 'AVGO', 'CRCL', 'BBAI', 'SLDB']
    
    # Initialize multi-company real-time PUT trading system
    trading_system = MultiCompanyRealTimeTrading(companies, capital_per_company=100.0)
    
    # Start multi-company real-time PUT trading
    trading_system.run_realtime_trading()


if __name__ == "__main__":
    main() 