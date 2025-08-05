import pandas as pd
import openpyxl
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def debug_aur_excel():
    """Debug the AUR Excel file to see what options are available"""
    
    excel_file = "realtime_output/multi_company_sep19/AUR_hourly_data.xlsx"
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file)
        
        logger.info(f"Excel file loaded: {excel_file}")
        
        # Check each hour sheet
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith('Hour_'):
                ws = wb[sheet_name]
                logger.info(f"\n=== Checking {sheet_name} ===")
                
                # Find column indices
                option_id_col = None
                strike_col = None
                option_type_col = None
                
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value == 'Option_ID':
                        option_id_col = col
                    elif cell_value == 'Strike':
                        strike_col = col
                    elif cell_value == 'Option Type':
                        option_type_col = col
                
                logger.info(f"Columns found - Option_ID: {option_id_col}, Strike: {strike_col}, Option Type: {option_type_col}")
                
                # List all options in this sheet
                logger.info("All options in this sheet:")
                for row in range(2, ws.max_row + 1):
                    option_id = ws.cell(row=row, column=option_id_col).value if option_id_col else None
                    strike = ws.cell(row=row, column=strike_col).value if strike_col else None
                    option_type = ws.cell(row=row, column=option_type_col).value if option_type_col else None
                    
                    if option_id:
                        logger.info(f"  Row {row}: {option_id} - Strike: {strike}, Type: {option_type}")
                
    except Exception as e:
        logger.error(f"Error checking Excel file: {e}")

if __name__ == "__main__":
    debug_aur_excel() 