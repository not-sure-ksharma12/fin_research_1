import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime, timedelta
import logging
import os
import sys
import time
import warnings
from typing import Dict, List, Tuple, Optional
import QuantLib as ql
import pytz
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Add scripts directory to path
sys.path.append(r"C:\Users\ksharma12\fin_research\scripts")

# Import the existing modules
from fetch_options_to_excel import (
    connect_to_bloomberg, get_option_chain, parse_option_ticker, 
    get_current_price, fetch_option_data, save_with_formatting
)
from heston_calculator import calibrate_heston, heston_price_row

warnings.filterwarnings('ignore')

# Configure logging with file handler
import logging.handlers

# Create logs directory
logs_dir = "scripts/realtime_output/multi_company_sep19/logs"
os.makedirs(logs_dir, exist_ok=True)

# Configure logging to both console and file
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# Create formatters
console_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

# Console handler
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_handler.setFormatter(console_formatter)

# File handler with rotation (daily)
log_filename = f"{logs_dir}/trading_activities.log"
file_handler = logging.handlers.TimedRotatingFileHandler(
    log_filename, 
    when='midnight', 
    interval=1, 
    backupCount=30,  # Keep 30 days of logs
    encoding='utf-8'
)
file_handler.setLevel(logging.INFO)
file_handler.setFormatter(file_formatter)

# Add handlers to logger
logger.addHandler(console_handler)
logger.addHandler(file_handler)

# Also create a separate detailed log for all activities
detailed_log_filename = f"{logs_dir}/detailed_activities.log"
detailed_handler = logging.FileHandler(detailed_log_filename, mode='a', encoding='utf-8')
detailed_handler.setLevel(logging.DEBUG)
detailed_handler.setFormatter(file_formatter)
logger.addHandler(detailed_handler)

logger.info("=" * 80)
logger.info("🚀 MULTI-COMPANY REAL-TIME TRADING SYSTEM STARTING")
logger.info("=" * 80)
logger.info(f"Log files will be saved to: {logs_dir}")
logger.info(f"Main log: {log_filename}")
logger.info(f"Detailed log: {detailed_log_filename}")
logger.info("=" * 80)

class CompanyRealTimeTrading:
    """Individual company real-time trading system"""
    
    def __init__(self, company: str, initial_capital: float = 100.0):
        """
        Initialize real-time trading system for a single company
        
        Args:
            company: Company ticker symbol
            initial_capital: Initial capital allocation
        """
        self.company = company
        self.initial_capital = initial_capital
        self.current_capital = initial_capital
        self.active_trades = {}  # {option_id: trade_info}
        self.trade_history = []
        self.last_processed_time = None
        
        # Trade tracking
        self.total_trades = 0
        self.profitable_trades = 0
        self.total_pnl = 0.0
        
        # Data saving and logging
        self.hourly_data_history = []
        self.daily_trades = []
        self.daily_pnl = 0.0
        self.daily_trade_count = 0
        
        # Market hours (PST)
        self.pst_tz = pytz.timezone('US/Pacific')
        
        logger.info(f"Real-time trading system initialized for {self.company} with ${initial_capital} capital")
        
        # Log initialization to terminal activities
        self.save_terminal_activities(
            "SYSTEM", 
            "Trading system initialized", 
            {"initial_capital": f"${initial_capital}", "company": self.company}
        )
    
    def is_market_open(self) -> bool:
        """
        Check if market is open (PST timezone)
        Market hours: 6:30 AM - 1:00 PM PST (9:30 AM - 4:00 PM EST)
        """
        now_pst = datetime.now(self.pst_tz)
        current_time = now_pst.time()
        
        # Market hours: 6:30 AM - 1:00 PM PST
        market_open = datetime.strptime("06:30:00", "%H:%M:%S").time()
        market_close = datetime.strptime("13:00:00", "%H:%M:%S").time()
        
        # Check if it's a weekday and within market hours
        is_weekday = now_pst.weekday() < 5  # Monday = 0, Friday = 4
        is_market_hours = market_open <= current_time <= market_close
        
        return is_weekday and is_market_hours
    
    def fetch_live_data(self, session, expiry_date: str = "2025-09-19") -> pd.DataFrame:
        """
        Fetch live options data from Bloomberg for this company
        
        Args:
            session: Bloomberg session
            expiry_date: Options expiration date
            
        Returns:
            DataFrame with current options data including Heston prices
        """
        try:
            logger.info(f"Fetching live {self.company} options data for expiry {expiry_date}")
            
            # Log data fetching to terminal activities
            self.save_terminal_activities(
                "DATA_FETCH", 
                "Fetching live options data", 
                {"expiry_date": expiry_date, "company": self.company}
            )
            
            # Get current option chain
            chain = get_option_chain(session, self.company)
            
            # Select options for the given expiry
            all_options = []
            for option in chain:
                parsed = parse_option_ticker(option)
                if parsed and parsed["Expiration"] == expiry_date:
                    all_options.append((option, parsed["Strike"], parsed["Option Type"]))
            
            logger.info(f"Found {len(all_options)} live options for {self.company}")
            
            # Get current stock price
            underlying_ticker = f"{self.company} US Equity"
            current_price = get_current_price(session, underlying_ticker)
            
            if not current_price:
                logger.warning(f"Could not get current price for {self.company}")
                return pd.DataFrame()
            
            # Fetch live option data
            option_data = fetch_option_data(session, all_options, current_price=current_price)
            df = pd.DataFrame(option_data)
            
            if df.empty:
                logger.warning(f"No live option data retrieved for {self.company}")
                return df
            
            # Ensure current price is filled
            df["Current Price"] = current_price
            
            # Filter for Call options only and sort by strike price
            df_calls = df[df["Option Type"] == "Call"].sort_values(by=["Strike"], ascending=True)
            
            if df_calls.empty:
                logger.warning(f"No Call options found for {self.company}")
                return pd.DataFrame()
            
            logger.info(f"Found {len(df_calls)} Call options for {self.company}")
            df_sorted = df_calls
            
            # Calculate live Heston prices
            df_with_heston = self.calculate_live_heston_prices(df_sorted)
            
            # Add current timestamp
            current_time = datetime.now()
            df_with_heston['Timestamp'] = current_time
            df_with_heston['Hour'] = current_time.hour
            
            return df_with_heston
            
        except Exception as e:
            logger.error(f"Error fetching live {self.company} data: {e}")
            return pd.DataFrame()
    
    def calculate_live_heston_prices(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Calculate live Heston model prices for options
        
        Args:
            df: DataFrame with live options data
            
        Returns:
            DataFrame with live Heston prices added
        """
        try:
            if df.empty:
                return df
            
            # Prepare data for Heston calculation
            valuation_date = datetime.today()
            risk_free_rate = 0.0453  # 4.53% from Bloomberg
            
            # Ensure required columns exist
            if "Implied Volatility" not in df.columns or df["Implied Volatility"].isna().all():
                logger.warning(f"No implied volatility data for {self.company}, using default")
                df["Implied Volatility"] = 0.3 * 100  # 30% default
            
            # Calibrate Heston model with live data
            params, heston_model, helpers, model_prices, market_prices = calibrate_heston(
                df, valuation_date, risk_free_rate=risk_free_rate
            )
            
            logger.info(f"Live Heston params for {self.company}: v0={params[0]:.4f}, kappa={params[1]:.4f}, theta={params[2]:.4f}, sigma={params[3]:.4f}, rho={params[4]:.4f}")
            
            # Calculate live Heston prices for all options
            heston_results = df.apply(
                lambda row: heston_price_row(row, heston_model, valuation_date, risk_free_rate), 
                axis=1
            )
            
            # Add Heston prices to DataFrame
            df_with_heston = pd.concat([df, heston_results], axis=1)
            
            # Calculate live mispricing metrics
            df_with_heston['Market_vs_Heston'] = df_with_heston['PX_LAST'] - df_with_heston['Heston_Price']
            df_with_heston['Heston_vs_Market'] = df_with_heston['Heston_Price'] - df_with_heston['PX_LAST']
            
            # Add unique option identifier
            df_with_heston['Option_ID'] = df_with_heston.apply(
                lambda row: f"{self.company}_{row['Strike']}_{row['Option Type']}", 
                axis=1
            )
            
            return df_with_heston
        
        except Exception as e:
            logger.error(f"Error calculating live Heston prices for {self.company}: {e}")
            return df
    
    def save_hourly_data(self, data: pd.DataFrame, hour: int):
        """
        Save hourly data with color coding for trades - updates single file
        """
        if data.empty:
            return
        
        # Create a copy for saving
        save_data = data.copy()
        
        # Add hour and timestamp
        save_data['Hour'] = hour
        save_data['Timestamp'] = datetime.now()
        save_data['Date'] = datetime.now().strftime('%Y-%m-%d')
        
        # Add trade status columns
        save_data['Trade_Status'] = 'No_Trade'
        save_data['Trade_Type'] = ''
        save_data['Position_Size'] = 0.0
        
        # Mark trades based on hour logic
        active_trades_marked = 0
        for option_id, trade in self.active_trades.items():
            # Check if this trade should be marked in this hour
            entry_time = trade.get('entry_time')
            if entry_time:
                try:
                    # Parse entry time to get hour
                    if isinstance(entry_time, str):
                        entry_hour = int(entry_time.split(' ')[1].split(':')[0])
                    else:
                        # entry_time is a datetime object
                        entry_hour = entry_time.hour
                    
                    # Only mark trades that were entered in this hour or earlier
                    if entry_hour <= hour:
                        # Try exact match first
                        mask = save_data['Option_ID'] == option_id
                        if mask.any():
                            save_data.loc[mask, 'Trade_Status'] = 'Entered'
                            save_data.loc[mask, 'Trade_Type'] = trade['trade_type']
                            save_data.loc[mask, 'Position_Size'] = trade['position_size']
                            active_trades_marked += 1
                            logger.info(f"Marked active trade in Excel (hour {hour}): {option_id} - {trade['trade_type']} - ${trade['position_size']:.2f} (entered hour {entry_hour})")
                        else:
                            # Try to find by parsing the option_id components
                            try:
                                # Parse option_id format: "COMPANY_STRIKE_OPTION_TYPE"
                                parts = option_id.split('_')
                                if len(parts) >= 3:
                                    company = parts[0]
                                    strike = float(parts[1])
                                    option_type = parts[2]
                                    
                                    # Find matching option in current data
                                    option_mask = (
                                        (save_data['Strike'] == strike) & 
                                        (save_data['Option Type'] == option_type)
                                    )
                                    
                                    if option_mask.any():
                                        save_data.loc[option_mask, 'Trade_Status'] = 'Entered'
                                        save_data.loc[option_mask, 'Trade_Type'] = trade['trade_type']
                                        save_data.loc[option_mask, 'Position_Size'] = trade['position_size']
                                        active_trades_marked += 1
                                        logger.info(f"Marked active trade by components (hour {hour}): {option_id} - {trade['trade_type']} - ${trade['position_size']:.2f} (entered hour {entry_hour})")
                                    else:
                                        logger.warning(f"Active trade {option_id} not found in current data for {self.company} (no matching strike/type)")
                                else:
                                    logger.warning(f"Active trade {option_id} has invalid format for {self.company}")
                            except Exception as e:
                                logger.warning(f"Error parsing active trade {option_id} for {self.company}: {e}")
                    else:
                        logger.info(f"Skipping trade {option_id} for hour {hour} (entered at hour {entry_hour})")
                except Exception as e:
                    logger.warning(f"Error parsing entry time for trade {option_id}: {e}")
                    # Fallback: mark the trade if we can't parse the time
                    mask = save_data['Option_ID'] == option_id
                    if mask.any():
                        save_data.loc[mask, 'Trade_Status'] = 'Entered'
                        save_data.loc[mask, 'Trade_Type'] = trade['trade_type']
                        save_data.loc[mask, 'Position_Size'] = trade['position_size']
                        active_trades_marked += 1
                        logger.info(f"Marked active trade (fallback): {option_id} - {trade['trade_type']} - ${trade['position_size']:.2f}")
        
        # Mark trades that were exited this hour
        exited_trades_marked = 0
        for trade in self.trade_history:
            if trade.get('exit_hour') == hour:
                mask = save_data['Option_ID'] == trade['option_id']
                if mask.any():
                    save_data.loc[mask, 'Trade_Status'] = 'Exited'
                    save_data.loc[mask, 'Trade_Type'] = trade['trade_type']
                    save_data.loc[mask, 'Position_Size'] = trade['position_size']
                    exited_trades_marked += 1
                    logger.info(f"Marked exited trade in Excel: {trade['option_id']} - {trade['trade_type']} - ${trade['position_size']:.2f}")
        
        # Add to history
        self.hourly_data_history.append(save_data)
        
        # Update single Excel file with new hour data
        self.update_hourly_excel(save_data, hour)
        
        # Log summary of what was saved
        logger.info(f"Saved hourly data for {self.company} hour {hour}: {active_trades_marked} active trades marked, {exited_trades_marked} exited trades marked")
        
        # If we have active trades but couldn't mark them in current data, force regenerate
        if len(self.active_trades) > 0 and active_trades_marked < len(self.active_trades):
            logger.warning(f"Only marked {active_trades_marked}/{len(self.active_trades)} active trades for {self.company}. Forcing regeneration...")
            self.regenerate_hourly_excel_with_active_trades()
    
    def update_hourly_excel(self, data: pd.DataFrame, hour: int):
        """
        Update single Excel file with new hour data and color coding
        """
        # Single filename for the company
        filename = f"scripts/realtime_output/multi_company_sep19/{self.company}_hourly_data.xlsx"
        
        try:
            # Try to load existing workbook
            from openpyxl import load_workbook
            wb = load_workbook(filename)
            
            # Check if hour sheet already exists
            if f"Hour_{hour}" in wb.sheetnames:
                # Remove existing sheet for this hour
                wb.remove(wb[f"Hour_{hour}"])
            
        except FileNotFoundError:
            # Create new workbook if file doesn't exist
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
        
        # Create new sheet for this hour
        ws = wb.create_sheet(f"Hour_{hour}")
        
        # Define colors
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # BUY entered
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # SELL entered
        light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # BUY profitable exit
        light_red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # BUY loss exit
        dark_green_fill = PatternFill(start_color='006400', end_color='006400', fill_type='solid')  # SELL profitable exit
        dark_red_fill = PatternFill(start_color='8B0000', end_color='8B0000', fill_type='solid')  # SELL loss exit
        
        # Write data
        for r in dataframe_to_rows(data, index=False, header=True):
            ws.append(r)
        
        # Apply color coding
        for row in range(2, len(data) + 2):  # Skip header
            trade_status = ws.cell(row=row, column=data.columns.get_loc('Trade_Status') + 1).value
            trade_type = ws.cell(row=row, column=data.columns.get_loc('Trade_Type') + 1).value
            
            if trade_status == 'Entered':
                if trade_type == 'BUY':
                    for col in range(1, len(data.columns) + 1):
                        ws.cell(row=row, column=col).fill = yellow_fill
                elif trade_type == 'SELL':
                    for col in range(1, len(data.columns) + 1):
                        ws.cell(row=row, column=col).fill = orange_fill
            
            elif trade_status == 'Exited':
                # Check if profitable by looking at PnL
                pnl_col = data.columns.get_loc('pnl') + 1 if 'pnl' in data.columns else None
                if pnl_col:
                    pnl_value = ws.cell(row=row, column=pnl_col).value
                    if pnl_value is not None:
                        if trade_type == 'BUY':
                            fill_color = light_green_fill if pnl_value > 0 else light_red_fill
                        else:  # SELL
                            fill_color = dark_green_fill if pnl_value > 0 else dark_red_fill
                        
                        for col in range(1, len(data.columns) + 1):
                            ws.cell(row=row, column=col).fill = fill_color
        
        # Save file
        wb.save(filename)
        logger.info(f"Hourly data updated for {self.company} hour {hour}: {filename}")
    
    def save_hourly_excel(self, data: pd.DataFrame, hour: int):
        """
        Legacy method - now calls update_hourly_excel
        """
        self.update_hourly_excel(data, hour)
    
    def log_trade_to_text(self, trade_info: dict, action: str):
        """
        Log trade information to text file and terminal activities
        """
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log_entry = f"[{timestamp}] {action.upper()} - {self.company} - {trade_info['option_id']} - {trade_info['trade_type']} - ${trade_info['position_size']:.2f}"
        
        if action == 'EXIT':
            log_entry += f" - PnL: ${trade_info.get('pnl', 0):.2f} - Return: {trade_info.get('return_pct', 0):.2f}%"
        
        # Save to text file
        log_filename = f"scripts/realtime_output/multi_company_sep19/{self.company}_trades.log"
        with open(log_filename, 'a') as f:
            f.write(log_entry + '\n')
        
        # Save to terminal activities log
        terminal_log_filename = f"scripts/realtime_output/multi_company_sep19/logs/{self.company}_terminal_activities.log"
        with open(terminal_log_filename, 'a') as f:
            f.write(log_entry + '\n')
        
        logger.info(log_entry)
    
    def save_terminal_activities(self, activity_type: str, message: str, details: dict = None):
        """
        Save terminal activities to log files
        
        Args:
            activity_type: Type of activity (TRADE, SYSTEM, ERROR, etc.)
            message: Activity message
            details: Additional details as dictionary
        """
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Create detailed log entry
        log_entry = f"[{timestamp}] {activity_type.upper()} - {self.company} - {message}"
        
        if details:
            details_str = " | ".join([f"{k}: {v}" for k, v in details.items()])
            log_entry += f" | {details_str}"
        
        # Save to company-specific terminal activities log
        terminal_log_filename = f"scripts/realtime_output/multi_company_sep19/logs/{self.company}_terminal_activities.log"
        with open(terminal_log_filename, 'a') as f:
            f.write(log_entry + '\n')
        
        # Also log to main system log
        logger.info(log_entry)
    
    def update_daily_stats(self, trade_info: dict = None):
        """
        Update daily trading statistics
        """
        today = datetime.now().strftime('%Y-%m-%d')
        
        # Reset daily stats if it's a new day
        if not self.daily_trades or self.daily_trades[0].get('date') != today:
            self.daily_trades = []
            self.daily_pnl = 0.0
            self.daily_trade_count = 0
        
        if trade_info:
            self.daily_trades.append(trade_info)
            self.daily_pnl += trade_info.get('pnl', 0)
            self.daily_trade_count += 1
    
    def select_live_trading_opportunities(self, data: pd.DataFrame) -> Tuple[List[str], List[str]]:
        """
        Select trading opportunities based on current active positions
        - If no active trades: select 2 BUY and 2 SELL opportunities
        - If some trades active: only select opportunities to fill remaining slots
        
        Args:
            data: Current live options data (already filtered for Call options)
            
        Returns:
            Tuple of (undervalued_option_ids, overvalued_option_ids)
        """
        # Filter for valid data (non-null prices) - data is already Call options only
        valid_data = data.dropna(subset=['PX_LAST', 'Heston_Price'])
        
        if len(valid_data) < 4:
            logger.warning(f"Insufficient live data for {self.company}: only {len(valid_data)} valid Call options")
            return [], []
        
        # Find undervalued options (Market < Heston)
        undervalued = valid_data[valid_data['Market_vs_Heston'] < 0].copy()
        undervalued = undervalued.sort_values('Heston_vs_Market', ascending=False)
        
        # Find overvalued options (Market > Heston)
        overvalued = valid_data[valid_data['Market_vs_Heston'] > 0].copy()
        overvalued = overvalued.sort_values('Market_vs_Heston', ascending=False)
        
        # Count current active trades by type
        active_buy_trades = sum(1 for trade in self.active_trades.values() if trade['trade_type'] == 'BUY')
        active_sell_trades = sum(1 for trade in self.active_trades.values() if trade['trade_type'] == 'SELL')
        
        # Calculate how many more trades we need of each type
        needed_buy_trades = max(0, 2 - active_buy_trades)
        needed_sell_trades = max(0, 2 - active_sell_trades)
        
        # Select opportunities based on what we need
        undervalued_ids = undervalued['Option_ID'].head(needed_buy_trades).tolist()
        overvalued_ids = overvalued['Option_ID'].head(needed_sell_trades).tolist()
        
        logger.info(f"Live selection for {self.company}: {len(undervalued_ids)} BUY opportunities, {len(overvalued_ids)} SELL opportunities")
        logger.info(f"Current active trades: {active_buy_trades} BUY, {active_sell_trades} SELL")
        
        return undervalued_ids, overvalued_ids
    
    def enter_live_trade(self, option_id: str, trade_type: str, data: pd.DataFrame) -> bool:
        """
        Enter a new live trade if not already in position and max 4 trades not reached
        Position size is dynamic based on available capital (up to $25)
        
        Args:
            option_id: Unique option identifier
            trade_type: 'BUY' for undervalued, 'SELL' for overvalued
            data: Current live data
            
        Returns:
            True if trade entered, False otherwise
        """
        if option_id in self.active_trades:
            logger.debug(f"Already in live trade for {option_id}")
            return False
        
        # Check if we already have 4 active trades (max limit)
        if len(self.active_trades) >= 4:
            logger.debug(f"Maximum 4 trades already active for {self.company}, cannot enter new trade")
            return False
        
        option_data = data[data['Option_ID'] == option_id].iloc[0]
        
        # Calculate available capital for this trade
        # If we have losses, use maximum available amount up to $25
        # If we have profits, use standard $25
        # For the 4th trade, use all remaining capital (up to $25)
        
        # Check if this is the 4th trade (last trade)
        is_last_trade = (len(self.active_trades) == 3)  # 0-based, so 3 means 4th trade
        
        if self.current_capital >= self.initial_capital:
            # We have profits or break-even, use standard $25
            position_size = 25.0
        else:
            # We have losses
            if is_last_trade:
                # 4th trade: use all remaining capital (up to $25)
                # Calculate how much we've already allocated
                allocated_capital = sum(trade['position_size'] for trade in self.active_trades.values())
                remaining_capital = self.current_capital - allocated_capital
                position_size = min(25.0, remaining_capital)
                
                logger.info(f"4th trade detected for {self.company}: allocated=${allocated_capital:.2f}, remaining=${remaining_capital:.2f}, position_size=${position_size:.2f}")
            else:
                # First 3 trades: use maximum of $25 or capital/4, but never exceed available capital
                if self.current_capital >= 100.0:
                    # We have enough capital for 4 trades of $25 each
                    position_size = 25.0
                else:
                    # Use maximum of $25 or capital/4, but never exceed available capital
                    safe_per_trade = self.current_capital / 4
                    position_size = min(25.0, safe_per_trade)
        
        # Ensure we have minimum capital to trade
        if position_size < 1.0:
            logger.warning(f"Insufficient capital for {self.company}: ${self.current_capital:.2f}, minimum required: $1.00")
            return False
        
        # Calculate number of contracts (assuming $100 per contract)
        contracts = int(position_size / 100)
        if contracts == 0:
            contracts = 1  # Minimum 1 contract
        
        trade_info = {
            'option_id': option_id,
            'trade_type': trade_type,
            'entry_time': option_data['Timestamp'],
            'entry_hour': option_data['Hour'],
            'entry_market_price': option_data['PX_LAST'],
            'entry_heston_price': option_data['Heston_Price'],
            'strike_price': option_data['Strike'],
            'stock_price': option_data['Current Price'],
            'contracts': contracts,
            'position_size': position_size,
            'entry_mispricing': option_data['Market_vs_Heston'],
            'capital_at_entry': self.current_capital
        }
        
        self.active_trades[option_id] = trade_info
        self.total_trades += 1
        
        # Log trade to text file
        self.log_trade_to_text(trade_info, 'ENTER')
        
        # Enhanced logging for trade entry
        logger.info("=" * 80)
        logger.info(f"🎯 TRADE ENTERED: {trade_type} {option_id}")
        logger.info(f"   Company: {self.company}")
        logger.info(f"   Entry Time: {option_data['Timestamp'].strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"   Entry Price: ${option_data['PX_LAST']:.2f}")
        logger.info(f"   Heston Price: ${option_data['Heston_Price']:.2f}")
        logger.info(f"   Strike Price: ${option_data['Strike']:.2f}")
        logger.info(f"   Stock Price: ${option_data['Current Price']:.2f}")
        logger.info(f"   Contracts: {contracts}")
        logger.info(f"   Position Size: ${position_size:.2f}")
        logger.info(f"   Mispricing: ${option_data['Market_vs_Heston']:.2f}")
        logger.info(f"   Active Trades: {len(self.active_trades)}/4")
        logger.info(f"   Current Capital: ${self.current_capital:.2f}")
        logger.info(f"   Capital Status: {'PROFIT' if self.current_capital >= self.initial_capital else 'LOSS'}")
        if is_last_trade and self.current_capital < self.initial_capital:
            logger.info(f"   🎯 4TH TRADE: Using remaining capital allocation")
        logger.info("=" * 80)
        
        # Log trade entry to terminal activities
        self.save_terminal_activities(
            "TRADE_ENTRY",
            f"Trade entered: {trade_type} {option_id}",
            {
                "entry_price": f"${option_data['PX_LAST']:.2f}",
                "heston_price": f"${option_data['Heston_Price']:.2f}",
                "strike": f"${option_data['Strike']:.2f}",
                "contracts": contracts,
                "position_size": f"${position_size:.2f}",
                "mispricing": f"${option_data['Market_vs_Heston']:.2f}",
                "active_trades": f"{len(self.active_trades)}/4",
                "capital": f"${self.current_capital:.2f}"
            }
        )
        
        return True
    
    def check_live_exit_conditions(self, option_id: str, data: pd.DataFrame) -> bool:
        """
        Check if exit conditions are met for a live trade
        
        Args:
            option_id: Option identifier
            data: Current live data
            
        Returns:
            True if should exit, False otherwise
        """
        if option_id not in self.active_trades:
            return False
        
        trade = self.active_trades[option_id]
        option_data = data[data['Option_ID'] == option_id]
        
        if len(option_data) == 0:
            return False
        
        current_data = option_data.iloc[0]
        current_market_price = current_data['PX_LAST']
        current_heston_price = current_data['Heston_Price']
        
        # Exit conditions
        if trade['trade_type'] == 'BUY':
            # Exit when Market >= Heston (no longer undervalued)
            should_exit = current_market_price >= current_heston_price
        else:  # SELL
            # Exit when Market <= Heston (no longer overvalued)
            should_exit = current_market_price <= current_heston_price
        
        return should_exit
    
    def exit_live_trade(self, option_id: str, data: pd.DataFrame) -> bool:
        """
        Exit a live trade and calculate PnL
        
        Args:
            option_id: Option identifier
            data: Current live data
            
        Returns:
            True if trade exited, False otherwise
        """
        if option_id not in self.active_trades:
            return False
        
        trade = self.active_trades[option_id]
        option_data = data[data['Option_ID'] == option_id]
        
        if len(option_data) == 0:
            return False
        
        current_data = option_data.iloc[0]
        current_market_price = current_data['PX_LAST']
        
        # Calculate PnL
        if trade['trade_type'] == 'BUY':
            # Long position: profit = (exit_price - entry_price) * contracts
            pnl = (current_market_price - trade['entry_market_price']) * trade['contracts']
        else:  # SELL
            # Short position: profit = (entry_price - exit_price) * contracts
            pnl = (trade['entry_market_price'] - current_market_price) * trade['contracts']
        
        # Update trade info
        trade.update({
            'exit_time': current_data['Timestamp'],
            'exit_hour': current_data['Hour'],
            'exit_market_price': current_market_price,
            'exit_heston_price': current_data['Heston_Price'],
            'pnl': pnl,
            'return_pct': (pnl / trade['position_size']) * 100,
            'duration_hours': current_data['Hour'] - trade['entry_hour']
        })
        
        # Update portfolio statistics
        self.total_pnl += pnl
        if pnl > 0:
            self.profitable_trades += 1
        
        # Update capital
        self.current_capital += pnl
        
        # Move to trade history
        self.trade_history.append(trade)
        del self.active_trades[option_id]
        
        # Update daily stats
        self.update_daily_stats(trade)
        
        # Log trade to text file
        self.log_trade_to_text(trade, 'EXIT')
        
        # Enhanced logging for trade exit
        logger.info("=" * 80)
        logger.info(f"💰 TRADE EXITED: {trade['trade_type']} {option_id}")
        logger.info(f"   Company: {self.company}")
        logger.info(f"   Exit Time: {current_data['Timestamp'].strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"   Entry Time: {trade['entry_time'].strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"   Duration: {trade['duration_hours']} hours")
        logger.info(f"   Entry Price: ${trade['entry_market_price']:.2f}")
        logger.info(f"   Exit Price: ${current_market_price:.2f}")
        logger.info(f"   Entry Heston: ${trade['entry_heston_price']:.2f}")
        logger.info(f"   Exit Heston: ${current_data['Heston_Price']:.2f}")
        logger.info(f"   PnL: ${pnl:.2f}")
        logger.info(f"   Return: {trade['return_pct']:.2f}%")
        logger.info(f"   Contracts: {trade['contracts']}")
        logger.info(f"   Position Size: ${trade['position_size']:.2f}")
        logger.info(f"   Capital at Entry: ${trade['capital_at_entry']:.2f}")
        logger.info(f"   Capital at Exit: ${self.current_capital:.2f}")
        active_buy_trades = sum(1 for trade in self.active_trades.values() if trade['trade_type'] == 'BUY')
        active_sell_trades = sum(1 for trade in self.active_trades.values() if trade['trade_type'] == 'SELL')
        logger.info(f"   Active Trades: {len(self.active_trades)}/4 ({active_buy_trades} BUY, {active_sell_trades} SELL)")
        logger.info(f"   Total PnL: ${self.total_pnl:.2f}")
        logger.info(f"   Capital Status: {'PROFIT' if self.current_capital >= self.initial_capital else 'LOSS'}")
        logger.info("=" * 80)
        
        # Log trade exit to terminal activities
        self.save_terminal_activities(
            "TRADE_EXIT",
            f"Trade exited: {trade['trade_type']} {option_id}",
            {
                "entry_price": f"${trade['entry_market_price']:.2f}",
                "exit_price": f"${current_market_price:.2f}",
                "pnl": f"${pnl:.2f}",
                "return_pct": f"{trade['return_pct']:.2f}%",
                "duration_hours": trade['duration_hours'],
                "contracts": trade['contracts'],
                "active_trades": f"{len(self.active_trades)}/4",
                "capital": f"${self.current_capital:.2f}",
                "total_pnl": f"${self.total_pnl:.2f}"
            }
        )
        
        return True
    
    def process_live_hour(self, session):
        """
        Process one live trading hour for this company
        """
        current_time = datetime.now()
        current_hour = current_time.hour
        
        # Check if market is open
        if not self.is_market_open():
            logger.info(f"Market is closed for {self.company} at {current_time.strftime('%Y-%m-%d %H:%M:%S')} PST")
            return
        
        logger.info(f"Processing live hour for {self.company} at {current_time.strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Fetch fresh live data
        live_data = self.fetch_live_data(session)
        if live_data.empty:
            logger.error(f"No live data available for {self.company}, skipping this hour")
            return
        
        # Save hourly data with color coding
        self.save_hourly_data(live_data, current_hour)
        
        # Check exit conditions for existing trades
        options_to_exit = []
        for option_id in list(self.active_trades.keys()):
            if self.check_live_exit_conditions(option_id, live_data):
                options_to_exit.append(option_id)
        
        # Exit trades
        for option_id in options_to_exit:
            self.exit_live_trade(option_id, live_data)
        
        # Select new trading opportunities
        undervalued_ids, overvalued_ids = self.select_live_trading_opportunities(live_data)
        
        # Enter new trades
        for option_id in undervalued_ids:
            self.enter_live_trade(option_id, 'BUY', live_data)
        
        for option_id in overvalued_ids:
            self.enter_live_trade(option_id, 'SELL', live_data)
        
        # Log current portfolio status with detailed breakdown
        active_trades_count = len(self.active_trades)
        active_buy_trades = sum(1 for trade in self.active_trades.values() if trade['trade_type'] == 'BUY')
        active_sell_trades = sum(1 for trade in self.active_trades.values() if trade['trade_type'] == 'SELL')
        
        logger.info(f"Live hour complete for {self.company}: {active_trades_count}/4 active trades ({active_buy_trades} BUY, {active_sell_trades} SELL), Capital: ${self.current_capital:.2f}")
        
        # Force regenerate Excel to ensure all trades are properly recorded
        self.regenerate_hourly_excel_with_active_trades()
        
        # Log hour completion to terminal activities
        self.save_terminal_activities(
            "HOUR_COMPLETE",
            f"Live hour processing complete",
            {
                "active_trades": f"{active_trades_count}/4",
                "buy_trades": active_buy_trades,
                "sell_trades": active_sell_trades,
                "capital": f"${self.current_capital:.2f}",
                "hour": current_hour
            }
        )
        
        # Update last processed time
        self.last_processed_time = current_time
    
    def regenerate_hourly_excel_with_active_trades(self):
        """
        Regenerate hourly Excel file to show all active trades
        This is useful to fix the issue where trades don't show up in Excel
        """
        if not self.hourly_data_history:
            logger.warning(f"No hourly data history for {self.company}")
            return
        
        # Get the most recent hourly data
        latest_data = self.hourly_data_history[-1].copy()
        
        # Mark active trades based on hour logic
        active_trades_marked = 0
        hour = latest_data['Hour'].iloc[0] if 'Hour' in latest_data.columns else datetime.now().hour
        
        for option_id, trade in self.active_trades.items():
            # Check if this trade should be marked in this hour
            entry_time = trade.get('entry_time')
            if entry_time:
                try:
                    # Parse entry time to get hour
                    if isinstance(entry_time, str):
                        entry_hour = int(entry_time.split(' ')[1].split(':')[0])
                    else:
                        # entry_time is a datetime object
                        entry_hour = entry_time.hour
                    
                    # Only mark trades that were entered in this hour or earlier
                    if entry_hour <= hour:
                        # Try exact match first
                        mask = latest_data['Option_ID'] == option_id
                        if mask.any():
                            latest_data.loc[mask, 'Trade_Status'] = 'Entered'
                            latest_data.loc[mask, 'Trade_Type'] = trade['trade_type']
                            latest_data.loc[mask, 'Position_Size'] = trade['position_size']
                            active_trades_marked += 1
                            logger.info(f"Regenerated: Marked active trade in Excel (hour {hour}): {option_id} - {trade['trade_type']} - ${trade['position_size']:.2f} (entered hour {entry_hour})")
                        else:
                            # Try to find by parsing the option_id components
                            try:
                                # Parse option_id format: "COMPANY_STRIKE_OPTION_TYPE"
                                parts = option_id.split('_')
                                if len(parts) >= 3:
                                    company = parts[0]
                                    strike = float(parts[1])
                                    option_type = parts[2]
                                    
                                    # Find matching option in current data
                                    option_mask = (
                                        (latest_data['Strike'] == strike) & 
                                        (latest_data['Option Type'] == option_type)
                                    )
                                    
                                    if option_mask.any():
                                        latest_data.loc[option_mask, 'Trade_Status'] = 'Entered'
                                        latest_data.loc[option_mask, 'Trade_Type'] = trade['trade_type']
                                        latest_data.loc[option_mask, 'Position_Size'] = trade['position_size']
                                        active_trades_marked += 1
                                        logger.info(f"Regenerated: Marked active trade by components (hour {hour}): {option_id} - {trade['trade_type']} - ${trade['position_size']:.2f} (entered hour {entry_hour})")
                                    else:
                                        logger.warning(f"Regenerated: Active trade {option_id} not found in data for {self.company} (no matching strike/type)")
                                else:
                                    logger.warning(f"Regenerated: Active trade {option_id} has invalid format for {self.company}")
                            except Exception as e:
                                logger.warning(f"Regenerated: Error parsing active trade {option_id} for {self.company}: {e}")
                    else:
                        logger.info(f"Regenerated: Skipping trade {option_id} for hour {hour} (entered at hour {entry_hour})")
                except Exception as e:
                    logger.warning(f"Regenerated: Error parsing entry time for trade {option_id}: {e}")
                    # Fallback: mark the trade if we can't parse the time
                    mask = latest_data['Option_ID'] == option_id
                    if mask.any():
                        latest_data.loc[mask, 'Trade_Status'] = 'Entered'
                        latest_data.loc[mask, 'Trade_Type'] = trade['trade_type']
                        latest_data.loc[mask, 'Position_Size'] = trade['position_size']
                        active_trades_marked += 1
                        logger.info(f"Regenerated: Marked active trade (fallback): {option_id} - {trade['trade_type']} - ${trade['position_size']:.2f}")
        
        # Update the Excel file
        self.update_hourly_excel(latest_data, hour)
        
        logger.info(f"Regenerated hourly Excel for {self.company} with {active_trades_marked}/{len(self.active_trades)} active trades marked")
        
        # If we still couldn't mark all trades, log a warning
        if active_trades_marked < len(self.active_trades):
            logger.warning(f"Regenerated: Could only mark {active_trades_marked}/{len(self.active_trades)} active trades for {self.company}")
            logger.warning(f"Active trades: {list(self.active_trades.keys())}")
            logger.warning(f"Available options in data: {list(latest_data['Option_ID'].unique())}")
    
    def save_system_summary_to_terminal(self):
        """
        Save current system summary to terminal activities log
        """
        summary = {
            "total_trades": self.total_trades,
            "profitable_trades": self.profitable_trades,
            "total_pnl": f"${self.total_pnl:.2f}",
            "current_capital": f"${self.current_capital:.2f}",
            "active_trades": len(self.active_trades),
            "daily_pnl": f"${self.daily_pnl:.2f}",
            "daily_trades": self.daily_trade_count,
            "win_rate": f"{(self.profitable_trades / self.total_trades * 100) if self.total_trades > 0 else 0:.2f}%"
        }
        
        self.save_terminal_activities(
            "SYSTEM_SUMMARY",
            "Current system status summary",
            summary
        )


class MultiCompanyRealTimeTrading:
    """Multi-company real-time trading system"""
    
    def __init__(self, companies: List[str], capital_per_company: float = 100.0):
        """
        Initialize multi-company real-time trading system
        
        Args:
            companies: List of company ticker symbols
            capital_per_company: Capital allocation per company
        """
        self.companies = companies
        self.capital_per_company = capital_per_company
        self.session = None
        self.is_running = False
        
        # Initialize individual company trading systems
        self.company_systems = {}
        for company in companies:
            self.company_systems[company] = CompanyRealTimeTrading(company, capital_per_company)
        
        # Create output directories
        self.output_dir = "scripts/realtime_output/multi_company_sep19"
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Portfolio tracking
        self.total_portfolio_capital = capital_per_company * len(companies)
        self.current_portfolio_capital = self.total_portfolio_capital
        self.portfolio_trade_history = []
        
        logger.info(f"Multi-company real-time trading system initialized for {len(companies)} companies")
        logger.info(f"Companies: {', '.join(companies)}")
        logger.info(f"Capital per company: ${capital_per_company}")
        logger.info(f"Total portfolio capital: ${self.total_portfolio_capital}")
        logger.info(f"Portfolio allocation: {len(companies)} companies × ${capital_per_company} = ${self.total_portfolio_capital}")
    
    def is_market_open(self) -> bool:
        """
        Check if market is open (PST timezone)
        Market hours: 6:30 AM - 1:00 PM PST (9:30 AM - 4:00 PM EST)
        """
        now_pst = datetime.now(pytz.timezone('US/Pacific'))
        current_time = now_pst.time()
        
        # Market hours: 6:30 AM - 1:00 PM PST
        market_open = datetime.strptime("06:30:00", "%H:%M:%S").time()
        market_close = datetime.strptime("13:00:00", "%H:%M:%S").time()
        
        # Check if it's a weekday and within market hours
        is_weekday = now_pst.weekday() < 5  # Monday = 0, Friday = 4
        is_market_hours = market_open <= current_time <= market_close
        
        return is_weekday and is_market_hours
    
    def save_daily_summary(self):
        """
        Save daily trading summary for all companies
        """
        today = datetime.now().strftime('%Y-%m-%d')
        
        daily_summary = {
            'Date': [],
            'Company': [],
            'Number_of_Trades': [],
            'Daily_PnL': [],
            'Daily_Return_%': [],
            'Total_Trades': [],
            'Total_PnL': [],
            'Current_Capital': [],
            'Win_Rate_%': []
        }
        
        for company, system in self.company_systems.items():
            # Get daily stats
            daily_pnl = system.daily_pnl
            daily_trades = system.daily_trade_count
            
            # Calculate daily return
            daily_return = (daily_pnl / system.initial_capital * 100) if system.initial_capital > 0 else 0
            
            # Calculate win rate
            win_rate = (system.profitable_trades / system.total_trades * 100) if system.total_trades > 0 else 0
            
            daily_summary['Date'].append(today)
            daily_summary['Company'].append(company)
            daily_summary['Number_of_Trades'].append(daily_trades)
            daily_summary['Daily_PnL'].append(daily_pnl)
            daily_summary['Daily_Return_%'].append(daily_return)
            daily_summary['Total_Trades'].append(system.total_trades)
            daily_summary['Total_PnL'].append(system.total_pnl)
            daily_summary['Current_Capital'].append(system.current_capital)
            daily_summary['Win_Rate_%'].append(win_rate)
        
        # Create DataFrame
        df_daily = pd.DataFrame(daily_summary)
        
        # Save to Excel
        daily_filename = f"{self.output_dir}/daily_summary_{today}.xlsx"
        df_daily.to_excel(daily_filename, index=False)
        
        logger.info(f"Daily summary saved: {daily_filename}")
        
        # Also append to master daily log
        master_daily_file = f"{self.output_dir}/master_daily_log.xlsx"
        
        try:
            # Try to read existing file
            existing_df = pd.read_excel(master_daily_file)
            combined_df = pd.concat([existing_df, df_daily], ignore_index=True)
        except FileNotFoundError:
            # Create new file if it doesn't exist
            combined_df = df_daily
        
        combined_df.to_excel(master_daily_file, index=False)
        logger.info(f"Master daily log updated: {master_daily_file}")
    
    def connect_bloomberg(self):
        """Connect to Bloomberg Terminal"""
        try:
            self.session = connect_to_bloomberg()
            logger.info("Successfully connected to Bloomberg Terminal")
            return True
        except Exception as e:
            logger.error(f"Failed to connect to Bloomberg: {e}")
            return False
    
    def process_live_hour_all_companies(self):
        """Process one live trading hour for all companies"""
        current_time = datetime.now()
        logger.info("=" * 80)
        logger.info(f"🔄 PROCESSING LIVE HOUR FOR ALL COMPANIES at {current_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info("=" * 80)
        
        # Process each company
        for company in self.companies:
            try:
                self.company_systems[company].process_live_hour(self.session)
            except Exception as e:
                logger.error(f"Error processing {company}: {e}")
        
        # Update portfolio capital
        self.current_portfolio_capital = sum(
            system.current_capital for system in self.company_systems.values()
        )
        
        # Log portfolio status
        total_active_trades = sum(
            len(system.active_trades) for system in self.company_systems.values()
        )
        total_trades = sum(
            system.total_trades for system in self.company_systems.values()
        )
        total_pnl = sum(
            system.total_pnl for system in self.company_systems.values()
        )
        
        logger.info("=" * 80)
        logger.info(f"📊 PORTFOLIO STATUS:")
        logger.info(f"   Total Active Trades: {total_active_trades}")
        logger.info(f"   Total Trades Executed: {total_trades}")
        logger.info(f"   Total Portfolio PnL: ${total_pnl:.2f}")
        logger.info(f"   Current Portfolio Capital: ${self.current_portfolio_capital:.2f}")
        logger.info(f"   Portfolio Return: {((self.current_portfolio_capital - self.total_portfolio_capital) / self.total_portfolio_capital * 100):.2f}%")
        logger.info("=" * 80)
        
        # Save incremental results
        self.save_incremental_results()
        
        # Save daily summary (check if it's end of day)
        current_time = datetime.now()
        if current_time.hour >= 13:  # After 1 PM PST (market close)
            self.save_daily_summary()
    
    def save_incremental_results(self):
        """Save incremental results after each hour"""
        
        # Save individual company results
        for company, system in self.company_systems.items():
            if system.trade_history:
                df_trades = pd.DataFrame(system.trade_history)
                trade_file = f"{self.output_dir}/{company}_realtime_trades.xlsx"
                df_trades.to_excel(trade_file, index=False)
        
        # Save portfolio summary
        portfolio_summary = {
            'Company': [],
            'Initial Capital': [],
            'Current Capital': [],
            'Total PnL': [],
            'Total Return %': [],
            'Total Trades': [],
            'Profitable Trades': [],
            'Win Rate %': [],
            'Active Trades': [],
            'Last Updated': []
        }
        
        for company, system in self.company_systems.items():
            win_rate = (system.profitable_trades / system.total_trades * 100) if system.total_trades > 0 else 0
            total_return = ((system.current_capital - system.initial_capital) / system.initial_capital * 100)
            
            portfolio_summary['Company'].append(company)
            portfolio_summary['Initial Capital'].append(f"${system.initial_capital:.2f}")
            portfolio_summary['Current Capital'].append(f"${system.current_capital:.2f}")
            portfolio_summary['Total PnL'].append(f"${system.total_pnl:.2f}")
            portfolio_summary['Total Return %'].append(f"{total_return:.2f}%")
            portfolio_summary['Total Trades'].append(system.total_trades)
            portfolio_summary['Profitable Trades'].append(system.profitable_trades)
            portfolio_summary['Win Rate %'].append(f"{win_rate:.2f}%")
            portfolio_summary['Active Trades'].append(len(system.active_trades))
            portfolio_summary['Last Updated'].append(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        df_portfolio = pd.DataFrame(portfolio_summary)
        portfolio_file = f"{self.output_dir}/portfolio_summary.xlsx"
        df_portfolio.to_excel(portfolio_file, index=False)
        
        logger.info(f"Incremental portfolio results saved at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    def save_comprehensive_results(self):
        """Save comprehensive results with detailed analysis"""
        
        logger.info("=" * 80)
        logger.info("📊 SAVING COMPREHENSIVE PORTFOLIO RESULTS...")
        logger.info("=" * 80)
        
        # Save detailed trade logs for each company
        for company, system in self.company_systems.items():
            if system.trade_history:
                df_trades = pd.DataFrame(system.trade_history)
                
                # Add additional analysis columns
                df_trades['cumulative_pnl'] = df_trades['pnl'].cumsum()
                df_trades['cumulative_return'] = (df_trades['cumulative_pnl'] / system.initial_capital) * 100
                df_trades['trade_number'] = range(1, len(df_trades) + 1)
                
                # Save detailed trade log
                detailed_trade_file = f"{self.output_dir}/{company}_detailed_trades.xlsx"
                df_trades.to_excel(detailed_trade_file, index=False)
                logger.info(f"Detailed trade log saved for {company}: {detailed_trade_file}")
                
                # Save active trades summary
                if system.active_trades:
                    active_trades_data = []
                    for option_id, trade in system.active_trades.items():
                        active_trades_data.append({
                            'option_id': option_id,
                            'trade_type': trade['trade_type'],
                            'entry_time': trade['entry_time'],
                            'entry_price': trade['entry_market_price'],
                            'strike_price': trade['strike_price'],
                            'contracts': trade['contracts'],
                            'position_size': trade['position_size'],
                            'duration_hours': (datetime.now() - trade['entry_time']).total_seconds() / 3600
                        })
                    
                    df_active = pd.DataFrame(active_trades_data)
                    active_trades_file = f"{self.output_dir}/{company}_active_trades.xlsx"
                    df_active.to_excel(active_trades_file, index=False)
                    logger.info(f"Active trades saved for {company}: {active_trades_file}")
        
        # Save comprehensive portfolio summary
        portfolio_summary_data = {
            'Metric': [
                'Total Companies', 'Initial Portfolio Capital', 'Final Portfolio Capital', 'Total Portfolio PnL', 'Total Portfolio Return %',
                'Total Trades', 'Total Profitable Trades', 'Overall Win Rate %', 'Total Active Trades',
                'Best Performing Company', 'Worst Performing Company', 'Start Time', 'End Time', 'Total Runtime (hours)',
                'Last Updated'
            ],
            'Value': [
                len(self.companies),
                f"${self.total_portfolio_capital:.2f}",
                f"${self.current_portfolio_capital:.2f}",
                f"${self.current_portfolio_capital - self.total_portfolio_capital:.2f}",
                f"{((self.current_portfolio_capital - self.total_portfolio_capital) / self.total_portfolio_capital * 100):.2f}%",
                sum(system.total_trades for system in self.company_systems.values()),
                sum(system.profitable_trades for system in self.company_systems.values()),
                f"{(sum(system.profitable_trades for system in self.company_systems.values()) / sum(system.total_trades for system in self.company_systems.values()) * 100) if sum(system.total_trades for system in self.company_systems.values()) > 0 else 0:.2f}%",
                sum(len(system.active_trades) for system in self.company_systems.values()),
                max(self.company_systems.items(), key=lambda x: x[1].total_pnl)[0] if any(system.total_pnl > 0 for system in self.company_systems.values()) else "N/A",
                min(self.company_systems.items(), key=lambda x: x[1].total_pnl)[0] if any(system.total_pnl < 0 for system in self.company_systems.values()) else "N/A",
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "0.0",
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
        }
        
        df_portfolio_summary = pd.DataFrame(portfolio_summary_data)
        comprehensive_portfolio_file = f"{self.output_dir}/comprehensive_portfolio_summary.xlsx"
        df_portfolio_summary.to_excel(comprehensive_portfolio_file, index=False)
        logger.info(f"Comprehensive portfolio summary saved: {comprehensive_portfolio_file}")
        
        logger.info("=" * 80)
        logger.info("✅ COMPREHENSIVE PORTFOLIO RESULTS SAVED")
        logger.info("=" * 80)
    
    def generate_detailed_analysis(self):
        """Generate detailed analysis with plots and insights"""
        
        logger.info("=" * 80)
        logger.info("📈 GENERATING DETAILED PORTFOLIO ANALYSIS...")
        logger.info("=" * 80)
        
        # Create comprehensive analysis plots
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
        fig.suptitle('Multi-Company Real-Time Trading Analysis', fontsize=16)
        
        # Plot 1: Company Performance Comparison
        companies = list(self.company_systems.keys())
        returns = []
        pnls = []
        
        for company in companies:
            system = self.company_systems[company]
            total_return = ((system.current_capital - system.initial_capital) / system.initial_capital * 100)
            returns.append(total_return)
            pnls.append(system.total_pnl)
        
        colors = ['green' if ret > 0 else 'red' for ret in returns]
        bars1 = ax1.bar(companies, returns, color=colors, alpha=0.7)
        ax1.set_title('Company Performance Comparison')
        ax1.set_xlabel('Company')
        ax1.set_ylabel('Total Return (%)')
        ax1.grid(True, alpha=0.3)
        ax1.axhline(y=0, color='black', linestyle='-', alpha=0.5)
        
        # Plot 2: Total PnL by Company
        bars2 = ax2.bar(companies, pnls, color=colors, alpha=0.7)
        ax2.set_title('Total PnL by Company')
        ax2.set_xlabel('Company')
        ax2.set_ylabel('Total PnL ($)')
        ax2.grid(True, alpha=0.3)
        ax2.axhline(y=0, color='black', linestyle='-', alpha=0.5)
        
        # Plot 3: Trade Count by Company
        trade_counts = [system.total_trades for system in self.company_systems.values()]
        bars3 = ax3.bar(companies, trade_counts, alpha=0.7)
        ax3.set_title('Total Trades by Company')
        ax3.set_xlabel('Company')
        ax3.set_ylabel('Number of Trades')
        ax3.grid(True, alpha=0.3)
        
        # Plot 4: Win Rate by Company
        win_rates = []
        for system in self.company_systems.values():
            win_rate = (system.profitable_trades / system.total_trades * 100) if system.total_trades > 0 else 0
            win_rates.append(win_rate)
        
        bars4 = ax4.bar(companies, win_rates, alpha=0.7)
        ax4.set_title('Win Rate by Company')
        ax4.set_xlabel('Company')
        ax4.set_ylabel('Win Rate (%)')
        ax4.grid(True, alpha=0.3)
        
        plt.tight_layout()
        
        # Save analysis plot
        analysis_plot_file = f"{self.output_dir}/portfolio_detailed_analysis.png"
        plt.savefig(analysis_plot_file, dpi=300, bbox_inches='tight')
        plt.close()
        
        logger.info(f"Detailed portfolio analysis plot saved: {analysis_plot_file}")
        
        # Generate insights
        logger.info("=" * 80)
        logger.info("📊 PORTFOLIO TRADING INSIGHTS:")
        logger.info("=" * 80)
        
        # Calculate portfolio insights
        total_portfolio_return = ((self.current_portfolio_capital - self.total_portfolio_capital) / self.total_portfolio_capital) * 100
        total_trades = sum(system.total_trades for system in self.company_systems.values())
        total_profitable_trades = sum(system.profitable_trades for system in self.company_systems.values())
        overall_win_rate = (total_profitable_trades / total_trades * 100) if total_trades > 0 else 0
        
        best_company = max(self.company_systems.items(), key=lambda x: x[1].total_pnl)[0]
        worst_company = min(self.company_systems.items(), key=lambda x: x[1].total_pnl)[0]
        
        logger.info(f"💰 Total Portfolio Return: {total_portfolio_return:.2f}%")
        logger.info(f"🎯 Overall Win Rate: {overall_win_rate:.2f}%")
        logger.info(f"📊 Total Trades: {total_trades}")
        logger.info(f"✅ Total Profitable Trades: {total_profitable_trades}")
        logger.info(f"🚀 Best Performing Company: {best_company}")
        logger.info(f"📉 Worst Performing Company: {worst_company}")
        logger.info(f"💼 Total Active Positions: {sum(len(system.active_trades) for system in self.company_systems.values())}")
        
        logger.info("=" * 80)
        logger.info("✅ DETAILED PORTFOLIO ANALYSIS COMPLETE")
        logger.info("=" * 80)
    
    def run_realtime_trading(self):
        """
        Run the multi-company real-time trading system continuously
        """
        logger.info(f"Starting multi-company real-time trading system for {len(self.companies)} companies")
        
        # Connect to Bloomberg
        if not self.connect_bloomberg():
            logger.error("Failed to connect to Bloomberg, cannot start real-time trading")
            return
        
        self.is_running = True
        logger.info("Multi-company real-time trading system is now running. Press Ctrl+C to stop.")
        
        try:
            while self.is_running:
                current_time = datetime.now()
                
                # Check if market is open
                if self.is_market_open():
                    # Process the current hour for all companies
                    self.process_live_hour_all_companies()
                    
                    # Wait for the next hour (3600 seconds = 1 hour)
                    logger.info(f"Waiting for next hour... (will process at {(current_time + timedelta(hours=1)).strftime('%Y-%m-%d %H:%M:%S')} PST)")
                    time.sleep(3600)  # Wait 1 hour
                else:
                    # Market is closed, wait longer
                    logger.info(f"Market is closed. Waiting 30 minutes before checking again...")
                    time.sleep(1800)  # Wait 30 minutes
                
        except KeyboardInterrupt:
            logger.info("=" * 80)
            logger.info("🛑 MULTI-COMPANY REAL-TIME TRADING STOPPED BY USER (Ctrl+C)")
            logger.info("=" * 80)
        except Exception as e:
            logger.error(f"Error in multi-company real-time trading: {e}")
        finally:
            # Close any remaining positions for all companies
            logger.info("=" * 80)
            logger.info("🔚 CLOSING REMAINING POSITIONS FOR ALL COMPANIES...")
            logger.info("=" * 80)
            
            try:
                for company, system in self.company_systems.items():
                    logger.info(f"Closing positions for {company}...")
                    live_data = system.fetch_live_data(self.session)
                    if not live_data.empty:
                        for option_id in list(system.active_trades.keys()):
                            system.exit_live_trade(option_id, live_data)
                    else:
                        logger.warning(f"Could not fetch live data for {company} position closure")
            except Exception as e:
                logger.error(f"Error closing positions: {e}")
            
            # Close Bloomberg session
            if self.session:
                self.session.stop()
                logger.info("Bloomberg session closed")
            
            # Save comprehensive final results
            self.save_comprehensive_results()
            
            # Generate detailed final analysis
            self.generate_detailed_analysis()
            
            logger.info("=" * 80)
            logger.info("✅ MULTI-COMPANY REAL-TIME TRADING SYSTEM STOPPED")
            logger.info("=" * 80)
    
    def regenerate_all_excel_files(self):
        """
        Regenerate all Excel files to show current active trades
        This fixes the issue where trades don't show up in Excel files
        """
        logger.info("=" * 80)
        logger.info("🔄 REGENERATING ALL EXCEL FILES WITH ACTIVE TRADES...")
        logger.info("=" * 80)
        
        for company, system in self.company_systems.items():
            try:
                system.regenerate_hourly_excel_with_active_trades()
            except Exception as e:
                logger.error(f"Error regenerating Excel for {company}: {e}")
        
        logger.info("=" * 80)
        logger.info("✅ ALL EXCEL FILES REGENERATED")
        logger.info("=" * 80)


def main():
    """Main function to run the multi-company real-time trading system"""
    
    # Define the 9 companies
    companies = ['NVDA', 'AUR', 'TSLA', 'SOFI', 'SOUN', 'AMD', 'AVGO', 'CRCL', 'BBAI', 'SLDB']
    
    # Initialize multi-company real-time trading system
    trading_system = MultiCompanyRealTimeTrading(companies, capital_per_company=100.0)
    
    # Start multi-company real-time trading
    trading_system.run_realtime_trading()


if __name__ == "__main__":
    main() 