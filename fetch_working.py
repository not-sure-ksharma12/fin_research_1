import blpapi
import pandas as pd
from datetime import datetime
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# ---------- Bloomberg Setup ----------
def connect_to_bloomberg():
    from blpapi.sessionoptions import SessionOptions
    from blpapi.session import Session
    options = SessionOptions()
    options.setServerHost("localhost")
    options.setServerPort(8194)
    session = Session(options)
    if not session.start():
        raise RuntimeError("Failed to start session.")
    if not session.openService("//blp/refdata"):
        raise RuntimeError("Failed to open //blp/refdata")
    return session

# ---------- Get spot prices ----------
def get_current_price(session, tickers):
    prices = {}
    ref_data = session.getService("//blp/refdata")
    request = ref_data.createRequest("ReferenceDataRequest")
    for t in tickers:
        request.append("securities", f"{t} US Equity")
    request.append("fields", "PX_LAST")
    session.sendRequest(request)

    while True:
        ev = session.nextEvent(500)
        for msg in ev:
            if msg.messageType() == "ReferenceDataResponse":
                for sec in msg.getElement("securityData").values():
                    name = sec.getElementAsString("security").split()[0]
                    px = sec.getElement("fieldData").getElementAsFloat("PX_LAST")
                    prices[name] = px
        if ev.eventType() == blpapi.Event.EventType.RESPONSE:
            break
    return prices

# ---------- Get risk-free rate ----------
def get_risk_free_rate(session):
    ref_data = session.getService("//blp/refdata")
    request = ref_data.createRequest("ReferenceDataRequest")
    request.append("securities", "USGG3M Index")
    request.append("fields", "PX_LAST")
    session.sendRequest(request)

    while True:
        ev = session.nextEvent(500)
        for msg in ev:
            if msg.messageType() == "ReferenceDataResponse":
                sec_data = msg.getElement("securityData")
                field_data = sec_data.getValueAsElement(0).getElement("fieldData")
                return field_data.getElementAsFloat("PX_LAST") / 100
        if ev.eventType() == blpapi.Event.RESPONSE:
            break
    return 0.05

# ---------- Get valid Bloomberg option tickers from OPT_CHAIN ----------
def get_option_chain(session, ticker):
    ref_data = session.getService("//blp/refdata")
    request = ref_data.createRequest("ReferenceDataRequest")
    request.append("securities", f"{ticker} US Equity")
    request.append("fields", "OPT_CHAIN")
    session.sendRequest(request)
    chain_tickers = []

    while True:
        ev = session.nextEvent(500)
        for msg in ev:
            if msg.messageType() == "ReferenceDataResponse":
                sec_data = msg.getElement("securityData")
                for sec in sec_data.values():
                    field_data = sec.getElement("fieldData")
                    if field_data.hasElement("OPT_CHAIN"):
                        chain = field_data.getElement("OPT_CHAIN")
                        for i in range(chain.numValues()):
                            opt_elem = chain.getValue(i)
                            print(f"[DEBUG] OPT_CHAIN element {i} fields:")
                            for j in range(opt_elem.numElements()):
                                try:
                                    name = opt_elem.getElement(j).name()
                                    value = opt_elem.getElement(j).getValueAsString() if opt_elem.getElement(j).isNull() == False else str(opt_elem.getElement(j))
                                except Exception:
                                    name = str(j)
                                    value = str(opt_elem.getElement(j))
                                print(f"    {name}: {value}")
                            if opt_elem.hasElement("Security Description"):
                                chain_tickers.append(opt_elem.getElementAsString("Security Description"))
        if ev.eventType() == blpapi.Event.RESPONSE:
            break
    return chain_tickers

# ---------- Parse ticker structure ----------
def parse_option_ticker(ticker):
    parts = ticker.split()
    if len(parts) < 5:
        return None
    expiry = parts[2]
    cp_flag = parts[3][0]
    strike = float(parts[3][1:])
    exp_date = datetime.strptime(expiry, "%m/%d/%y").strftime("%Y-%m-%d")
    option_type = 'Call' if cp_flag == 'C' else 'Put'
    return {
        "Ticker": ticker,
        "Strike": strike,
        "Expiration": exp_date,
        "Option Type": option_type
    }

# ---------- Filter top 25 calls and puts expiring on target ----------
def select_top_25_each(option_list, target_exp="2025-12-19"):
    calls = []
    puts = []
    for o in option_list:
        parsed = parse_option_ticker(o)
        if parsed and parsed["Expiration"] == target_exp:
            if parsed["Option Type"] == "Call":
                calls.append((o, parsed["Strike"], "Call"))
            elif parsed["Option Type"] == "Put":
                puts.append((o, parsed["Strike"], "Put"))
    # Sort by strike and take top 25
    calls = sorted(calls, key=lambda x: x[1])[-25:]
    puts = sorted(puts, key=lambda x: x[1])[:25]
    return calls + puts

# ---------- Fetch option data ----------
def fetch_option_data(session, option_tickers):
    ref_data = session.getService("//blp/refdata")
    results = []

    for i in range(0, len(option_tickers), 50):
        batch = option_tickers[i:i+50]
        request = ref_data.createRequest("ReferenceDataRequest")
        for ticker, _, _ in batch:
            request.append("securities", ticker)
        fields = ["PX_LAST", "DELTA", "GAMMA", "VEGA", "THETA", "RHO", "IMPVOL_MID", "OPT_EXPIRE_DT", "OPT_CONTRACT_SIZE"]
        for f in fields:
            request.append("fields", f)
        session.sendRequest(request)

        while True:
            ev = session.nextEvent(500)
            for msg in ev:
                if msg.messageType() == "ReferenceDataResponse":
                    sec_data = msg.getElement("securityData")
                    for j in range(sec_data.numValues()):
                        sec = sec_data.getValueAsElement(j)
                        name = sec.getElementAsString("security")
                        field_data = sec.getElement("fieldData")
                        parts = name.split()
                        if len(parts) < 5:
                            continue
                        cp_flag = parts[3][0]
                        strike = float(parts[3][1:])
                        option_type = 'Call' if cp_flag == 'C' else 'Put'
                        # Fix: handle OPT_EXPIRE_DT as date or datetime
                        if field_data.hasElement("OPT_EXPIRE_DT"):
                            exp_val = field_data.getElementAsDatetime("OPT_EXPIRE_DT")
                            if hasattr(exp_val, 'isoformat'):
                                expiration = exp_val.isoformat()
                            else:
                                expiration = str(exp_val)
                        else:
                            expiration = None
                        results.append({
                            "Option Ticker": name,
                            "Strike": strike,
                            "Option Type": option_type,
                            "PX_LAST": field_data.getElementAsFloat("PX_LAST") if field_data.hasElement("PX_LAST") else None,
                            "DELTA": field_data.getElementAsFloat("DELTA") if field_data.hasElement("DELTA") else None,
                            "GAMMA": field_data.getElementAsFloat("GAMMA") if field_data.hasElement("GAMMA") else None,
                            "VEGA": field_data.getElementAsFloat("VEGA") if field_data.hasElement("VEGA") else None,
                            "THETA": field_data.getElementAsFloat("THETA") if field_data.hasElement("THETA") else None,
                            "RHO": field_data.getElementAsFloat("RHO") if field_data.hasElement("RHO") else None,
                            "IMPVOL_MID": field_data.getElementAsFloat("IMPVOL_MID") if field_data.hasElement("IMPVOL_MID") else None,
                            "Expiration": expiration,
                            "Contract Size": field_data.getElementAsFloat("OPT_CONTRACT_SIZE") if field_data.hasElement("OPT_CONTRACT_SIZE") else None,
                        })
            if ev.eventType() == blpapi.Event.RESPONSE:
                break
        time.sleep(0.1)
    return results

# ---------- Save with formatting ----------
def save_with_formatting(df, file_path):
    df["BS vs Market"] = None  # placeholder for future model comparison
    df["Heston vs Market"] = None
    df["BS vs Heston"] = None

    df.to_excel(file_path, index=False)
    wb = load_workbook(file_path)
    ws = wb.active
    if ws is None:
        raise ValueError("Active worksheet is None. Check the Excel file and openpyxl version.")

    header_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    # Use iter_rows to get the first row (header)
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        cell.fill = header_fill

    # Auto-size columns using iter_cols
    for idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), 1):
        max_length = max(len(str(cell.value) if cell.value else "") for cell in col_cells)
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_path)

# ========== MAIN ==========
if __name__ == "__main__":
    session = connect_to_bloomberg()
    tickers = ['NVDA', 'SOUN', 'BBAI', 'CRCL', 'AVGO', 'AMD']
    all_options = []

    print("Fetching current prices...")
    prices = get_current_price(session, tickers)
    print("Fetching risk-free rate...")
    rfr = get_risk_free_rate(session)

    print("Fetching real option tickers from OPT_CHAIN...")
    for ticker in tickers:
        chain = get_option_chain(session, ticker)
        selected = select_top_25_each(chain)
        for entry in selected:
            all_options.append(entry)

    print(f"Total options selected: {len(all_options)}")
    print("Fetching market data for selected options...")
    data = fetch_option_data(session, all_options)

    for row in data:
        row["Underlying"] = row["Option Ticker"].split()[0]
        row["Current Price"] = prices.get(row["Underlying"])
        row["Risk-Free Rate"] = rfr

    df = pd.DataFrame(data)
    save_with_formatting(df, "bloomberg_options_top50.xlsx")
    print("✅ Excel file saved: bloomberg_options_top50.xlsx")
