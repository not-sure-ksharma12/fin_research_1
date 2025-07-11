import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import time
import schedule
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import os
import QuantLib as ql
import math
import blpapi

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class BloombergOptionsDataFetcher:
    def __init__(self):
        self.session = None
        self.ref_data_service = None
        self.tickers = ['NVDA US Equity', 'SOUN US Equity', 'BBAI US Equity', 
                       'CRCL US Equity', 'AVGO US Equity', 'AMD US Equity']
        self.expiration_date = '2025-12-19'  # Options expiration date (updated)
        self.excel_file = 'options_data.xlsx'
        self.risk_free_rate = 0.05  # 5% default, will be fetched from Bloomberg
        
        # QuantLib setup
        try:
            self.calendar = ql.UnitedStates(ql.UnitedStates.NYSE)
            self.day_counter = ql.Actual365Fixed()
            self.calculation_date = ql.Date.todaysDate()
            ql.Settings.instance().evaluationDate = self.calculation_date
        except Exception as e:
            logger.error(f"Error initializing QuantLib: {e}")
            # Fallback to basic setup
            self.calendar = None
            self.day_counter = None
            self.calculation_date = None
        
    def connect_to_bloomberg(self):
        """Establish connection to Bloomberg Terminal"""
        try:
            session_options = blpapi.SessionOptions()
            session_options.setServerHost('localhost')
            session_options.setServerPort(8194)
            
            self.session = blpapi.Session(session_options)
            if not self.session.start():
                logger.error("Failed to start Bloomberg session")
                return False
                
            if not self.session.openService("//blp/refdata"):
                logger.error("Failed to open Bloomberg reference data service")
                return False
                
            self.ref_data_service = self.session.getService("//blp/refdata")
            logger.info("Successfully connected to Bloomberg Terminal")
            return True
            
        except Exception as e:
            logger.error(f"Error connecting to Bloomberg: {e}")
            return False
    
    def get_current_stock_price(self, ticker):
        """Get current stock price for a ticker"""
        try:
            request = self.ref_data_service.createRequest("ReferenceDataRequest")
            request.append("securities", ticker)
            request.append("fields", "PX_LAST")
            
            self.session.sendRequest(request)
            
            while True:
                event = self.session.nextEvent(500)
                if event.eventType() == blpapi.Event.RESPONSE:
                    for msg in event:
                        security_data = msg.getElement("securityData")
                        for i in range(security_data.numValues()):
                            security = security_data.getValueAsElement(i)
                            field_data = security.getElement("fieldData")
                            if field_data.hasElement("PX_LAST"):
                                return field_data.getElementAsFloat("PX_LAST")
                    break
                    
        except Exception as e:
            logger.error(f"Error getting stock price for {ticker}: {e}")
            return None
    
    def get_risk_free_rate(self, ticker=None):
        """Get risk-free rate from Bloomberg (per ticker, default to US 3-month Treasury for USD)."""
        try:
            # Default: use USGG3M Index for USD tickers
            # Extend here for other currencies if needed
            rate_ticker = "USGG3M Index"
            request = self.ref_data_service.createRequest("ReferenceDataRequest")
            request.append("securities", rate_ticker)
            request.append("fields", "PX_LAST")
            self.session.sendRequest(request)
            while True:
                event = self.session.nextEvent(500)
                if event.eventType() == blpapi.Event.RESPONSE:
                    for msg in event:
                        security_data = msg.getElement("securityData")
                        for i in range(security_data.numValues()):
                            security = security_data.getValueAsElement(i)
                            field_data = security.getElement("fieldData")
                            if field_data.hasElement("PX_LAST"):
                                return field_data.getElementAsFloat("PX_LAST") / 100
                    break
        except Exception as e:
            logger.error(f"Error getting risk-free rate for {ticker}: {e}")
            return self.risk_free_rate
    
    def get_volatility(self, ticker):
        """Get implied volatility for the ticker"""
        try:
            request = self.ref_data_service.createRequest("ReferenceDataRequest")
            request.append("securities", ticker)
            request.append("fields", "VOLATILITY_30D")
            
            self.session.sendRequest(request)
            
            while True:
                event = self.session.nextEvent(500)
                if event.eventType() == blpapi.Event.RESPONSE:
                    for msg in event:
                        security_data = msg.getElement("securityData")
                        for i in range(security_data.numValues()):
                            security = security_data.getValueAsElement(i)
                            field_data = security.getElement("fieldData")
                            if field_data.hasElement("VOLATILITY_30D"):
                                return field_data.getElementAsFloat("VOLATILITY_30D") / 100
                    break
                    
        except Exception as e:
            logger.error(f"Error getting volatility for {ticker}: {e}")
            return 0.3  # Default 30% volatility
    
    def get_option_data(self, ticker, strike_price, option_type, expiration_date):
        """Get option data including Greeks for specific strike and type"""
        try:
            # Create option ticker format for Bloomberg
            ticker_base = ticker.replace(' US Equity', '')
            exp_date = expiration_date.replace('-', '')
            option_ticker = f"{ticker_base} {exp_date}{'C' if option_type == 'Call' else 'P'}{strike_price:08.2f} Equity"
            
            request = self.ref_data_service.createRequest("ReferenceDataRequest")
            request.append("securities", option_ticker)
            
            # Request option price and Greeks
            fields = ["PX_LAST", "DELTA", "GAMMA", "THETA", "VEGA", "RHO", "IMPVOL_MID"]
            for field in fields:
                request.append("fields", field)
            
            self.session.sendRequest(request)
            
            option_data = {}
            
            while True:
                event = self.session.nextEvent(500)
                if event.eventType() == blpapi.Event.RESPONSE:
                    for msg in event:
                        security_data = msg.getElement("securityData")
                        for i in range(security_data.numValues()):
                            security = security_data.getValueAsElement(i)
                            field_data = security.getElement("fieldData")
                            
                            # Extract all available fields
                            for field in fields:
                                if field_data.hasElement(field):
                                    option_data[field] = field_data.getElementAsFloat(field)
                                else:
                                    option_data[field] = None
                    break
                    
            return option_data
                    
        except Exception as e:
            logger.error(f"Error getting option data for {ticker} strike {strike_price}: {e}")
            return None
    
    def quantlib_black_scholes_pricing(self, S, K, T, r, sigma, option_type, dividend_yield=0.0):
        """Calculate Black-Scholes option price and Greeks using QuantLib"""
        try:
            if T <= 0 or sigma <= 0:
                return {
                    'price': 0,
                    'delta': 0,
                    'gamma': 0,
                    'theta': 0,
                    'vega': 0,
                    'rho': 0
                }
            
            # Convert time to expiration to QuantLib date
            expiry_date = self.calculation_date + ql.Period(int(T * 365), ql.Days)
            
            # Create QuantLib objects
            underlying = ql.SimpleQuote(S)
            volatility = ql.BlackConstantVol(self.calculation_date, self.calendar, sigma, self.day_counter)
            risk_free_curve = ql.FlatForward(self.calculation_date, r, self.day_counter)
            dividend_curve = ql.FlatForward(self.calculation_date, dividend_yield, self.day_counter)
            
            # Create handles
            underlying_handle = ql.QuoteHandle(underlying)
            volatility_handle = ql.BlackVolTermStructureHandle(volatility)
            risk_free_handle = ql.YieldTermStructureHandle(risk_free_curve)
            dividend_handle = ql.YieldTermStructureHandle(dividend_curve)
            
            # Create the Black-Scholes process
            bs_process = ql.BlackScholesMertonProcess(underlying_handle, dividend_handle, 
                                                     risk_free_handle, volatility_handle)
            
            # Create the option
            exercise = ql.EuropeanExercise(expiry_date)
            if option_type == 'Call':
                payoff = ql.PlainVanillaPayoff(ql.Option.Call, K)
            else:
                payoff = ql.PlainVanillaPayoff(ql.Option.Put, K)
            
            option = ql.VanillaOption(payoff, exercise)
            
            # Create the Black-Scholes analytical pricing engine
            engine = ql.AnalyticEuropeanEngine(bs_process)
            option.setPricingEngine(engine)
            
            # Calculate price and Greeks
            price = option.NPV()
            delta = option.delta()
            gamma = option.gamma()
            theta = option.theta() / 365  # Convert to per-day
            vega = option.vega() / 100    # Convert to per 1% vol change
            rho = option.rho() / 100      # Convert to per 1% rate change
            
            return {
                'price': price,
                'delta': delta,
                'gamma': gamma,
                'theta': theta,
                'vega': vega,
                'rho': rho
            }
            
        except Exception as e:
            logger.error(f"Error in QuantLib Black-Scholes pricing: {e}")
            return {
                'price': 0,
                'delta': 0,
                'gamma': 0,
                'theta': 0,
                'vega': 0,
                'rho': 0
            }
    
    def quantlib_heston_pricing(self, S, K, T, r, sigma, option_type, dividend_yield=0.0):
        """Calculate Heston model option price and Greeks using QuantLib (finite difference engine for Greeks)"""
        try:
            if T <= 0 or sigma <= 0:
                return {
                    'price': 0,
                    'delta': 0,
                    'gamma': 0,
                    'theta': 0,
                    'vega': 0,
                    'rho': 0
                }
            
            expiry_date = self.calculation_date + ql.Period(int(T * 365), ql.Days)
            underlying = ql.SimpleQuote(S)
            risk_free_curve = ql.FlatForward(self.calculation_date, r, self.day_counter)
            dividend_curve = ql.FlatForward(self.calculation_date, dividend_yield, self.day_counter)
            underlying_handle = ql.QuoteHandle(underlying)
            risk_free_handle = ql.YieldTermStructureHandle(risk_free_curve)
            dividend_handle = ql.YieldTermStructureHandle(dividend_curve)
            v0 = sigma * sigma
            kappa = 2.0
            theta = v0
            sigma_v = 0.3
            rho_heston = -0.5
            heston_process = ql.HestonProcess(risk_free_handle, dividend_handle, underlying_handle, v0, kappa, theta, sigma_v, rho_heston)
            exercise = ql.EuropeanExercise(expiry_date)
            if option_type == 'Call':
                payoff = ql.PlainVanillaPayoff(ql.Option.Call, K)
            else:
                payoff = ql.PlainVanillaPayoff(ql.Option.Put, K)
            option = ql.VanillaOption(payoff, exercise)
            # Use finite difference engine for Greeks
            engine = ql.FdHestonVanillaEngine(ql.HestonModel(heston_process))
            option.setPricingEngine(engine)
            price = option.NPV()
            try:
                delta = option.delta()
            except:
                delta = 0
            try:
                gamma = option.gamma()
            except:
                gamma = 0
            try:
                theta = option.theta() / 365
            except:
                theta = 0
            try:
                vega = option.vega() / 100
            except:
                vega = 0
            try:
                rho = option.rho() / 100
            except:
                rho = 0
            return {
                'price': price,
                'delta': delta,
                'gamma': gamma,
                'theta': theta,
                'vega': vega,
                'rho': rho
            }
        except Exception as e:
            logger.error(f"Error in QuantLib Heston pricing: {e}")
            return {
                'price': 0,
                'delta': 0,
                'gamma': 0,
                'theta': 0,
                'vega': 0,
                'rho': 0
            }

    def quantlib_implied_volatility(self, S, K, T, r, market_price, option_type, dividend_yield=0.0):
        """Calculate implied volatility using QuantLib Black-Scholes"""
        try:
            if T <= 0 or market_price <= 0:
                logger.debug(f"Implied vol calc skipped: T={T}, market_price={market_price}")
                return None
            expiry_date = self.calculation_date + ql.Period(int(T * 365), ql.Days)
            underlying = ql.SimpleQuote(S)
            risk_free_curve = ql.FlatForward(self.calculation_date, r, self.day_counter)
            dividend_curve = ql.FlatForward(self.calculation_date, dividend_yield, self.day_counter)
            underlying_handle = ql.QuoteHandle(underlying)
            risk_free_handle = ql.YieldTermStructureHandle(risk_free_curve)
            dividend_handle = ql.YieldTermStructureHandle(dividend_curve)
            exercise = ql.EuropeanExercise(expiry_date)
            if option_type == 'Call':
                payoff = ql.PlainVanillaPayoff(ql.Option.Call, K)
            else:
                payoff = ql.PlainVanillaPayoff(ql.Option.Put, K)
            option = ql.VanillaOption(payoff, exercise)
            # Improved guess and bounds
            vol_guess = 0.25
            min_vol = 0.01
            max_vol = 1.0
            try:
                implied_vol = option.impliedVolatility(
                    market_price,
                    ql.BlackScholesMertonProcess(
                        underlying_handle, dividend_handle, risk_free_handle,
                        ql.BlackVolTermStructureHandle(
                            ql.BlackConstantVol(self.calculation_date, self.calendar, vol_guess, self.day_counter)
                        )
                    ),
                    1e-4,  # accuracy
                    100,   # max evaluations
                    min_vol,
                    max_vol
                )
                logger.debug(f"Implied vol success: S={S}, K={K}, T={T}, r={r}, market_price={market_price}, result={implied_vol}")
                return implied_vol
            except Exception as e:
                logger.debug(f"Implied vol failed: S={S}, K={K}, T={T}, r={r}, market_price={market_price}, error={e}")
                return None
        except Exception as e:
            logger.error(f"Error calculating implied volatility: {e}")
            return None
    
    def calculate_time_to_expiration(self, expiration_date):
        """Calculate time to expiration in years"""
        exp_date = datetime.strptime(expiration_date, '%Y-%m-%d')
        current_date = datetime.now()
        time_diff = exp_date - current_date
        return time_diff.days / 365.25
    
    def generate_strike_prices(self, current_price, num_strikes=25):
        """Generate top 25 strike prices around current price with $10 intervals"""
        # Start from a strike price below current price
        base_strike = (current_price // 10) * 10 - 120  # Start 12 strikes below
        strikes = []
        
        for i in range(num_strikes):
            strike = base_strike + (i * 10)
            if strike > 0:  # Only positive strikes
                strikes.append(strike)
                
        return strikes[:25]  # Ensure only top 25 strikes
    
    def get_listed_option_tickers(self, ticker, expiration_date):
        """Fetch listed option tickers for a given ticker and expiry from Bloomberg"""
        try:
            request = self.ref_data_service.createRequest("ReferenceDataRequest")
            request.append("securities", ticker)
            request.append("fields", "OPT_CHAIN")
            self.session.sendRequest(request)
            option_tickers = []
            
            # Convert expiration date to Bloomberg format (MM/DD/YY)
            exp_date = datetime.strptime(expiration_date, '%Y-%m-%d')
            bloomberg_expiry = exp_date.strftime('%m/%d/%y')
            
            while True:
                event = self.session.nextEvent(500)
                if event.eventType() == blpapi.Event.RESPONSE:
                    for msg in event:
                        logger.info(f"Raw Bloomberg OPT_CHAIN response: {msg}")  # Debug print
                        security_data = msg.getElement("securityData")
                        for i in range(security_data.numValues()):
                            sec = security_data.getValueAsElement(i)
                            field_data = sec.getElement("fieldData")
                            if field_data.hasElement("OPT_CHAIN"):
                                chain = field_data.getElement("OPT_CHAIN")
                                for j in range(chain.numValues()):
                                    opt = chain.getValueAsElement(j)
                                    opt_ticker = opt.getElementAsString("Security Description")
                                    # Only keep tickers for the desired expiry
                                    if bloomberg_expiry in opt_ticker:
                                        option_tickers.append(opt_ticker)
                    break
            return option_tickers
        except Exception as e:
            logger.error(f"Error fetching option chain for {ticker}: {e}")
            return []

    def parse_bloomberg_option_ticker(self, opt_ticker):
        """Parse Bloomberg option ticker format like 'AMD US 07/18/25 C45 Equity'"""
        try:
            # Example: "AMD US 07/18/25 C45 Equity"
            parts = opt_ticker.split()
            if len(parts) < 5:
                logger.warning(f"Malformed option ticker: {opt_ticker}")
                return None, None, None
            
            # Extract components
            underlying = parts[0]  # AMD
            country = parts[1]     # US
            expiry_date = parts[2] # 07/18/25
            cp_flag = parts[3][0]  # C or P
            strike_str = parts[3][1:]  # 45
            option_type = 'Call' if cp_flag == 'C' else 'Put'
            
            # Parse strike price
            strike = float(strike_str)
            
            # Convert expiry to standard format
            exp_date = datetime.strptime(expiry_date, '%m/%d/%y')
            standard_expiry = exp_date.strftime('%Y-%m-%d')
            
            return strike, option_type, standard_expiry
            
        except Exception as e:
            logger.warning(f"Failed to parse option ticker {opt_ticker}: {e}")
            return None, None, None

    def get_last_nonzero_historical_price(self, option_ticker, days=30):
        """Fetch the last nonzero historical PX_LAST for an option ticker from Bloomberg."""
        try:
            request = self.ref_data_service.createRequest("HistoricalDataRequest")
            request.append("securities", option_ticker)
            request.append("fields", "PX_LAST")
            today = datetime.now().strftime('%Y%m%d')
            start_date = (datetime.now() - timedelta(days=days)).strftime('%Y%m%d')
            request.set("startDate", start_date)
            request.set("endDate", today)
            request.set("periodicitySelection", "DAILY")
            self.session.sendRequest(request)
            last_nonzero = None
            while True:
                event = self.session.nextEvent(500)
                if event.eventType() == blpapi.Event.RESPONSE:
                    for msg in event:
                        security_data = msg.getElement("securityData")
                        if security_data.hasElement("fieldData"):
                            field_data_array = security_data.getElement("fieldData")
                            for j in range(field_data_array.numValues()):
                                field_data = field_data_array.getValueAsElement(j)
                                if field_data.hasElement("PX_LAST"):
                                    px = field_data.getElementAsFloat("PX_LAST")
                                    if px and px > 0:
                                        last_nonzero = px
                    break
            return last_nonzero
        except Exception as e:
            logger.error(f"Error fetching historical price for {option_ticker}: {e}")
            return None

    def get_last_nonzero_historical_ivm(self, option_ticker, days=30):
        """Fetch the last nonzero historical IMPVOL_MID for an option ticker from Bloomberg."""
        try:
            request = self.ref_data_service.createRequest("HistoricalDataRequest")
            request.append("securities", option_ticker)
            request.append("fields", "IMPVOL_MID")
            today = datetime.now().strftime('%Y%m%d')
            start_date = (datetime.now() - timedelta(days=days)).strftime('%Y%m%d')
            request.set("startDate", start_date)
            request.set("endDate", today)
            request.set("periodicitySelection", "DAILY")
            self.session.sendRequest(request)
            last_nonzero = None
            while True:
                event = self.session.nextEvent(500)
                if event.eventType() == blpapi.Event.RESPONSE:
                    for msg in event:
                        security_data = msg.getElement("securityData")
                        if security_data.hasElement("fieldData"):
                            field_data_array = security_data.getElement("fieldData")
                            for j in range(field_data_array.numValues()):
                                field_data = field_data_array.getValueAsElement(j)
                                if field_data.hasElement("IMPVOL_MID"):
                                    ivm = field_data.getElementAsFloat("IMPVOL_MID")
                                    if ivm and ivm > 0:
                                        last_nonzero = ivm
                    break
            return last_nonzero
        except Exception as e:
            logger.error(f"Error fetching historical IVM for {option_ticker}: {e}")
            return None

    def fetch_all_options_data(self):
        """Fetch all options data for all tickers using only listed strikes"""
        all_data = []
        try:
            for ticker in self.tickers:
                logger.info(f"Fetching listed options for {ticker}")
                risk_free_rate = self.get_risk_free_rate(ticker)
                stock_price = self.get_current_stock_price(ticker)
                logger.info(f"Stock price for {ticker}: {stock_price}")
                if stock_price is None:
                    logger.warning(f"Could not get stock price for {ticker}, skipping")
                    continue
                listed_option_tickers = self.get_listed_option_tickers(ticker, self.expiration_date)
                # Parse strikes from tickers
                parsed_options = []
                for opt_ticker in listed_option_tickers:
                    strike, option_type, expiry_date = self.parse_bloomberg_option_ticker(opt_ticker)
                    if strike is not None:
                        parsed_options.append((strike, option_type, expiry_date, opt_ticker))
                # Sort by strike
                parsed_options.sort(key=lambda x: x[0])
                # Find index of closest strike to stock price
                closest_idx = min(range(len(parsed_options)), key=lambda i: abs(parsed_options[i][0] - stock_price)) if parsed_options else 0
                # Get 25 below and 25 above (including closest)
                start = max(0, closest_idx - 25)
                end = min(len(parsed_options), closest_idx + 26)  # +1 for inclusive
                selected_options = parsed_options[start:end]
                logger.info(f"Selected {len(selected_options)} options for {ticker} (25 below and 25 above)")
                volatility = self.get_volatility(ticker)
                logger.info(f"Volatility for {ticker}: {volatility}")
                for strike, option_type, expiry_date, opt_ticker in selected_options:
                    try:
                        time_to_expiration = self.calculate_time_to_expiration(expiry_date)
                        if time_to_expiration <= 0:
                            logger.warning(f"Skipping expired option: {opt_ticker} (T={time_to_expiration})")
                            continue
                        logger.info(f"Inputs: Ticker={ticker}, OptionTicker={opt_ticker}, Strike={strike}, Type={option_type}, S={stock_price}, Vol={volatility}, T={time_to_expiration}, r={risk_free_rate}")
                        option_data = self.get_option_data(ticker, strike, option_type, expiry_date)
                        logger.info(f"Bloomberg option data for {opt_ticker}: {option_data}")
                        option_price = option_data.get('PX_LAST', 0)
                        if not option_price or option_price == 0:
                            # Try to get last nonzero historical price
                            hist_price = self.get_last_nonzero_historical_price(opt_ticker)
                            if hist_price:
                                option_price = hist_price
                            else:
                                option_price = strike
                        bbg_delta = option_data.get('DELTA')
                        bbg_gamma = option_data.get('GAMMA')
                        bbg_theta = option_data.get('THETA')
                        bbg_vega = option_data.get('VEGA')
                        bbg_rho = option_data.get('RHO')
                        implied_vol = option_data.get('IMPVOL_MID')
                        if not implied_vol or implied_vol == 0:
                            hist_ivm = self.get_last_nonzero_historical_ivm(opt_ticker)
                            if hist_ivm:
                                implied_vol = hist_ivm
                        vol_for_pricing = (implied_vol / 100) if implied_vol else volatility
                        logger.info(f"QuantLib Inputs: S={stock_price}, K={strike}, T={time_to_expiration}, r={risk_free_rate}, sigma={vol_for_pricing}, type={option_type}")
                        bs_results = self.quantlib_black_scholes_pricing(stock_price, strike, time_to_expiration, risk_free_rate, vol_for_pricing, option_type)
                        logger.info(f"BS Results: {bs_results}")
                        heston_results = self.quantlib_heston_pricing(stock_price, strike, time_to_expiration, risk_free_rate, vol_for_pricing, option_type)
                        logger.info(f"Heston Results: {heston_results}")
                        calculated_iv = None
                        if option_price > 0:
                            calculated_iv = self.quantlib_implied_volatility(stock_price, strike, time_to_expiration, risk_free_rate, option_price, option_type)
                            logger.info(f"Implied Vol (Calc) for {opt_ticker}: {calculated_iv}")
                        row_data = {
                            'Ticker': ticker.replace(' US Equity', ''),
                            'Stock Price': stock_price,
                            'Strike Price': strike,
                            'Option Price (BBG)': option_price,
                            'Black-Scholes Price': bs_results['price'],
                            'Heston Price': heston_results['price'],
                            'BS vs Market': option_price - bs_results['price'],
                            'Heston vs Market': option_price - heston_results['price'],
                            'BS vs Heston': bs_results['price'] - heston_results['price'],
                            'Time to Expiration': time_to_expiration,
                            'Risk Free Rate': risk_free_rate,
                            'Implied Vol (BBG)': implied_vol / 100 if implied_vol else None,
                            'Implied Vol (Calc)': calculated_iv,
                            'Historical Volatility': volatility,
                            'Option Type': option_type,
                            'Delta (BBG)': bbg_delta,
                            'Gamma (BBG)': bbg_gamma,
                            'Theta (BBG)': bbg_theta,
                            'Vega (BBG)': bbg_vega,
                            'Rho (BBG)': bbg_rho,
                            'Delta (BS)': bs_results['delta'],
                            'Gamma (BS)': bs_results['gamma'],
                            'Theta (BS)': bs_results['theta'],
                            'Vega (BS)': bs_results['vega'],
                            'Rho (BS)': bs_results['rho'],
                            'Delta (Heston)': heston_results['delta'],
                            'Gamma (Heston)': heston_results['gamma'],
                            'Theta (Heston)': heston_results['theta'],
                            'Vega (Heston)': heston_results['vega'],
                            'Rho (Heston)': heston_results['rho'],
                            'Expiration Date': expiry_date,
                            'Last Updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        }
                        all_data.append(row_data)
                        time.sleep(0.1)
                    except Exception as e:
                        logger.warning(f"Error processing option ticker {opt_ticker}: {e}")
                        continue
        except Exception as e:
            logger.error(f"Error in fetch_all_options_data: {e}")
        return all_data
    
    def save_to_excel(self, data):
        """Save data to Excel with formatting and conditional formatting for error columns"""
        try:
            df = pd.DataFrame(data)
            with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Options Data', index=False)
                workbook = writer.book
                worksheet = writer.sheets['Options Data']
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                for row in range(2, len(df) + 2):
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = border
                        if col in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30]:
                            if col in [11, 12, 13, 14]:
                                cell.number_format = '0.00%'
                            elif col in [16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30]:
                                cell.number_format = '0.0000'
                            else:
                                cell.number_format = '0.00'
                # Conditional formatting for error columns
                from openpyxl.formatting.rule import ColorScaleRule
                # BS vs Market (col 7), BS vs Heston (col 9)
                worksheet.conditional_formatting.add(f'G2:G{len(df)+1}', ColorScaleRule(start_type='min', start_color='FFFFFF', mid_type='percentile', mid_value=50, mid_color='FFFF00', end_type='max', end_color='FF0000'))
                worksheet.conditional_formatting.add(f'I2:I{len(df)+1}', ColorScaleRule(start_type='min', start_color='FFFFFF', mid_type='percentile', mid_value=50, mid_color='FFFF00', end_type='max', end_color='FF0000'))
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            logger.info(f"Data saved to {self.excel_file}")
        except Exception as e:
            logger.error(f"Error saving to Excel: {e}")
    
    def run_data_fetch(self):
        """Run the data fetching process"""
        try:
            if not self.connect_to_bloomberg():
                logger.error("Could not connect to Bloomberg Terminal")
                return
            
            logger.info("Starting data fetch...")
            data = self.fetch_all_options_data()
            
            if data:
                self.save_to_excel(data)
                logger.info(f"Successfully fetched data for {len(data)} options")
            else:
                logger.warning("No data was fetched")
                
        except Exception as e:
            logger.error(f"Error in data fetch: {e}")
        finally:
            if self.session:
                self.session.stop()
    
    def start_scheduler(self):
        """Start the hourly scheduler"""
        logger.info("Starting options data fetcher with 5-hourly updates")
        
        # Run immediately
        self.run_data_fetch()
        
        # Schedule to run every 5 hours
        schedule.every(5).hours.do(self.run_data_fetch)
        
        while True:
            schedule.run_pending()
            time.sleep(60)  # Check every minute

def main():
    """Main function to run the options data fetcher"""
    fetcher = BloombergOptionsDataFetcher()
    
    try:
        fetcher.start_scheduler()
    except KeyboardInterrupt:
        logger.info("Shutting down options data fetcher")
    except Exception as e:
        logger.error(f"Unexpected error: {e}")

if __name__ == "__main__":
    main() 